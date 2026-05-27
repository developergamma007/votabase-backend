import json
import traceback
import os
from io import BytesIO
import re
import tempfile
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import boto3
import jwt
from botocore.exceptions import NoCredentialsError
from fastapi import Depends, FastAPI, File, HTTPException, Query, Request, UploadFile, BackgroundTasks
from starlette.background import BackgroundTask
from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path
from fastapi.responses import JSONResponse, FileResponse
import base64
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import Boolean, DateTime, Double, ForeignKey, Integer, String, and_, asc, case, create_engine, desc, func, or_, text, not_
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker
import subprocess
import shutil

from app.extract import router as extract_router


# ---------------------------
# Config
# ---------------------------
def _load_dotenv_from_python_backend() -> None:
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            # Keep backend deterministic: values in votabase-backend/.env
            # should override inherited shell variables.
            os.environ[key.strip()] = value.strip().strip("\"'")


_load_dotenv_from_python_backend()

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL is required. Example: postgresql+psycopg2://user:pass@host:5432/dbname")

CONTEXT_PATH = os.getenv("CONTEXT_PATH", "/votebase/v1")
AWS_REGION = os.getenv("AWS_REGION", "ap-south-1")
AWS_BUCKET = os.getenv("AWS_S3_BUCKET", "")
AWS_ACCESS_KEY = os.getenv("AWS_ACCESS_KEY_ID", "")
AWS_SECRET_KEY = os.getenv("AWS_SECRET_ACCESS_KEY", "")
PROFILE_UPLOAD_DIR = os.getenv("AWS_S3_PROFILE_PICS", "profile_pics")
SUPERADMIN_USERNAME = os.getenv("SUPERADMIN_USERNAME", "admin@iswot.io")

JWT_SECRET = "supersecretkeysupersecretkeysupersecretkey"
JWT_EXPIRATION_SECONDS = 60 * 60 * 24 * 365
SNAPSHOT_CACHE_TTL_SECONDS = 60 * 15


# In-memory cache for large snapshot payloads served via short-lived links.
_snapshot_cache: Dict[str, Dict[str, Any]] = {}


def _safe_db_target(db_url: str) -> str:
    # Print DB target without exposing credentials.
    m = re.match(r"postgresql(?:\+psycopg2)?://[^@]+@([^:/]+)(?::(\d+))?/([^?]+)", db_url or "")
    if not m:
        return "unknown"
    host = m.group(1)
    port = m.group(2) or "5432"
    db = m.group(3)
    return f"{host}:{port}/{db}"


def _cleanup_snapshot_cache() -> None:
    now_ts = int(datetime.now(timezone.utc).timestamp())
    expired = [
        k
        for k, v in _snapshot_cache.items()
        if now_ts - int(v.get("created_ts", 0)) > SNAPSHOT_CACHE_TTL_SECONDS
    ]
    for k in expired:
        _snapshot_cache.pop(k, None)


def _cache_snapshot(snapshot: Dict[str, Any]) -> str:
    _cleanup_snapshot_cache()
    snapshot_id = uuid.uuid4().hex
    _snapshot_cache[snapshot_id] = {
        "created_ts": int(datetime.now(timezone.utc).timestamp()),
        "payload": snapshot,
    }
    return snapshot_id


# ---------------------------
# DB models
# ---------------------------
class Base(DeclarativeBase):
    pass


class Tenant(Base):
    __tablename__ = "tenant"
    __table_args__ = {"schema": "metastore"}

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_id: Mapped[str] = mapped_column(String(20), unique=True)
    name: Mapped[str] = mapped_column(String(255))
    description: Mapped[Optional[str]] = mapped_column(String)
    contact_email: Mapped[str] = mapped_column(String(255), unique=True)
    contact_phone: Mapped[Optional[str]] = mapped_column(String(50))
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[Optional[datetime]] = mapped_column(DateTime)


class User(Base):
    __tablename__ = "users"
    __table_args__ = {"schema": "metastore"}

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_ref: Mapped[Optional[int]] = mapped_column("tenant_id", ForeignKey("metastore.tenant.id"))
    role: Mapped[str] = mapped_column(String(30))
    assignment_type: Mapped[Optional[str]] = mapped_column(String(30))
    assignment_id: Mapped[Optional[int]] = mapped_column(Integer)
    first_name: Mapped[str] = mapped_column(String(100), unique=True)
    phone: Mapped[str] = mapped_column(String(10), unique=True)
    profile_pic_url: Mapped[Optional[str]] = mapped_column(String)
    blocked: Mapped[bool] = mapped_column(Boolean, default=False)
    deleted: Mapped[bool] = mapped_column(Boolean, default=False)

    tenant: Mapped[Optional[Tenant]] = relationship(Tenant)


class VolunteerUser(Base):
    __tablename__ = "volunteer_users"
    __table_args__ = {"schema": "metastore"}

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_id: Mapped[str] = mapped_column(String(20))
    role: Mapped[str] = mapped_column(String(30), default="USER")
    working_level: Mapped[Optional[str]] = mapped_column(String(30))
    assignment_type: Mapped[Optional[str]] = mapped_column(String(30))
    assignment_id: Mapped[Optional[str]] = mapped_column(String)
    assembly_ids: Mapped[Optional[str]] = mapped_column(String)
    ward_ids: Mapped[Optional[str]] = mapped_column(String)
    booth_ids: Mapped[Optional[str]] = mapped_column(String)
    first_name: Mapped[str] = mapped_column(String(100))
    phone: Mapped[str] = mapped_column(String(10), unique=True)
    profile_pic_url: Mapped[Optional[str]] = mapped_column(String)
    blocked: Mapped[bool] = mapped_column(Boolean, default=False)
    deleted: Mapped[bool] = mapped_column(Boolean, default=False)

class Assembly(Base):
    __tablename__ = "assembly"
    __table_args__ = {"schema": "data"}

    assembly_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_id: Mapped[str] = mapped_column(String(20))
    assembly_name_en: Mapped[Optional[str]] = mapped_column(String(255))
    assembly_name_local: Mapped[Optional[str]] = mapped_column(String(255))
    assembly_code: Mapped[str] = mapped_column(String(12), unique=True)


class Ward(Base):
    __tablename__ = "wards"
    __table_args__ = {"schema": "data"}

    ward_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    assembly_id: Mapped[int] = mapped_column(ForeignKey("data.assembly.assembly_id"))
    tenant_id: Mapped[str] = mapped_column(String(20))
    ward_name_en: Mapped[Optional[str]] = mapped_column(String(255))
    ward_name_local: Mapped[Optional[str]] = mapped_column(String(255))
    ward_code: Mapped[str] = mapped_column(String(20), unique=True)


class Booth(Base):
    __tablename__ = "booths"
    __table_args__ = {"schema": "data"}

    booth_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    ward_id: Mapped[int] = mapped_column(ForeignKey("data.wards.ward_id"))
    tenant_id: Mapped[str] = mapped_column(String(20))
    polling_station_adr_en: Mapped[Optional[str]] = mapped_column(String(255))
    polling_station_adr_local: Mapped[Optional[str]] = mapped_column(String(255))
    booth_no: Mapped[Optional[str]] = mapped_column(String(20))
    ward_code: Mapped[Optional[str]] = mapped_column(String(20))
    booth_add_en: Mapped[Optional[str]] = mapped_column(String)
    booth_add_local: Mapped[Optional[str]] = mapped_column(String)


class Association(Base):
    __tablename__ = "association"
    __table_args__ = {"schema": "data"}

    association_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    association_name: Mapped[str] = mapped_column(String(100))
    booth_id: Mapped[Optional[int]] = mapped_column(ForeignKey("data.booths.booth_id"))
    association_address: Mapped[Optional[str]] = mapped_column(String)
    association_head_name: Mapped[Optional[str]] = mapped_column(String)
    phone: Mapped[Optional[str]] = mapped_column(String)
    latitude: Mapped[Optional[float]] = mapped_column(Double)
    longitude: Mapped[Optional[float]] = mapped_column(Double)
    tenant_id: Mapped[str] = mapped_column(String(20))


class Voter(Base):
    __tablename__ = "voters"
    __table_args__ = {"schema": "data"}

    voter_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_id: Mapped[str] = mapped_column(String(20))
    booth_id: Mapped[int] = mapped_column("booth_id", ForeignKey("data.booths.booth_id"))
    sr_no: Mapped[Optional[int]] = mapped_column(Integer)
    epic_no: Mapped[Optional[str]] = mapped_column(String(20), unique=True)
    first_middle_name_en: Mapped[Optional[str]] = mapped_column(String(150))
    last_name_en: Mapped[Optional[str]] = mapped_column(String(100))
    first_middle_name_local: Mapped[Optional[str]] = mapped_column(String(150))
    last_name_local: Mapped[Optional[str]] = mapped_column(String(100))
    relation_type: Mapped[Optional[str]] = mapped_column(String(20))
    relation_first_middle_name_en: Mapped[Optional[str]] = mapped_column(String(150))
    relation_last_name_en: Mapped[Optional[str]] = mapped_column(String(100))
    relation_first_middle_name_local: Mapped[Optional[str]] = mapped_column(String(150))
    relation_last_name_local: Mapped[Optional[str]] = mapped_column(String(100))
    house_no_en: Mapped[Optional[str]] = mapped_column(String(50))
    house_no_local: Mapped[Optional[str]] = mapped_column(String(50))
    gender: Mapped[Optional[str]] = mapped_column(String(10))
    age: Mapped[Optional[int]] = mapped_column(Integer)
    dob: Mapped[Optional[datetime]] = mapped_column(DateTime)
    mobile: Mapped[Optional[str]] = mapped_column(String(15))
    address_en: Mapped[Optional[str]] = mapped_column(String(255))
    address_local: Mapped[Optional[str]] = mapped_column(String(255))
    status: Mapped[Optional[str]] = mapped_column(String(20))
    community: Mapped[Optional[str]] = mapped_column(String(100))
    caste: Mapped[Optional[str]] = mapped_column(String(100))
    residence_type: Mapped[Optional[str]] = mapped_column(String(100))
    civic_issue: Mapped[Optional[str]] = mapped_column(String(255))
    mother_tongue: Mapped[Optional[str]] = mapped_column(String(100))
    team: Mapped[Optional[str]] = mapped_column(String(100))
    ownership: Mapped[Optional[str]] = mapped_column(String(20))
    education: Mapped[Optional[str]] = mapped_column(String(20))
    nature_of_voter: Mapped[Optional[str]] = mapped_column(String(100))
    latitude: Mapped[Optional[float]] = mapped_column(Double)
    longitude: Mapped[Optional[float]] = mapped_column(Double)


class Family(Base):
    __tablename__ = "family"
    __table_args__ = {"schema": "data"}

    familyId: Mapped[int] = mapped_column("family_id", Integer, primary_key=True)
    tenant_id: Mapped[str] = mapped_column(String(20))
    family_name: Mapped[str] = mapped_column(String(255))
    family_address: Mapped[Optional[str]] = mapped_column(String(555))
    building_name: Mapped[Optional[str]] = mapped_column(String(255))
    building_address: Mapped[Optional[str]] = mapped_column(String(555))
    has_association: Mapped[Optional[bool]] = mapped_column(Boolean, default=False)
    association_name: Mapped[Optional[str]] = mapped_column(String(255))
    association_head_name: Mapped[Optional[str]] = mapped_column(String(255))
    association_head_phone: Mapped[Optional[str]] = mapped_column(String(30))
    head_voter_id: Mapped[Optional[int]] = mapped_column(ForeignKey("data.family_members.member_id"))
    phone: Mapped[Optional[str]] = mapped_column(String(15))
    points: Mapped[Optional[int]] = mapped_column(Integer)
    points_provided: Mapped[Optional[int]] = mapped_column(Integer)
    latitude: Mapped[Optional[float]] = mapped_column(Double)
    longitude: Mapped[Optional[float]] = mapped_column(Double)
    booth_id: Mapped[int] = mapped_column(ForeignKey("data.booths.booth_id"))
    association_id: Mapped[Optional[int]] = mapped_column(ForeignKey("data.association.association_id"))
    economic_status: Mapped[Optional[str]] = mapped_column(String(50))
    family_nature: Mapped[Optional[str]] = mapped_column(String(50))
    road_name: Mapped[Optional[str]] = mapped_column(String(255))
    building_number: Mapped[Optional[str]] = mapped_column(String(100))
    flat_number: Mapped[Optional[str]] = mapped_column(String(100))
    family_number: Mapped[Optional[str]] = mapped_column(String(100))
    tag_leader: Mapped[Optional[str]] = mapped_column(String(255))
    family_availability: Mapped[Optional[str]] = mapped_column(String(50))
    deleted: Mapped[bool] = mapped_column(Boolean, default=False)
    created_at: Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    updated_at: Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    created_by: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    created_by_name: Mapped[Optional[str]] = mapped_column(String(255), nullable=True)
    created_by_phone: Mapped[Optional[str]] = mapped_column(String(50), nullable=True)
    updated_by: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    updated_by_name: Mapped[Optional[str]] = mapped_column(String(255), nullable=True)
    updated_by_phone: Mapped[Optional[str]] = mapped_column(String(50), nullable=True)


class FamilyMember(Base):
    __tablename__ = "family_members"
    __table_args__ = {"schema": "data"}

    member_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    family_id: Mapped[int] = mapped_column(ForeignKey("data.family.family_id"))
    voter_id: Mapped[int] = mapped_column(ForeignKey("data.voters.voter_id"))
    is_head: Mapped[bool] = mapped_column(Boolean, default=False)


class Meeting(Base):
    __tablename__ = "meetings"
    __table_args__ = {"schema": "data"}

    meeting_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_id: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    title: Mapped[str] = mapped_column(String(255))
    start_time: Mapped[Optional[str]] = mapped_column(String(50))
    end_time: Mapped[Optional[str]] = mapped_column(String(50))
    latitude: Mapped[Optional[float]] = mapped_column(Double)
    longitude: Mapped[Optional[float]] = mapped_column(Double)
    radius: Mapped[Optional[int]] = mapped_column(Integer)
    recipients: Mapped[Optional[str]] = mapped_column(String(555))
    channels: Mapped[Optional[str]] = mapped_column(String(255))
    created_at: Mapped[Optional[datetime]] = mapped_column(DateTime, default=func.now())


class MeetingAttendance(Base):
    __tablename__ = "meeting_attendance"
    __table_args__ = {"schema": "data"}

    attendance_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    meeting_id: Mapped[int] = mapped_column(ForeignKey("data.meetings.meeting_id"))
    voter_id: Mapped[Optional[int]] = mapped_column(ForeignKey("data.voters.voter_id"), nullable=True)
    volunteer_name: Mapped[Optional[str]] = mapped_column(String(255))
    volunteer_phone: Mapped[Optional[str]] = mapped_column(String(50))
    distance: Mapped[Optional[float]] = mapped_column(Double)
    attended_at: Mapped[Optional[datetime]] = mapped_column(DateTime, default=func.now())


class VoterSnapshot(Base):
    __tablename__ = "voter_snapshot"
    __table_args__ = {"schema": "snapshot"}

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_id: Mapped[str] = mapped_column(String(20))
    assembly_code: Mapped[str] = mapped_column(String(12))
    ward_code: Mapped[Optional[str]] = mapped_column(String(20))
    booth_id: Mapped[Optional[int]] = mapped_column(Integer)
    s3_url: Mapped[str] = mapped_column(String)
    snapshot_level: Mapped[str] = mapped_column(String(20))
    version: Mapped[int] = mapped_column(Integer, default=1)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[Optional[datetime]] = mapped_column(DateTime)


class VoterChangeLog(Base):
    __tablename__ = "voter_changelog"
    __table_args__ = {"schema": "data"}

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tenant_id: Mapped[str] = mapped_column(String(20))
    voter_id: Mapped[int] = mapped_column(ForeignKey("data.voters.voter_id"))
    updated_by: Mapped[int] = mapped_column(ForeignKey("metastore.users.id"))
    field_name: Mapped[str] = mapped_column(String(100))
    old_value: Mapped[Optional[str]] = mapped_column(String)
    new_value: Mapped[Optional[str]] = mapped_column(String)
    updated_at: Mapped[datetime] = mapped_column(DateTime)
    update_latitude: Mapped[Optional[float]] = mapped_column(Double)
    update_longitude: Mapped[Optional[float]] = mapped_column(Double)


class VoterEnrichment(Base):
    __tablename__ = "voter_enrichment"
    __table_args__ = {"schema": "public"}

    epic: Mapped[str] = mapped_column(String(20), primary_key=True)
    ward_code: Mapped[Optional[str]] = mapped_column(String(20))
    booth_no: Mapped[Optional[str]] = mapped_column(String(20))
    first_middle_name_en: Mapped[Optional[str]] = mapped_column(String)
    last_name_en: Mapped[Optional[str]] = mapped_column(String)
    first_middle_name_local: Mapped[Optional[str]] = mapped_column(String)
    last_name_local: Mapped[Optional[str]] = mapped_column(String)
    relation_type: Mapped[Optional[str]] = mapped_column(String)
    relation_first_middle_name_en: Mapped[Optional[str]] = mapped_column(String)
    relation_last_name_en: Mapped[Optional[str]] = mapped_column(String)
    relation_first_middle_name_local: Mapped[Optional[str]] = mapped_column(String)
    relation_last_name_local: Mapped[Optional[str]] = mapped_column(String)
    house_no_en: Mapped[Optional[str]] = mapped_column(String)
    house_no_local: Mapped[Optional[str]] = mapped_column(String)
    gender: Mapped[Optional[str]] = mapped_column(String(20))
    age: Mapped[Optional[int]] = mapped_column(Integer)
    dob: Mapped[Optional[str]] = mapped_column(String(50))
    mobile: Mapped[Optional[str]] = mapped_column(String(30))
    address_en: Mapped[Optional[str]] = mapped_column(String)
    address_local: Mapped[Optional[str]] = mapped_column(String)
    status: Mapped[Optional[str]] = mapped_column(String(100))
    community: Mapped[Optional[str]] = mapped_column(String(100))
    caste: Mapped[Optional[str]] = mapped_column(String(100))
    residence_type: Mapped[Optional[str]] = mapped_column(String(100))
    civic_issue: Mapped[Optional[str]] = mapped_column(String)
    mother_tongue: Mapped[Optional[str]] = mapped_column(String(100))
    team: Mapped[Optional[str]] = mapped_column(String(100))
    ownership: Mapped[Optional[str]] = mapped_column(String(100))
    education: Mapped[Optional[str]] = mapped_column(String(100))
    nature_of_voter: Mapped[Optional[str]] = mapped_column(String(100))
    voter_points: Mapped[Optional[str]] = mapped_column(String(100))
    govt_scheme_tracking: Mapped[Optional[str]] = mapped_column(String)
    engagement_potential: Mapped[Optional[str]] = mapped_column(String(100))
    if_shifted: Mapped[Optional[str]] = mapped_column(String(100))
    notes: Mapped[Optional[str]] = mapped_column(String)
    present_address: Mapped[Optional[str]] = mapped_column(String)
    new_ward: Mapped[Optional[str]] = mapped_column(String(100))
    new_booth_no: Mapped[Optional[str]] = mapped_column(String(100))
    new_serial_no: Mapped[Optional[str]] = mapped_column(String(100))
    not_available_reason: Mapped[Optional[str]] = mapped_column(String)
    latitude: Mapped[Optional[float]] = mapped_column(Double)
    longitude: Mapped[Optional[float]] = mapped_column(Double)
    updated_fields: Mapped[str] = mapped_column(String, default="[]")
    updated_by: Mapped[Optional[int]] = mapped_column(ForeignKey("metastore.users.id"))
    updated_by_name: Mapped[Optional[str]] = mapped_column(String(255))
    updated_by_phone: Mapped[Optional[str]] = mapped_column(String(50))
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class PollDayConfig(Base):
    __tablename__ = "poll_day_config"
    __table_args__ = {"schema": "public"}

    config_id: Mapped[int] = mapped_column(Integer, primary_key=True)
    assembly_id: Mapped[Optional[int]] = mapped_column(Integer)
    ward_id: Mapped[Optional[int]] = mapped_column(Integer)
    enabled: Mapped[bool] = mapped_column(Boolean, default=False)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=func.now())


class MessageTemplate(Base):
    __tablename__ = "message_templates"
    __table_args__ = {"schema": "metastore"}

    template_id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    tenant_id: Mapped[Optional[str]] = mapped_column(String(20))
    ward_id: Mapped[Optional[int]] = mapped_column(Integer)
    channel: Mapped[str] = mapped_column(String(20), default="WHATSAPP")
    authority_name: Mapped[Optional[str]] = mapped_column(String(255))
    election_name: Mapped[Optional[str]] = mapped_column(String(255))
    assembly_label: Mapped[Optional[str]] = mapped_column(String(255))
    ward_label: Mapped[Optional[str]] = mapped_column(String(255))
    candidate_name: Mapped[Optional[str]] = mapped_column(String(255))
    candidate_party: Mapped[Optional[str]] = mapped_column(String(255))
    candidate_ward_label: Mapped[Optional[str]] = mapped_column(String(255))
    vote_date: Mapped[Optional[str]] = mapped_column(String(50))
    vote_time: Mapped[Optional[str]] = mapped_column(String(50))
    social_link: Mapped[Optional[str]] = mapped_column(String(255))
    booth_location_link: Mapped[Optional[str]] = mapped_column(String(255))
    banner_url: Mapped[Optional[str]] = mapped_column(String)
    show_logo: Mapped[Optional[bool]] = mapped_column(Boolean, default=True)
    enabled: Mapped[Optional[bool]] = mapped_column(Boolean, default=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


engine = create_engine(DATABASE_URL, future=True, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------------------------
# API response helpers
# ---------------------------
def _normalize_assembly_code(code: Optional[Any]) -> Optional[str]:
    if not code:
        return None
    c = str(code).strip()
    if c.isdigit() and len(c) < 12:
        return c.zfill(12)
    return c


def api_success(message: str, result: Any) -> Dict[str, Any]:
    return {"success": True, "message": message, "data": {"result": result}}


def api_error(message: str, error: Any) -> Dict[str, Any]:
    return {"success": False, "message": message, "data": {"error": error}}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


class InvalidCredentialsException(Exception):
    def __init__(self, message: str):
        super().__init__(message)
        self.code = "INVALID_CREDENTIALS"
        self.details = message


class ResourceAlreadyExistsException(Exception):
    def __init__(self, resource_name: str, field_name: str, field_value: Any):
        super().__init__(f"{resource_name} already exists with {field_name}: '{field_value}'")
        self.code = "RESOURCE_ALREADY_EXISTS"
        self.resource_name = resource_name
        self.field_name = field_name
        self.field_value = field_value
        self.details = f"{resource_name} with {field_name} '{field_value}' already exists."


class ResourceNotFoundException(Exception):
    def __init__(self, resource_name: str, field_name: str, field_value: Any):
        super().__init__(f"{resource_name} not found with {field_name}: '{field_value}'")
        self.code = "RESOURCE_NOT_FOUND"
        self.resource_name = resource_name
        self.field_name = field_name
        self.field_value = field_value
        self.details = f"{resource_name} with {field_name} '{field_value}' was not found."


# ---------------------------
# Auth
# ---------------------------
@dataclass
class JwtUserDetails:
    phone: str
    firstName: str
    role: str
    tenantId: Optional[str]
    assignmentType: Optional[str]
    assignmentId: Optional[str]


def _generate_token(first_name: str, role: str, tenant_id: Optional[str], assignment_type: Optional[str], assignment_id: Optional[str], phone: str) -> str:
    payload = {
        "sub": first_name,
        "role": role,
        "tenantId": tenant_id,
        "firstName": first_name,
        "assignmentType": assignment_type,
        "assignmentId": assignment_id,
        "phone": phone,
        "iat": int(datetime.now(timezone.utc).timestamp()),
        "exp": int((datetime.now(timezone.utc) + timedelta(seconds=JWT_EXPIRATION_SECONDS)).timestamp()),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def _parse_token(token: str) -> Dict[str, Any]:
    return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])


def _external_base_url(request: Request) -> str:
    forwarded_proto = request.headers.get("x-forwarded-proto")
    forwarded_host = request.headers.get("x-forwarded-host")
    host = forwarded_host or request.headers.get("host")
    scheme = forwarded_proto or request.url.scheme
    if host:
        return f"{scheme}://{host}"
    return str(request.base_url).rstrip("/")


def _resolve_tenant_id(user: Optional[User] = None, payload: Optional[Dict[str, Any]] = None) -> Optional[str]:
    if user is not None and getattr(user, "tenant", None) is not None:
        tenant_value = getattr(user.tenant, "tenant_id", None)
        if tenant_value:
            return str(tenant_value)
    if payload is not None:
        tenant_value = payload.get("tenantId")
        if tenant_value:
            return str(tenant_value)
    return None


def _resolve_tenant_id_for_entity(entity: Any) -> Optional[str]:
    if entity is None:
        return None
    if isinstance(entity, VolunteerUser):
        return entity.tenant_id
    return _resolve_tenant_id(user=entity)


def _auth_user(request: Request, db: Session) -> JwtUserDetails:
    auth = request.headers.get("Authorization")
    if not auth or not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid Authorization header")

    token = auth[7:]
    try:
        payload = _parse_token(token)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    base_query = (
        db.query(User)
        .join(Tenant, User.tenant_ref == Tenant.id, isouter=True)
        .filter(User.first_name == payload.get("firstName"), User.phone == payload.get("phone"))
    )
    user = base_query.first()
    volunteer = None
    if not user:
        volunteer = (
            db.query(VolunteerUser)
            .filter(VolunteerUser.first_name == payload.get("firstName"), VolunteerUser.phone == payload.get("phone"))
            .first()
        )

    target = user or volunteer
    if target and (target.blocked or target.deleted):
        raise HTTPException(status_code=403, detail="Please contact Admin")

    return JwtUserDetails(
        phone=payload.get("phone"),
        firstName=payload.get("firstName"),
        role=payload.get("role"),
        tenantId=_resolve_tenant_id_for_entity(target) or payload.get("tenantId"),
        assignmentType=payload.get("assignmentType"),
        assignmentId=payload.get("assignmentId"),
    )


def get_current_user(request: Request, db: Session = Depends(get_db)) -> JwtUserDetails:
    return _auth_user(request, db)


def require_roles(*roles: str):
    def dep(user: JwtUserDetails = Depends(get_current_user)) -> JwtUserDetails:
        role = (user.role or "").replace("ROLE_", "")
        if role not in roles:
            if role in {"ASSEMBLY", "WARD", "BOOTH"} and "USER" in roles:
                return user
            raise HTTPException(status_code=401, detail="Access denied")
        return user

    return dep


# ---------------------------
# Schemas
# ---------------------------
class MeetingCreateRequest(BaseModel):
    title: str
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    radius: Optional[int] = None
    recipients: Optional[str] = None
    channels: Optional[str] = None

class LoginRequest(BaseModel):
    firstName: str
    phone: str


class UserDetailsIn(BaseModel):
    role: str
    tenantId: Optional[str] = None
    assignmentType: Optional[str] = None
    assignmentId: Optional[str] = None
    firstName: str
    phone: Optional[str] = None
    profilePicUrl: Optional[str] = None
    blocked: Optional[bool] = None
    deleted: Optional[bool] = None


class UserBlockRequest(BaseModel):
    firstName: Optional[str] = None
    phone: Optional[str] = None
    userEmail: Optional[str] = None
    block: bool


class UserDeleteRequest(BaseModel):
    firstName: Optional[str] = None
    phone: Optional[str] = None
    userEmail: Optional[str] = None
    delete: bool


class UserBulkActionRequest(BaseModel):
    userFirstNames: Optional[List[str]] = None
    userEmails: Optional[List[str]] = None
    action: bool


class TenantDtoIn(BaseModel):
    id: Optional[int] = None
    tenantId: Optional[str] = None
    name: str
    description: Optional[str] = None
    contactEmail: str
    contactPhone: Optional[str] = None
    active: Optional[bool] = True


class VoterUpdatePayload(BaseModel):
    updateLocationLat: Optional[float] = None
    updateLocationLng: Optional[float] = None
    updateRequest: Dict[str, Any]


class PublicVoterUpdatePayload(BaseModel):
    wardCode: Optional[str] = None
    boothNo: Optional[str] = None
    updateRequest: Dict[str, Any]


class CreateAssociationRequest(BaseModel):
    associationName: str
    associationAddress: Optional[str] = None
    associationHeadName: Optional[str] = None
    phone: Optional[str] = None
    longitude: Optional[float] = None
    latitude: Optional[float] = None
    boothId: int


class CreateFamilyRequest(BaseModel):
    familyName: str
    familyAddress: Optional[str] = None
    roadName: Optional[str] = None
    buildingNumber: Optional[str] = None
    buildingName: Optional[str] = None
    flatNumber: Optional[str] = None
    familyNumber: Optional[str] = None
    tagLeader: Optional[str] = None
    familyAvailability: Optional[str] = None
    buildingAddress: Optional[str] = None
    hasAssociation: Optional[bool] = None
    associationName: Optional[str] = None
    associationHeadName: Optional[str] = None
    associationHeadPhone: Optional[str] = None
    phone: Optional[str] = None
    points: Optional[int] = None
    pointsProvided: Optional[int] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    boothId: int
    wardId: Optional[int] = None
    associationId: Optional[int] = None
    headEpicNo: str
    memberEpicNos: List[str]
    economicStatus: Optional[str] = None
    familyNature: Optional[str] = None


class UpdateFamilyRequest(CreateFamilyRequest):
    pass


class MessageTemplatePayload(BaseModel):
    wardId: Optional[int] = None
    channel: str = "WHATSAPP"
    authorityName: Optional[str] = None
    electionName: Optional[str] = None
    assemblyLabel: Optional[str] = None
    wardLabel: Optional[str] = None
    candidateName: Optional[str] = None
    candidateParty: Optional[str] = None
    candidateWardLabel: Optional[str] = None
    voteDate: Optional[str] = None
    voteTime: Optional[str] = None
    socialLink: Optional[str] = None
    boothLocationLink: Optional[str] = None
    bannerUrl: Optional[str] = None
    showLogo: Optional[bool] = None
    enabled: Optional[bool] = None


class UserProfileDto(BaseModel):
    firstName: str
    phone: str
    profilePicUrl: Optional[str] = None
    tenantId: Optional[str] = None
    role: Optional[str] = None


class VolunteerCreateRequest(BaseModel):
    firstName: str
    phone: str
    workingLevel: str
    assemblyIds: Optional[List[int]] = None
    wardIds: Optional[List[int]] = None
    boothIds: Optional[List[int]] = None


class VolunteerUpdateRequest(VolunteerCreateRequest):
    pass


# ---------------------------
# Utility converters
# ---------------------------
def to_user_details(u: User) -> Dict[str, Any]:
    def parse_ids(value: Optional[str]) -> List[int]:
        if not value:
            return []
        return [int(v) for v in str(value).split(",") if str(v).strip().isdigit()]

    return {
        "id": u.id,
        "role": u.role,
        "tenantId": _resolve_tenant_id_for_entity(u),
        "assignmentType": u.assignment_type,
        "assignmentId": u.assignment_id,
        "workingLevel": getattr(u, "working_level", None) or u.assignment_type,
        "assemblyIds": parse_ids(getattr(u, "assembly_ids", None)),
        "wardIds": parse_ids(getattr(u, "ward_ids", None)),
        "boothIds": parse_ids(getattr(u, "booth_ids", None)),
        "firstName": u.first_name,
        "lastName": "",
        "userName": u.first_name,
        "phone": u.phone,
        "profilePicUrl": u.profile_pic_url,
        "blocked": u.blocked,
        "deleted": u.deleted,
    }


def to_tenant_dto(t: Tenant) -> Dict[str, Any]:
    return {
        "id": t.id,
        "tenantId": t.tenant_id,
        "name": t.name,
        "description": t.description,
        "contactEmail": t.contact_email,
        "contactPhone": t.contact_phone,
        "active": t.active,
    }


def _serialize_id_list(values: Optional[List[int]]) -> Optional[str]:
    if not values:
        return None
    return ",".join(str(int(v)) for v in values if v is not None)


def _first_id(values: Optional[List[int]]) -> Optional[int]:
    if not values:
        return None
    for v in values:
        if v is not None:
            return int(v)
    return None


def build_page(content: List[Any], page: int, size: int, total: int, sort_field: str = "", direction: str = "asc") -> Dict[str, Any]:
    total_pages = (total + size - 1) // size if size else 1
    return {
        "content": content,
        "pageable": {
            "pageNumber": page,
            "pageSize": size,
            "sort": {"sorted": bool(sort_field), "unsorted": not bool(sort_field), "empty": not bool(sort_field)},
            "offset": page * size,
            "paged": True,
            "unpaged": False,
        },
        "last": page >= max(total_pages - 1, 0),
        "totalPages": total_pages,
        "totalElements": total,
        "size": size,
        "number": page,
        "sort": {"sorted": bool(sort_field), "unsorted": not bool(sort_field), "empty": not bool(sort_field)},
        "first": page == 0,
        "numberOfElements": len(content),
        "empty": len(content) == 0,
    }


def _parse_id_list(raw: Optional[str]) -> List[int]:
    if not raw:
        return []
    items = []
    for part in str(raw).split(","):
        part = part.strip()
        if part.isdigit():
            items.append(int(part))
    return items


def _get_access_scope(db: Session, current: JwtUserDetails) -> Optional[Dict[str, Any]]:
    role = (current.role or "").replace("ROLE_", "")
    if role == "SUPER_ADMIN":
        return None

    volunteer = (
        db.query(VolunteerUser)
        .filter(VolunteerUser.first_name == current.firstName, VolunteerUser.phone == current.phone)
        .first()
    )
    assignment_type = normalize_optional_text(current.assignmentType) or (
        normalize_optional_text(volunteer.assignment_type) if volunteer else None
    ) or (normalize_optional_text(volunteer.working_level) if volunteer else None) or role
    assignment_type = (assignment_type or role).upper()

    assembly_ids = _parse_id_list(volunteer.assembly_ids) if volunteer else []
    ward_ids = _parse_id_list(volunteer.ward_ids) if volunteer else []
    booth_ids = _parse_id_list(volunteer.booth_ids) if volunteer else []

    if not (assembly_ids or ward_ids or booth_ids):
        fallback_ids = _parse_id_list(str(current.assignmentId or ""))
        if assignment_type == "ASSEMBLY":
            assembly_ids = [_normalize_assembly_code(aid) for aid in fallback_ids]
        elif assignment_type == "WARD":
            ward_ids = fallback_ids
        elif assignment_type == "BOOTH":
            booth_ids = fallback_ids

    return {
        "assignment_type": assignment_type,
        "assembly_ids": assembly_ids,
        "ward_ids": ward_ids,
        "booth_ids": booth_ids,
        "role": role,
    }


def _build_public_tenant_filter(current: JwtUserDetails) -> tuple[str, Dict[str, Any]]:
    if current.role == "SUPER_ADMIN":
        return "", {}
    return " AND (tenant_id = :tid OR tenant_id IS NULL OR tenant_id = '')", {"tid": current.tenantId}


def _resolve_access_scope_ids(db: Session, current: JwtUserDetails) -> Optional[Dict[str, Any]]:
    scope = _get_access_scope(db, current)
    if not scope:
        return None

    allowed_assembly_ids = set(scope.get("assembly_ids") or [])
    allowed_ward_ids = set(scope.get("ward_ids") or [])
    allowed_booth_ids = set(scope.get("booth_ids") or [])

    # booth_ids may be stored as booth_no values, so map both booth_id and booth_no to booth_id
    if allowed_booth_ids:
        booth_cols = _get_table_columns(db, "public", "booths")
        booth_pk_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
        booth_no_col = "booth_no" if "booth_no" in booth_cols else None
        if booth_pk_col:
            raw_values = [str(v).strip() for v in allowed_booth_ids if v is not None and str(v).strip() != ""]
            ids_int = sorted({int(v) for v in raw_values if v.isdigit()})
            ids_text = sorted({str(v) for v in raw_values if str(v).strip()})
            where_parts = []
            params: Dict[str, Any] = {}
            if ids_int:
                clause_id, params_id = _build_in_clause(booth_pk_col, ids_int, "scope_booth_id")
                where_parts.append(f"({clause_id})")
                params.update(params_id)
            if booth_no_col and ids_text:
                clause_no, params_no = _build_in_clause(f"CAST({booth_no_col} AS TEXT)", ids_text, "scope_booth_no")
                where_parts.append(f"({clause_no})")
                params.update(params_no)
            if not where_parts:
                where_parts.append("1=0")
            t_clause, t_params = _build_public_tenant_filter(current)
            params.update(t_params)
            where_clause = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
            rows = db.execute(
                text(
                    f"""
                    SELECT {booth_pk_col} AS booth_id, {booth_no_col if booth_no_col else booth_pk_col} AS booth_no
                    FROM public.booths
                    {where_clause} {t_clause}
                    """
                ),
                params,
            ).all()
            allowed_booth_ids = {row.booth_id for row in rows if row.booth_id is not None}

    # Hierarchical expansion using public schema to avoid sparse data from 'data' schema
    ward_cols = _get_table_columns(db, "public", "wards")
    booth_cols = _get_table_columns(db, "public", "booths")

    w_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
    w_asm_id_col = "assembly_id" if "assembly_id" in ward_cols else ("assembly_no" if "assembly_no" in ward_cols else None)
    
    b_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
    b_ward_id_col = "ward_id" if "ward_id" in booth_cols else None

    # Expansion: Assembly -> Wards
    if allowed_assembly_ids and w_id_col and w_asm_id_col:
        aids = sorted([int(v) for v in allowed_assembly_ids if v is not None])
        if aids:
            res = db.execute(text(f"SELECT {w_id_col} FROM public.wards WHERE {w_asm_id_col} IN :aids"), {"aids": tuple(aids)}).all()
            allowed_ward_ids.update([r[0] for r in res if r[0] is not None])

    # Expansion: Booths -> Wards (Upward)
    if allowed_booth_ids and b_id_col and b_ward_id_col:
        bids = sorted([int(v) for v in allowed_booth_ids if v is not None])
        if bids:
            res = db.execute(text(f"SELECT {b_ward_id_col} FROM public.booths WHERE {b_id_col} IN :bids"), {"bids": tuple(bids)}).all()
            allowed_ward_ids.update([r[0] for r in res if r[0] is not None])

    # Expansion: Wards -> Booths (Downward)
    if allowed_ward_ids and b_id_col and b_ward_id_col:
        wids = sorted([int(v) for v in allowed_ward_ids if v is not None])
        if wids:
            res = db.execute(text(f"SELECT {b_id_col} FROM public.booths WHERE {b_ward_id_col} IN :wids"), {"wids": tuple(wids)}).all()
            allowed_booth_ids.update([r[0] for r in res if r[0] is not None])

    # Expansion: Wards -> Assemblies (Upward)
    if allowed_ward_ids and w_id_col and w_asm_id_col:
        wids = sorted([int(v) for v in allowed_ward_ids if v is not None])
        if wids:
            res = db.execute(text(f"SELECT {w_asm_id_col} FROM public.wards WHERE {w_id_col} IN :wids"), {"wids": tuple(wids)}).all()
            allowed_assembly_ids.update([r[0] for r in res if r[0] is not None])

    allowed_assembly_ids = {v for v in allowed_assembly_ids if v is not None}
    allowed_ward_ids = {v for v in allowed_ward_ids if v is not None}
    allowed_booth_ids = {v for v in allowed_booth_ids if v is not None}

    return {
        **scope,
        "allowed_assembly_ids": allowed_assembly_ids,
        "allowed_ward_ids": allowed_ward_ids,
        "allowed_booth_ids": allowed_booth_ids,
    }


def _build_in_clause(column_expr: str, values: List[Any], prefix: str) -> tuple[str, Dict[str, Any]]:
    if not values:
        return "1=0", {}
    params: Dict[str, Any] = {}
    placeholders: List[str] = []
    for idx, value in enumerate(values):
        key = f"{prefix}_{idx}"
        params[key] = value
        placeholders.append(f":{key}")
    return f"{column_expr} IN ({', '.join(placeholders)})", params


def _build_comma_list_filter(column, values: List[int]):
    if not values:
        return None
    clauses = []
    for value in values:
        value_str = str(value)
        clauses.append(column == value_str)
        clauses.append(column.like(f"{value_str},%"))
        clauses.append(column.like(f"%,{value_str},%"))
        clauses.append(column.like(f"%,{value_str}"))
    return or_(*clauses) if clauses else None


def _get_s3_client():
    return boto3.client(
        "s3",
        region_name=AWS_REGION,
        aws_access_key_id=AWS_ACCESS_KEY or None,
        aws_secret_access_key=AWS_SECRET_KEY or None,
    )


def s3_extract_key(url: str) -> str:
    idx = url.find(".com/")
    if idx == -1:
        raise ValueError(f"Invalid S3 URL: {url}")
    return url[idx + 5 :]


def s3_presigned_url(key: str, minutes: int, fallback_url: Optional[str] = None) -> str:
    client = _get_s3_client()
    try:
        return client.generate_presigned_url(
            "get_object",
            Params={"Bucket": AWS_BUCKET, "Key": key},
            ExpiresIn=minutes * 60,
        )
    except NoCredentialsError:
        if fallback_url:
            return fallback_url
        if AWS_BUCKET:
            return f"https://{AWS_BUCKET}.s3.amazonaws.com/{key}"
        raise


def s3_upload_bytes(content: bytes, content_type: str, key: str) -> str:
    # Use Base64 storage instead of S3 as requested
    try:
        b64 = base64.b64encode(content).decode("utf-8")
        return f"data:{content_type};base64,{b64}"
    except Exception as e:
        print(f"Base64 conversion failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Image processing failed: {str(e)}")


def normalize_assembly_code(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if len(s) == 12:
        return s
    return f"{int(s):012d}"


def normalize_optional_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    trimmed = value.strip()
    return trimmed if trimmed else None


def normalize_phone(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return digits if digits else None


def parse_optional_bool(value: Optional[str | bool]) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    cleaned = normalize_optional_text(value)
    if cleaned is None:
        return None
    lowered = cleaned.lower()
    if lowered in {"true", "1", "yes"}:
        return True
    if lowered in {"false", "0", "no"}:
        return False
    raise ValueError(f"Invalid boolean value: {value}")


def resolve_user_first_name(first_name: Optional[str], fallback_username: Optional[str]) -> Optional[str]:
    return normalize_optional_text(first_name) or normalize_optional_text(fallback_username)


def resolve_bulk_usernames(payload: UserBulkActionRequest) -> List[str]:
    candidates = payload.userFirstNames if payload.userFirstNames is not None else payload.userEmails
    if not candidates:
        raise ValueError("At least one user identifier is required")
    usernames = [name.strip() for name in candidates if name and name.strip()]
    if not usernames:
        raise ValueError("At least one valid user identifier is required")
    return usernames


# ---------------------------
# App + exception handlers
# ---------------------------
app = FastAPI(title="VoteBase Python API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(extract_router, prefix=CONTEXT_PATH)


@app.middleware("http")
async def log_all_requests(request: Request, call_next):
    started_at = datetime.now(timezone.utc)
    query = request.url.query
    path_with_query = f"{request.url.path}?{query}" if query else request.url.path
    auth_present = bool(request.headers.get("Authorization"))
    body_log = None

    target_paths = {
        f"{CONTEXT_PATH}/api/auth/login",
        f"{CONTEXT_PATH}/api/booth",
        f"{CONTEXT_PATH}/api/voters/snapshot",
    }

    def mask_sensitive(value: Any) -> Any:
        if isinstance(value, dict):
            out: Dict[str, Any] = {}
            for k, v in value.items():
                lk = k.lower()
                if lk in {"phone", "token", "password", "authorization"}:
                    out[k] = "***"
                else:
                    out[k] = mask_sensitive(v)
            return out
        if isinstance(value, list):
            return [mask_sensitive(v) for v in value]
        return value

    if request.url.path in target_paths:
        content_type = (request.headers.get("content-type") or "").lower()
        if "application/json" in content_type:
            try:
                raw_body = await request.body()
                if raw_body:
                    parsed = json.loads(raw_body.decode("utf-8"))
                    body_log = json.dumps(mask_sensitive(parsed), ensure_ascii=False)

                    # Re-inject the already-read body so downstream handlers can read it.
                    async def receive():
                        return {"type": "http.request", "body": raw_body, "more_body": False}

                    request = Request(request.scope, receive)
            except Exception:
                body_log = "<unparseable-json>"

    try:
        response = await call_next(request)
        duration_ms = int((datetime.now(timezone.utc) - started_at).total_seconds() * 1000)
        line = (
            f"[API_HIT] method={request.method} path={path_with_query} "
            f"status={response.status_code} auth={'yes' if auth_present else 'no'} "
            f"durationMs={duration_ms}"
        )
        if body_log is not None:
            line += f" body={body_log}"
        print(line)
        return response
    except Exception as ex:
        duration_ms = int((datetime.now(timezone.utc) - started_at).total_seconds() * 1000)
        line = (
            f"[API_HIT] method={request.method} path={path_with_query} "
            f"status=500 auth={'yes' if auth_present else 'no'} durationMs={duration_ms} "
            f"error={type(ex).__name__}"
        )
        if body_log is not None:
            line += f" body={body_log}"
        print(line)
        raise


@app.exception_handler(InvalidCredentialsException)
async def handle_invalid_credentials(_: Request, ex: InvalidCredentialsException):
    return JSONResponse(
        status_code=401,
        content=api_error(ex.args[0], {"timestamp": now_iso(), "code": ex.code, "details": ex.details}),
    )


@app.exception_handler(ResourceAlreadyExistsException)
async def handle_resource_exists(_: Request, ex: ResourceAlreadyExistsException):
    return JSONResponse(
        status_code=409,
        content=api_error(
            ex.args[0],
            {
                "timestamp": now_iso(),
                "code": ex.code,
                "resource": ex.resource_name,
                "field": ex.field_name,
                "value": ex.field_value,
                "details": ex.details,
            },
        ),
    )


@app.exception_handler(ResourceNotFoundException)
async def handle_resource_not_found(_: Request, ex: ResourceNotFoundException):
    return JSONResponse(
        status_code=404,
        content=api_error(
            "Resource not found",
            {
                "timestamp": now_iso(),
                "code": ex.code,
                "resource": ex.resource_name,
                "field": ex.field_name,
                "value": ex.field_value,
                "details": ex.details,
            },
        ),
    )


@app.exception_handler(ValueError)
async def handle_validation(_: Request, ex: ValueError):
    return JSONResponse(
        status_code=400,
        content=api_error("Validation failed", {"code": "INVALID_PARAMETER", "details": str(ex), "timestamp": now_iso()}),
    )


@app.exception_handler(Exception)
async def handle_generic(_: Request, ex: Exception):
    if isinstance(ex, HTTPException):
        return JSONResponse(status_code=ex.status_code, content=api_error(ex.detail, {"timestamp": now_iso(), "code": "HTTP_ERROR", "details": ex.detail}))
    return JSONResponse(
        status_code=500,
        content=api_error("An unexpected error occurred", {"timestamp": now_iso(), "code": "INTERNAL_SERVER_ERROR", "details": str(ex)}),
    )


# ---------------------------
# Startup: super admin seed
# ---------------------------
@app.on_event("startup")
def startup_seed_super_admin() -> None:
    print(f"[DB_TARGET] { _safe_db_target(DATABASE_URL) }")
    db = SessionLocal()
    try:
        existing = db.query(User).filter(User.role == "SUPER_ADMIN").first()
        if not existing:
            super_admin = User(
                first_name=SUPERADMIN_USERNAME,
                phone="8867038709",
                role="SUPER_ADMIN",
                tenant_ref=None,
                assignment_type=None,
                assignment_id=None,
                blocked=False,
                deleted=False,
            )
            db.add(super_admin)
            db.commit()
    finally:
        db.close()


@app.on_event("startup")
def startup_ensure_voter_enrichment() -> None:
    VoterEnrichment.__table__.create(bind=engine, checkfirst=True)
    VolunteerUser.__table__.create(bind=engine, checkfirst=True)
    MessageTemplate.__table__.create(bind=engine, checkfirst=True)
    db = SessionLocal()
    try:
        enrichment_cols = {
            row[0]
            for row in db.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = 'public'
                      AND table_name = 'voter_enrichment'
                    """
                )
            ).all()
        }
        if "updated_by_name" not in enrichment_cols:
            db.execute(text("ALTER TABLE public.voter_enrichment ADD COLUMN updated_by_name varchar(255)"))
        if "updated_by_phone" not in enrichment_cols:
            db.execute(text("ALTER TABLE public.voter_enrichment ADD COLUMN updated_by_phone varchar(50)"))

        existing_cols = {
            row[0]
            for row in db.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = 'metastore'
                      AND table_name = 'volunteer_users'
                    """
                )
            ).all()
        }
        if "working_level" not in existing_cols:
            db.execute(text("ALTER TABLE metastore.volunteer_users ADD COLUMN working_level varchar(30)"))
        if "assembly_ids" not in existing_cols:
            db.execute(text("ALTER TABLE metastore.volunteer_users ADD COLUMN assembly_ids text"))
        if "ward_ids" not in existing_cols:
            db.execute(text("ALTER TABLE metastore.volunteer_users ADD COLUMN ward_ids text"))
        if "booth_ids" not in existing_cols:
            db.execute(text("ALTER TABLE metastore.volunteer_users ADD COLUMN booth_ids text"))

        template_cols = {
            row[0]
            for row in db.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = 'metastore'
                      AND table_name = 'message_templates'
                    """
                )
            ).all()
        }
        if "authority_name" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN authority_name varchar(255)"))
        if "election_name" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN election_name varchar(255)"))
        if "assembly_label" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN assembly_label varchar(255)"))
        if "ward_label" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN ward_label varchar(255)"))
        if "candidate_name" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN candidate_name varchar(255)"))
        if "candidate_party" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN candidate_party varchar(255)"))
        if "candidate_ward_label" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN candidate_ward_label varchar(255)"))
        if "vote_date" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN vote_date varchar(50)"))
        if "vote_time" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN vote_time varchar(50)"))
        if "social_link" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN social_link varchar(255)"))
        if "booth_location_link" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN booth_location_link varchar(255)"))
        if "banner_url" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN banner_url text"))
        else:
            db.execute(text("ALTER TABLE metastore.message_templates ALTER COLUMN banner_url TYPE text"))
        if "show_logo" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN show_logo boolean DEFAULT TRUE"))
        if "enabled" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN enabled boolean DEFAULT FALSE"))
        if "created_at" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN created_at timestamp DEFAULT now()"))
        if "updated_at" not in template_cols:
            db.execute(text("ALTER TABLE metastore.message_templates ADD COLUMN updated_at timestamp DEFAULT now()"))

        # Transition profile pic column for Base64 support
        db.execute(text("ALTER TABLE metastore.users ALTER COLUMN profile_pic_url TYPE text"))

        family_cols = {
            row[0]
            for row in db.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = 'data'
                      AND table_name = 'family'
                    """
                )
            ).all()
        }
        if "building_name" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN building_name varchar(255)"))
        if "building_address" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN building_address varchar(555)"))
        if "has_association" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN has_association boolean DEFAULT FALSE"))
        if "association_name" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN association_name varchar(255)"))
        if "association_head_name" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN association_head_name varchar(255)"))
        if "association_head_phone" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN association_head_phone varchar(30)"))
        if "road_name" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN road_name varchar(255)"))
        if "building_number" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN building_number varchar(100)"))
        if "flat_number" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN flat_number varchar(100)"))
        if "family_number" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN family_number varchar(100)"))
        if "tag_leader" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN tag_leader varchar(255)"))
        if "family_availability" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN family_availability varchar(50)"))
        if "created_at" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN created_at timestamp"))
        if "updated_at" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN updated_at timestamp"))
        if "created_by" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN created_by integer"))
        if "created_by_name" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN created_by_name varchar(255)"))
        if "created_by_phone" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN created_by_phone varchar(50)"))
        if "updated_by" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN updated_by integer"))
        if "updated_by_name" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN updated_by_name varchar(255)"))
        if "updated_by_phone" not in family_cols:
            db.execute(text("ALTER TABLE data.family ADD COLUMN updated_by_phone varchar(50)"))

        col = db.execute(
            text(
                """
                SELECT data_type
                FROM information_schema.columns
                WHERE table_schema = 'metastore'
                  AND table_name = 'volunteer_users'
                  AND column_name = 'assignment_id'
                """
            )
        ).scalar()
        if col and col not in {"character varying", "text"}:
            db.execute(text("ALTER TABLE metastore.volunteer_users ALTER COLUMN assignment_id TYPE text USING assignment_id::text"))
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


# ---------------------------
# Routes
# ---------------------------
@app.post(f"{CONTEXT_PATH}/api/auth/login")
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    normalized_name = normalize_optional_text(payload.firstName) or ""
    normalized_phone = normalize_phone(payload.phone) or ""
    user = db.query(User).filter(func.lower(User.first_name) == normalized_name.lower(), User.phone == normalized_phone).first()
    if user and user.role not in {"SUPER_ADMIN", "ADMIN"} and not user.tenant:
        user = None
    volunteer = None
    if not user:
        volunteer = db.query(VolunteerUser).filter(func.lower(VolunteerUser.first_name) == normalized_name.lower(), VolunteerUser.phone == normalized_phone).first()
        if not volunteer and normalized_phone:
            # Fallback: allow login by phone only if the number matches a unique volunteer.
            volunteer = db.query(VolunteerUser).filter(VolunteerUser.phone == normalized_phone).first()
        if not volunteer and not user:
            raise InvalidCredentialsException("Invalid firstname or phone")

    tenant_id = None
    assignment_type = None
    assignment_id = 0
    if user:
        if user.tenant:
            tenant_id = user.tenant.tenant_id
        if user.role != "SUPER_ADMIN" and not tenant_id:
            raise InvalidCredentialsException("Tenant information missing for user")

        if user.role != "ADMIN" and user.role != "SUPER_ADMIN":
            if user.assignment_type is None or user.assignment_id == -1:
                raise InvalidCredentialsException("Assignment information missing for user")
        assignment_type = user.assignment_type
        assignment_id = user.assignment_id
    elif volunteer:
        if volunteer.blocked or volunteer.deleted:
            raise HTTPException(status_code=403, detail="Please contact Admin")
        tenant_id = volunteer.tenant_id
        assignment_type = volunteer.assignment_type
        assignment_id = volunteer.assignment_id

    effective = user or volunteer
    token = _generate_token(effective.first_name, effective.role, tenant_id, assignment_type, assignment_id, effective.phone)
    return api_success(
        "Login successful",
        {
            "userName": payload.firstName,
            "token": token,
            "role": effective.role,
            "tenantId": tenant_id,
            "assignmentType": assignment_type,
            "assignmentId": assignment_id,
            "assemblyIds": _parse_id_list(getattr(effective, "assembly_ids", "")) if hasattr(effective, "assembly_ids") else [],
            "wardIds": _parse_id_list(getattr(effective, "ward_ids", "")) if hasattr(effective, "ward_ids") else [],
            "boothIds": _parse_id_list(getattr(effective, "booth_ids", "")) if hasattr(effective, "booth_ids") else [],
        },
    )

@app.get(f"{CONTEXT_PATH}/api/me")
def get_me(db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    volunteer = (
        db.query(VolunteerUser)
        .filter(VolunteerUser.first_name == current.firstName, VolunteerUser.phone == current.phone)
        .first() if current.phone else None
    )
    user = (
        db.query(User)
        .filter(User.first_name == current.firstName, User.phone == current.phone)
        .first() if current.phone else None
    )
    effective = user or volunteer
    return api_success(
        "User profile fetched",
        {
            "userName": current.firstName,
            "role": current.role,
            "tenantId": current.tenantId,
            "assignmentType": current.assignmentType,
            "assignmentId": current.assignmentId,
            "assemblyIds": _parse_id_list(getattr(effective, "assembly_ids", "")) if hasattr(effective, "assembly_ids") else [],
            "wardIds": _parse_id_list(getattr(effective, "ward_ids", "")) if hasattr(effective, "ward_ids") else [],
            "boothIds": _parse_id_list(getattr(effective, "booth_ids", "")) if hasattr(effective, "booth_ids") else [],
        },
    )


@app.get(f"{CONTEXT_PATH}/api/admin/db-dump")
def admin_db_dump(
    background_tasks: BackgroundTasks,
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN"))
):
    if not DATABASE_URL:
        raise HTTPException(status_code=500, detail="DATABASE_URL not set")
    fd, path = tempfile.mkstemp(suffix=".sql")
    os.close(fd)
    
    try:
        clean_url = DATABASE_URL.replace("postgresql+psycopg2://", "postgresql://")
        # Try to find pg_dump in common paths if not in PATH
        pg_dump_path = "pg_dump"
        if not shutil.which("pg_dump"):
            for p in ["/opt/homebrew/bin/pg_dump", "/usr/local/bin/pg_dump", "/usr/bin/pg_dump"]:
                if os.path.exists(p):
                    pg_dump_path = p
                    break
        
        subprocess.run(
            [pg_dump_path, "--dbname", clean_url, "-f", path, "--no-owner", "--no-acl"],
            check=True
        )
    except Exception as e:
        if os.path.exists(path):
            os.remove(path)
        raise HTTPException(status_code=500, detail=f"Database dump failed: {str(e)}")
        
    return FileResponse(
        path=path,
        media_type="application/sql",
        filename=f"votabase_dump_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
        background=BackgroundTask(os.remove, path)
    )


@app.get(f"{CONTEXT_PATH}/api/admin/dashboard")
def get_dashboard(user: JwtUserDetails = Depends(require_roles("ADMIN"))):
    return api_success(f"Welcome, ['ROLE_{user.role}']", {"user": user.firstName})


@app.post(f"{CONTEXT_PATH}/api/user/register")
def register_user(
    payload: UserDetailsIn,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN")),
):
    if current.role not in {"ADMIN", "SUPER_ADMIN"}:
        raise HTTPException(status_code=403, detail="You are not authorized to create users")

    requested_role = (payload.role or "").strip().upper()
    role_to_create = "USER"
    if requested_role not in {"", "USER"}:
        raise ValueError("Invalid role in register request")

    tenant_id = current.tenantId if current.role == "ADMIN" else (payload.tenantId or "").strip()
    if not tenant_id and current.role == "SUPER_ADMIN":
        tenants = db.query(Tenant).all()
        if len(tenants) == 1:
            tenant_id = tenants[0].tenant_id
        else:
            raise HTTPException(status_code=400, detail="tenantId is required for SUPER_ADMIN")

    tenant = db.query(Tenant).filter(Tenant.tenant_id == tenant_id).first()
    if not tenant:
        raise ResourceNotFoundException("Tenant", "tenantId", tenant_id)

    exists = db.query(VolunteerUser).filter(VolunteerUser.first_name == payload.firstName, VolunteerUser.phone == payload.phone).first()
    if exists:
        raise ResourceAlreadyExistsException("registerUser", "userName", payload.firstName)

    volunteer = VolunteerUser(
        role=role_to_create,
        tenant_id=tenant.tenant_id,
        working_level=payload.assignmentType,
        assignment_type=payload.assignmentType,
        assignment_id=payload.assignmentId,
        first_name=payload.firstName,
        phone=payload.phone,
        blocked=False,
        deleted=False,
    )
    db.add(volunteer)
    db.commit()

    out = payload.model_dump()
    out["role"] = role_to_create
    out["tenantId"] = tenant_id
    out["userName"] = payload.firstName
    return api_success("User registered successfully", out)


@app.post(f"{CONTEXT_PATH}/api/volunteers")
def create_volunteer(
    payload: VolunteerCreateRequest,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD")),
):
    if current.role not in {"ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD"}:
        raise HTTPException(status_code=403, detail="You are not authorized to create volunteers")

    working_level = (payload.workingLevel or "").strip().upper()
    if working_level not in {"ASSEMBLY", "WARD", "BOOTH"}:
        raise ValueError("workingLevel must be ASSEMBLY, WARD, or BOOTH")

    phone = normalize_optional_text(payload.phone)
    if not phone or not phone.isdigit() or len(phone) != 10:
        raise ValueError("phone must be a 10 digit number")

    tenant_id = current.tenantId
    if not tenant_id:
        tenants = db.query(Tenant).all()
        if len(tenants) == 1:
            tenant_id = tenants[0].tenant_id
        elif len(tenants) == 0:
            tenant_id = ""
        else:
            raise HTTPException(status_code=400, detail="tenantId is required for SUPER_ADMIN")

    exists = db.query(VolunteerUser).filter(VolunteerUser.phone == phone).first()
    if exists:
        return JSONResponse(
            status_code=409,
            content=api_error("Volunteer already exists", {"phone": phone})
        )

    assembly_ids = payload.assemblyIds or []
    ward_ids = payload.wardIds or []
    booth_ids = payload.boothIds or []
    assignment_id = _first_id(assembly_ids) or _first_id(ward_ids) or _first_id(booth_ids)

    volunteer = VolunteerUser(
        role=working_level,
        tenant_id=tenant_id,
        working_level=working_level,
        assignment_type=working_level,
        assignment_id=str(assignment_id) if assignment_id is not None else None,
        assembly_ids=_serialize_id_list(assembly_ids),
        ward_ids=_serialize_id_list(ward_ids),
        booth_ids=_serialize_id_list(booth_ids),
        first_name=payload.firstName,
        phone=phone,
        blocked=False,
        deleted=False,
    )
    db.add(volunteer)
    db.commit()

    return api_success(
        "Volunteer created successfully",
        to_user_details(volunteer),
    )


@app.put(f"{CONTEXT_PATH}/api/volunteers")
def update_volunteer(
    payload: VolunteerUpdateRequest,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD")),
):
    phone = normalize_optional_text(payload.phone)
    if not phone or not phone.isdigit() or len(phone) != 10:
        raise ValueError("phone must be a 10 digit number")

    q = db.query(VolunteerUser).filter(VolunteerUser.phone == phone)
    if current.role != "SUPER_ADMIN" and current.tenantId:
        q = q.filter(VolunteerUser.tenant_id == current.tenantId)
    volunteer = q.first()
    if not volunteer:
        return api_error("Volunteer not found", {"details": f"Volunteer not found with phone: '{phone}'"})

    working_level = (payload.workingLevel or "").strip().upper()
    if working_level not in {"ASSEMBLY", "WARD", "BOOTH"}:
        raise ValueError("workingLevel must be ASSEMBLY, WARD, or BOOTH")

    assembly_ids = payload.assemblyIds or []
    ward_ids = payload.wardIds or []
    booth_ids = payload.boothIds or []
    assignment_id = _first_id(assembly_ids) or _first_id(ward_ids) or _first_id(booth_ids)

    volunteer.first_name = payload.firstName
    volunteer.working_level = working_level
    volunteer.assignment_type = working_level
    volunteer.assignment_id = str(assignment_id) if assignment_id is not None else None
    volunteer.assembly_ids = _serialize_id_list(assembly_ids)
    volunteer.ward_ids = _serialize_id_list(ward_ids)
    volunteer.booth_ids = _serialize_id_list(booth_ids)

    db.commit()
    db.refresh(volunteer)
    return api_success("Volunteer updated successfully", to_user_details(volunteer))


@app.get(f"{CONTEXT_PATH}/api/volunteers")
def list_volunteers(
    page: int = 0,
    size: int = 10,
    search: Optional[str] = None,
    blocked: Optional[str] = None,
    deleted: Optional[str] = None,
    workingLevel: Optional[str] = None,
    assemblyCode: Optional[str] = None,
    sortBy: str = "firstName",
    direction: str = "asc",
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD")),
):
    q = db.query(VolunteerUser).filter(VolunteerUser.role.in_(["USER", "ASSEMBLY", "WARD", "BOOTH"]))
    
    # Resolve tenant from assemblyCode if provided
    resolved_tenant_id = None
    if assemblyCode:
        normalized_code = _normalize_assembly_code(assemblyCode)
        assembly_row = db.execute(
            text("SELECT tenant_id FROM public.assembly WHERE assembly_code = :c OR CAST(assembly_no AS TEXT) = :n LIMIT 1"),
            {"c": normalized_code, "n": str(int(normalized_code)) if normalized_code and normalized_code.isdigit() else normalized_code}
        ).first()
        if assembly_row:
            resolved_tenant_id = assembly_row.tenant_id

    if current.role != "SUPER_ADMIN" and current.tenantId:
        q = q.filter(VolunteerUser.tenant_id == current.tenantId)
    elif resolved_tenant_id:
        # For SUPER_ADMIN or users without tenant, filter by resolved tenant OR allow legacy NULLs
        q = q.filter((VolunteerUser.tenant_id == resolved_tenant_id) | (VolunteerUser.tenant_id.is_(None)) | (VolunteerUser.tenant_id == ""))

    if assemblyCode:
        normalized_code = _normalize_assembly_code(assemblyCode)
        unpadded = str(int(normalized_code)) if normalized_code.isdigit() else normalized_code
        # Further restrict to ensure even if tenant_id is NULL, they belong to this assembly
        # Check both assignment_id and assembly_ids CSV string
        q = q.filter(
            (VolunteerUser.tenant_id == resolved_tenant_id) | 
            (VolunteerUser.assembly_ids.like(f"%{normalized_code}%")) |
            (VolunteerUser.assembly_ids.like(f"%{unpadded}%")) |
            (VolunteerUser.assignment_id == normalized_code) |
            (VolunteerUser.assignment_id == unpadded)
        )

    if current.role == "WARD":
        scope = _resolve_access_scope_ids(db, current)
        allowed_ward_ids = sorted(scope.get("allowed_ward_ids") or []) if scope else []
        allowed_booth_ids = sorted(scope.get("allowed_booth_ids") or []) if scope else []
        ward_filter = _build_comma_list_filter(VolunteerUser.ward_ids, allowed_ward_ids)
        booth_filter = _build_comma_list_filter(VolunteerUser.booth_ids, allowed_booth_ids)
        if ward_filter is not None and booth_filter is not None:
            q = q.filter(or_(ward_filter, booth_filter))
        elif ward_filter is not None:
            q = q.filter(ward_filter)
        elif booth_filter is not None:
            q = q.filter(booth_filter)
        else:
            q = q.filter(text("1=0"))

    blocked_filter = parse_optional_bool(blocked)
    deleted_filter = parse_optional_bool(deleted)
    level_filter = normalize_optional_text(workingLevel)

    if blocked_filter is not None:
        q = q.filter(VolunteerUser.blocked == blocked_filter)
    if deleted_filter is not None:
        q = q.filter(VolunteerUser.deleted == deleted_filter)
    if level_filter is not None:
        q = q.filter(VolunteerUser.working_level == level_filter)
    if search and search.strip():
        s = f"%{search.lower()}%"
        q = q.filter(or_(func.lower(VolunteerUser.first_name).like(s), VolunteerUser.phone.like(f"%{search}%")))

    sort_map = {
        "id": VolunteerUser.id,
        "firstName": VolunteerUser.first_name,
        "phone": VolunteerUser.phone,
        "role": VolunteerUser.role,
        "workingLevel": VolunteerUser.working_level,
        "assignmentType": VolunteerUser.assignment_type,
        "assignmentId": VolunteerUser.assignment_id,
        "blocked": VolunteerUser.blocked,
        "deleted": VolunteerUser.deleted,
    }
    sort_col = sort_map.get(sortBy, VolunteerUser.first_name)
    q = q.order_by(desc(sort_col) if direction.lower() == "desc" else asc(sort_col))

    total = q.count()
    users = q.offset(page * size).limit(size).all()
    def parse_ids(value: Optional[str]) -> List[int]:
        if not value:
            return []
        return [int(v) for v in str(value).split(",") if str(v).strip().isdigit()]

    ward_ids_all: set[int] = set()
    booth_ids_all: set[int] = set()
    for u in users:
        ward_ids_all.update(parse_ids(getattr(u, "ward_ids", None)))
        booth_ids_all.update(parse_ids(getattr(u, "booth_ids", None)))

    ward_name_map: Dict[int, str] = {}
    booth_name_map: Dict[int, str] = {}

    if ward_ids_all:
        # Fallback to public.wards just like the dropdowns do
        ward_cols = _get_table_columns(db, "public", "wards")
        id_col = "ward_id" if "ward_id" in ward_cols else ("ward_no" if "ward_no" in ward_cols else ("id" if "id" in ward_cols else None))
        name_col = "ward_name_en" if "ward_name_en" in ward_cols else ("name_en" if "name_en" in ward_cols else ("ward_name_local" if "ward_name_local" in ward_cols else None))
        code_col = "ward_code" if "ward_code" in ward_cols else None

        if id_col and name_col:
            clause, params = _build_in_clause(id_col, [str(m) for m in ward_ids_all], "mc")
            raw_rows = db.execute(
                text(f"SELECT {id_col} AS wid, {name_col} AS wname FROM public.wards WHERE {clause}"),
                params,
            ).all()
            for row in raw_rows:
                if row.wname: ward_name_map[int(row.wid)] = row.wname
        
        remaining_wards = ward_ids_all - set(ward_name_map.keys())
        if remaining_wards and code_col and name_col:
            clause, params = _build_in_clause(code_col, [str(m) for m in remaining_wards], "wc")
            raw_rows = db.execute(
                text(f"SELECT {code_col} AS wcode, {name_col} AS wname FROM public.wards WHERE {clause}"),
                params,
            ).all()
            for row in raw_rows:
                try:
                    if row.wname: ward_name_map[int(row.wcode)] = row.wname
                except Exception:
                    pass
        
        still_missing_wards = ward_ids_all - set(ward_name_map.keys())
        if still_missing_wards:
            ward_q = db.query(Ward.ward_id, Ward.ward_name_en)
            if current.tenantId is not None and current.role != "SUPER_ADMIN":
                ward_q = ward_q.filter(Ward.tenant_id == current.tenantId)
            ward_rows = ward_q.filter(Ward.ward_id.in_(list(still_missing_wards))).all()
            for r in ward_rows:
                if r.ward_name_en: ward_name_map[r.ward_id] = r.ward_name_en

    if booth_ids_all:
        booth_cols = _get_table_columns(db, "public", "booths")
        id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
        booth_no_col = "booth_no" if "booth_no" in booth_cols else ("id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None))
        ward_ref = "ward_id" if "ward_id" in booth_cols else ("ward_no" if "ward_no" in booth_cols else None)
        name_col = "booth_add_en" if "booth_add_en" in booth_cols else ("polling_station_adr_en" if "polling_station_adr_en" in booth_cols else None)
        
        if name_col:
            # 1. Try direct ID match
            if id_col:
                clause, params = _build_in_clause(id_col, [str(m) for m in booth_ids_all], "bc")
                rows = db.execute(text(f"SELECT {id_col} AS bid, {name_col} AS bname FROM public.booths WHERE {clause}"), params).all()
                for r in rows:
                    if r.bname: booth_name_map[int(r.bid)] = r.bname
            
            # 2. Try composite match (ward_id * 10000 + booth_no)
            remaining_booths = booth_ids_all - set(booth_name_map.keys())
            if remaining_booths and ward_ref and booth_no_col:
                potential_composites = [m for m in remaining_booths if m > 1000]
                if potential_composites:
                    potential_ward_ids = list(set([m // 10000 for m in potential_composites]))
                    if potential_ward_ids:
                        clause, params = _build_in_clause(ward_ref, [str(m) for m in potential_ward_ids], "wc_b")
                        rows = db.execute(text(f"SELECT {ward_ref} AS wid, {booth_no_col} AS bno, {name_col} AS bname FROM public.booths WHERE {clause}"), params).all()
                        for r in rows:
                            if r.wid is not None and r.bno is not None:
                                cid = int(r.wid) * 10000 + int(r.bno)
                                if cid in remaining_booths:
                                    prefix = f"{r.bno} - " if r.bno and r.bname and not str(r.bname).startswith(str(r.bno)) else ""
                                    booth_name_map[cid] = prefix + (r.bname or f"Booth {r.bno}")

            # 3. Try fallback booth_no match for any still remaining
            remaining_booths = booth_ids_all - set(booth_name_map.keys())
            if remaining_booths and booth_no_col:
                clause, params = _build_in_clause(booth_no_col, [str(m) for m in remaining_booths], "bno_f")
                rows = db.execute(text(f"SELECT {booth_no_col} AS bno, {name_col} AS bname FROM public.booths WHERE {clause}"), params).all()
                for r in rows:
                    if r.bname: booth_name_map[int(r.bno)] = r.bname

        # 4. Final fallback to data schema Booth table
        still_missing_booths = booth_ids_all - set(booth_name_map.keys())
        if still_missing_booths:
            booth_q = db.query(Booth.booth_id, Booth.polling_station_adr_en)
            if current.tenantId is not None and current.role != "SUPER_ADMIN":
                booth_q = booth_q.filter(Booth.tenant_id == current.tenantId)
            booth_rows = booth_q.filter(Booth.booth_id.in_(list(still_missing_booths))).all()
            for r in booth_rows:
                if r.polling_station_adr_en: booth_name_map[r.booth_id] = r.polling_station_adr_en

    content = []
    for u in users:
        item = to_user_details(u)
        u_ward_ids = parse_ids(getattr(u, "ward_ids", None))
        u_booth_ids = parse_ids(getattr(u, "booth_ids", None))
        item["wardNames"] = [ward_name_map.get(i, f"Ward {i}") for i in u_ward_ids]
        item["boothNames"] = [booth_name_map.get(i, f"Booth {i}") for i in u_booth_ids]
        content.append(item)
    return build_page(content, page, size, total, sortBy, direction)


def _enrichment_has_value(enrichment: VoterEnrichment, api_field: str) -> bool:
    column_name = VOTER_ENRICHMENT_FIELD_MAP.get(api_field, api_field)
    if not hasattr(enrichment, column_name):
        return False
    value = getattr(enrichment, column_name, None)
    if value is None:
        return False
    if isinstance(value, str):
        if api_field in VOTER_ENRICHMENT_JSON_FIELDS:
            try:
                parsed = json.loads(value)
                return bool(parsed)
            except Exception:
                return bool(value.strip())
        return bool(value.strip())
    return True


@app.get(f"{CONTEXT_PATH}/api/volunteers/analysis")
def volunteer_analysis(
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD")),
    wardId: Optional[int] = None,
    mode: Optional[str] = None,
    assemblyCode: Optional[str] = None,
):
    scope = _resolve_access_scope_ids(db, current)

    allowed_ward_codes: set[str] = set()
    allowed_booth_nos: set[str] = set()

    if scope:
        allowed_ward_ids = sorted(scope.get("allowed_ward_ids") or [])
        allowed_booth_ids = sorted(scope.get("allowed_booth_ids") or [])

        if allowed_ward_ids:
            ward_cols = _get_table_columns(db, "public", "wards")
            ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
            ward_code_col = "ward_code" if "ward_code" in ward_cols else None
            if ward_id_col and ward_code_col:
                clause, params = _build_in_clause(ward_id_col, allowed_ward_ids, "scope_ward_id")
                rows = db.execute(
                    text(
                        f"""
                        SELECT {ward_code_col} AS ward_code
                        FROM public.wards
                        WHERE {clause}
                        """
                    ),
                    params,
                ).all()
                allowed_ward_codes.update([str(r.ward_code) for r in rows if r.ward_code is not None])

        if allowed_booth_ids:
            booth_cols = _get_table_columns(db, "public", "booths")
            booth_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
            booth_no_col = "booth_no" if "booth_no" in booth_cols else None
            if booth_id_col and booth_no_col:
                clause, params = _build_in_clause(booth_id_col, allowed_booth_ids, "scope_booth_id")
                rows = db.execute(
                    text(
                        f"""
                        SELECT {booth_no_col} AS booth_no
                        FROM public.booths
                        WHERE {clause}
                        """
                    ),
                    params,
                ).all()
                allowed_booth_nos.update([str(r.booth_no) for r in rows if r.booth_no is not None])

    q = db.query(VoterEnrichment).filter(VoterEnrichment.updated_by.isnot(None))
    # Filter out admin entries from analysis
    q = q.filter(not_(VoterEnrichment.updated_by_name.in_(["admin@iswot.in", "admin@iswot.io"])))
    if scope:
        filters = []
        if allowed_ward_codes:
            filters.append(VoterEnrichment.ward_code.in_(allowed_ward_codes))
        if allowed_booth_nos:
            filters.append(VoterEnrichment.booth_no.in_(allowed_booth_nos))
        if filters:
            q = q.filter(or_(*filters))
        else:
            q = q.filter(text("1=0"))

    if wardId:
        ward_cols = _get_table_columns(db, "public", "wards")
        ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
        ward_code_col = "ward_code" if "ward_code" in ward_cols else None
        if ward_id_col and ward_code_col:
            row = db.execute(
                text(
                    f"""
                    SELECT {ward_code_col} AS ward_code
                    FROM public.wards
                    WHERE {ward_id_col} = :ward_id
                    LIMIT 1
                    """
                ),
                {"ward_id": wardId},
            ).first()
            if row and row.ward_code is not None:
                q = q.filter(VoterEnrichment.ward_code == str(row.ward_code))
            else:
                q = q.filter(text("1=0"))

    q = _apply_enrichment_assembly_filter(q, db, assemblyCode)

    enrichments = q.all()
    if not enrichments:
        return api_success("Volunteer analysis fetched", [])

    def _label_from_key(key: str) -> str:
        return " ".join([part.capitalize() for part in key.replace("_", " ").split()])

    exclude_keys = {
        "firstMiddleNameEn",
        "lastNameEn",
        "firstMiddleNameLocal",
        "lastNameLocal",
        "relationType",
        "relationFirstMiddleNameEn",
        "relationLastNameEn",
        "relationFirstMiddleNameLocal",
        "relationLastNameLocal",
        "houseNoEn",
        "houseNoLocal",
        "gender",
        "age",
    }
    enrichment_keys = [key for key in VOTER_ENRICHMENT_FIELD_MAP.keys() if key not in exclude_keys]
    extra_keys = ["ward_code", "booth_no", "updated_fields"]
    analysis_fields = [{"key": key, "label": _label_from_key(key)} for key in enrichment_keys + extra_keys]

    mode_key = (mode or "agent").lower()

    counters: Dict[int, Dict[str, Any]] = {}
    for enrichment in enrichments:
        user_id = enrichment.updated_by
        if user_id is None:
            continue
        bucket = counters.setdefault(
            int(user_id),
            {
                "userId": int(user_id),
                "counts": {item["key"]: 0 for item in analysis_fields},
                "total": 0,
                "agentName": None,
                "phone": None,
                "lastUpdatedAt": None,
            },
        )
        if not bucket.get("agentName") and enrichment.updated_by_name:
            bucket["agentName"] = enrichment.updated_by_name
        if not bucket.get("phone") and enrichment.updated_by_phone:
            bucket["phone"] = enrichment.updated_by_phone
        bucket["total"] += 1
        if enrichment.updated_at:
            existing = bucket.get("lastUpdatedAt")
            if not existing or enrichment.updated_at > existing:
                bucket["lastUpdatedAt"] = enrichment.updated_at
        for item in analysis_fields:
            if _enrichment_has_value(enrichment, item["key"]):
                bucket["counts"][item["key"]] += 1

    if mode_key == "agent":
        user_rows = db.query(User).filter(User.id.in_(list(counters.keys()))).all()
        user_map = {u.id: u for u in user_rows}

        results = []
        for user_id, bucket in counters.items():
            user = user_map.get(user_id)
            results.append(
                {
                    "userId": user_id,
                    "agentName": user.first_name if user else (bucket.get("agentName") or f"User {user_id}"),
                    "phone": user.phone if user else (bucket.get("phone") or ""),
                    "total": bucket["total"],
                    "counts": bucket["counts"],
                    "lastUpdatedAt": bucket.get("lastUpdatedAt").isoformat() if bucket.get("lastUpdatedAt") else None,
                }
            )
        results.sort(key=lambda item: item.get("agentName") or "")
        return api_success("Volunteer analysis fetched", {"fields": analysis_fields, "rows": results, "mode": mode_key})

    ward_name_map: Dict[str, str] = {}
    if mode_key == "ward":
        ward_cols = _get_table_columns(db, "public", "wards")
        ward_code_col = "ward_code" if "ward_code" in ward_cols else None
        ward_name_col = (
            "ward_name_en"
            if "ward_name_en" in ward_cols
            else ("name_en" if "name_en" in ward_cols else ("ward_name_local" if "ward_name_local" in ward_cols else None))
        )
        if ward_code_col and ward_name_col:
            t_clause, t_params = _build_public_tenant_filter(current)
            ward_rows = db.execute(
                text(
                    f"""
                    SELECT {ward_code_col} AS ward_code, {ward_name_col} AS ward_name
                    FROM public.wards
                    WHERE 1=1 {t_clause}
                    """
                ),
                t_params
            ).all()
            ward_name_map = {
                str(r.ward_code): str(r.ward_name)
                for r in ward_rows
                if r.ward_code is not None and r.ward_name is not None
            }

    group_buckets: Dict[str, Dict[str, Any]] = {}
    for enrichment in enrichments:
        if mode_key == "date":
            if not enrichment.updated_at:
                continue
            group_key = enrichment.updated_at.date().isoformat()
            group_label = enrichment.updated_at.date().isoformat()
        elif mode_key == "ward":
            group_key = str(enrichment.ward_code or "")
            if not group_key:
                continue
            group_label = ward_name_map.get(group_key) or f"Ward {group_key}"
        elif mode_key == "booth":
            group_key = str(enrichment.booth_no or "")
            if not group_key:
                continue
            group_label = f"Booth {group_key}"
        else:
            group_key = "all"
            group_label = "All"

        bucket = group_buckets.setdefault(
            group_key,
            {
                "groupKey": group_key,
                "label": group_label,
                "counts": {item["key"]: 0 for item in analysis_fields},
                "total": 0,
                "agents": set(),
                "booths": set(),
                "lastUpdatedAt": None,
            },
        )
        bucket["total"] += 1
        if enrichment.updated_by:
            bucket["agents"].add(enrichment.updated_by)
        if enrichment.booth_no:
            bucket["booths"].add(str(enrichment.booth_no))
        if enrichment.updated_at:
            existing = bucket.get("lastUpdatedAt")
            if not existing or enrichment.updated_at > existing:
                bucket["lastUpdatedAt"] = enrichment.updated_at
        for item in analysis_fields:
            if _enrichment_has_value(enrichment, item["key"]):
                bucket["counts"][item["key"]] += 1

    grouped_rows = []
    for bucket in group_buckets.values():
        grouped_rows.append(
            {
                "groupKey": bucket["groupKey"],
                "label": bucket["label"],
                "agentsWorked": len(bucket["agents"]),
                "boothsCovered": len(bucket["booths"]),
                "total": bucket["total"],
                "counts": bucket["counts"],
                "lastUpdatedAt": bucket.get("lastUpdatedAt").isoformat() if bucket.get("lastUpdatedAt") else None,
            }
        )

    if mode_key == "date":
        grouped_rows.sort(key=lambda item: item.get("groupKey") or "")
    else:
        grouped_rows.sort(key=lambda item: item.get("label") or "")
    return api_success("Volunteer analysis fetched", {"fields": analysis_fields, "rows": grouped_rows, "mode": mode_key})


@app.get(f"{CONTEXT_PATH}/api/volunteers/analysis/enrichment")
def volunteer_analysis_enrichment(
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "WARD", "ASSEMBLY")),
    wardId: Optional[int] = None,
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    page: Optional[int] = None,
    size: Optional[int] = None,
    assemblyCode: Optional[str] = None,
):
    exclude_keys = {
        "firstMiddleNameEn",
        "lastNameEn",
        "firstMiddleNameLocal",
        "lastNameLocal",
        "relationType",
        "relationFirstMiddleNameEn",
        "relationLastNameEn",
        "relationFirstMiddleNameLocal",
        "relationLastNameLocal",
        "houseNoEn",
        "houseNoLocal",
        "addressEn",
        "addressLocal",
        "team",
    }
    ordered_keys = [
        "serialNumber",
        "wardName",
        "name",
        "epicNo",
        "boothNo",
        "voterSerialNo",
        "lastUpdatedAt",
    ]
    ward_name_map: Dict[str, str] = {}
    ward_cols = _get_table_columns(db, "public", "wards")
    ward_code_col = "ward_code" if "ward_code" in ward_cols else None
    ward_name_col = (
        "ward_name_en"
        if "ward_name_en" in ward_cols
        else ("name_en" if "name_en" in ward_cols else ("ward_name_local" if "ward_name_local" in ward_cols else None))
    )
    if ward_code_col and ward_name_col:
        t_clause, t_params = _build_public_tenant_filter(current)
        ward_rows = db.execute(
            text(
                f"""
                SELECT {ward_code_col} AS ward_code, {ward_name_col} AS ward_name
                FROM public.wards
                WHERE 1=1 {t_clause}
                """
            ),
            t_params
        ).all()
        ward_name_map = {
            str(r.ward_code): str(r.ward_name) for r in ward_rows if r.ward_code is not None and r.ward_name is not None
        }

    enrichments_q = db.query(VoterEnrichment).filter(not_(VoterEnrichment.updated_by_name.in_(["admin@iswot.in", "admin@iswot.io"])))

    # For non-super-admin users, scope to their allowed ward codes
    scope = _resolve_access_scope_ids(db, current)
    if scope:
        allowed_ward_ids = sorted(scope.get("allowed_ward_ids") or [])
        if allowed_ward_ids:
            ward_cols = _get_table_columns(db, "public", "wards")
            scope_ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
            scope_ward_code_col = "ward_code" if "ward_code" in ward_cols else None
            if scope_ward_id_col and scope_ward_code_col:
                clause, params = _build_in_clause(scope_ward_id_col, allowed_ward_ids, "scope_enrich_ward_id")
                t_clause, t_params = _build_public_tenant_filter(current)
                scope_ward_rows = db.execute(
                    text(f"SELECT {scope_ward_code_col} AS ward_code FROM public.wards WHERE {clause} {t_clause}"),
                    {**params, **t_params},
                ).all()
                allowed_ward_codes = [str(r.ward_code) for r in scope_ward_rows if r.ward_code is not None]
                if allowed_ward_codes:
                    enrichments_q = enrichments_q.filter(VoterEnrichment.ward_code.in_(allowed_ward_codes))
                else:
                    enrichments_q = enrichments_q.filter(text("1=0"))

    if wardId:
        ward_cols = _get_table_columns(db, "public", "wards")
        ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
        ward_code_col = "ward_code" if "ward_code" in ward_cols else None
        if ward_id_col and ward_code_col:
            row = db.execute(
                text(
                    f"""
                    SELECT {ward_code_col} AS ward_code
                    FROM public.wards
                    WHERE {ward_id_col} = :ward_id
                    LIMIT 1
                    """
                ),
                {"ward_id": wardId},
            ).first()
            if row and row.ward_code is not None:
                enrichments_q = enrichments_q.filter(VoterEnrichment.ward_code == str(row.ward_code))
            else:
                enrichments_q = enrichments_q.filter(text("1=0"))

    enrichments_q = _apply_enrichment_assembly_filter(enrichments_q, db, assemblyCode)

    def _parse_date(value: Optional[str], end_of_day: bool = False) -> Optional[datetime]:
        if not value:
            return None
        try:
            parsed = datetime.fromisoformat(value)
        except Exception:
            try:
                parsed = datetime.strptime(value, "%Y-%m-%d")
            except Exception:
                return None
        if end_of_day and parsed.hour == 0 and parsed.minute == 0 and parsed.second == 0 and len(value) <= 10:
            parsed = parsed + timedelta(hours=23, minutes=59, seconds=59)
        return parsed

    from_dt = _parse_date(updatedFrom)
    to_dt = _parse_date(updatedTo, end_of_day=True)
    if from_dt:
        enrichments_q = enrichments_q.filter(VoterEnrichment.updated_at >= from_dt)
    if to_dt:
        enrichments_q = enrichments_q.filter(VoterEnrichment.updated_at <= to_dt)
    
    enrichments_q = enrichments_q.order_by(desc(VoterEnrichment.updated_at))
    if page is not None and size is not None:
        enrichments = enrichments_q.offset(page * size).limit(size).all()
        start_idx = (page * size) + 1
    else:
        enrichments = enrichments_q.all()
        start_idx = 1
        
    if not enrichments:
        return api_success("Volunteer enrichment fetched", [])

    def _normalize_epic(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        normalized = str(value).strip().upper()
        return normalized or None

    epics = [_normalize_epic(e.epic) for e in enrichments if _normalize_epic(e.epic)]
    voter_rows = (
        db.query(Voter)
        .filter(func.upper(func.trim(Voter.epic_no)).in_(epics))
        .all()
        if epics
        else []
    )
    voter_map = {(_normalize_epic(v.epic_no)): v for v in voter_rows if _normalize_epic(v.epic_no)}

    public_voter_map: Dict[str, Dict[str, Any]] = {}
    if epics:
        voter_cols = _get_table_columns(db, "public", "voters")
        epic_col = "epic" if "epic" in voter_cols else ("epic_no" if "epic_no" in voter_cols else None)
        name_col = "name_en" if "name_en" in voter_cols else ("name" if "name" in voter_cols else None)
        sl_col = "sl" if "sl" in voter_cols else ("sr_no" if "sr_no" in voter_cols else None)
        if epic_col and (name_col or sl_col):
            clause, params = _build_in_clause(f"UPPER(TRIM({epic_col}))", epics, "public_epic")
            select_cols = [
                f"{epic_col} AS epic",
                f"{name_col} AS name_en" if name_col else "NULL AS name_en",
                f"{sl_col} AS sl" if sl_col else "NULL AS sl",
            ]
            t_clause, t_params = _build_public_tenant_filter(current)
            public_rows = db.execute(
                text(
                    f"""
                    SELECT {', '.join(select_cols)}
                    FROM public.voters
                    WHERE {clause} {t_clause}
                    """
                ),
                {**params, **t_params},
            ).all()
            public_voter_map = {
                _normalize_epic(r.epic): {"name_en": r.name_en, "sl": r.sl}
                for r in public_rows
                if _normalize_epic(r.epic)
            }

    rows: List[Dict[str, Any]] = []
    for idx, enrichment in enumerate(enrichments, start=start_idx):
        payload = _build_voter_enrichment_payload(enrichment)
        ward_name = ward_name_map.get(str(enrichment.ward_code), None) if enrichment.ward_code is not None else None
        epic_key = _normalize_epic(enrichment.epic) or _normalize_epic(payload.get("epicNo"))
        voter = voter_map.get(epic_key)
        public_voter = public_voter_map.get(epic_key) if epic_key else None
        name_parts = [
            payload.get("firstMiddleNameEn") or (voter.first_middle_name_en if voter else None),
            payload.get("lastNameEn") or (voter.last_name_en if voter else None),
        ]
        full_name = " ".join([part for part in name_parts if part]).strip() or None
        if not full_name and public_voter:
            full_name = normalize_optional_text(public_voter.get("name_en"))
        ordered: Dict[str, Any] = {}
        ordered["serialNumber"] = idx
        ordered["wardName"] = ward_name
        ordered["name"] = full_name
        ordered["epicNo"] = payload.get("epicNo")
        ordered["boothNo"] = payload.get("boothNo")
        # Prefer sr_no from Voter table; fall back to enrichment newSerialNo
        ordered["voterSerialNo"] = (
            voter.sr_no
            if voter and voter.sr_no is not None
            else (public_voter.get("sl") if public_voter else None)
        ) or payload.get("newSerialNo")
        ordered["lastUpdatedAt"] = enrichment.updated_at.isoformat() if enrichment.updated_at else None

        remaining_keys = [
            key
            for key in (
                ["wardCode", "boothNo", "updatedFields", "updatedByName", "updatedByPhone"]
                + list(VOTER_ENRICHMENT_FIELD_MAP.keys())
            )
            if key not in ordered_keys and key not in exclude_keys and key != "updatedFields"
        ]
        for key in remaining_keys:
            if key in payload:
                ordered[key] = payload.get(key)
        rows.append(ordered)

    return api_success("Volunteer enrichment fetched", rows)


@app.get(f"{CONTEXT_PATH}/api/volunteers/analysis/locations")
def volunteer_analysis_locations(
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD")),
    wardId: Optional[int] = None,
    assemblyCode: Optional[str] = None,
):
    scope = _resolve_access_scope_ids(db, current)

    allowed_ward_codes: set[str] = set()
    allowed_booth_nos: set[str] = set()

    if scope:
        allowed_ward_ids = sorted(scope.get("allowed_ward_ids") or [])
        allowed_booth_ids = sorted(scope.get("allowed_booth_ids") or [])

        if allowed_ward_ids:
            ward_cols = _get_table_columns(db, "public", "wards")
            ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
            ward_code_col = "ward_code" if "ward_code" in ward_cols else None
            if ward_id_col and ward_code_col:
                clause, params = _build_in_clause(ward_id_col, allowed_ward_ids, "scope_ward_id")
                rows = db.execute(
                    text(
                        f"""
                        SELECT {ward_code_col} AS ward_code
                        FROM public.wards
                        WHERE {clause}
                        """
                    ),
                    params,
                ).all()
                allowed_ward_codes.update([str(r.ward_code) for r in rows if r.ward_code is not None])

        if allowed_booth_ids:
            booth_cols = _get_table_columns(db, "public", "booths")
            booth_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
            booth_no_col = "booth_no" if "booth_no" in booth_cols else None
            if booth_id_col and booth_no_col:
                clause, params = _build_in_clause(booth_id_col, allowed_booth_ids, "scope_booth_id")
                rows = db.execute(
                    text(
                        f"""
                        SELECT {booth_no_col} AS booth_no
                        FROM public.booths
                        WHERE {clause}
                        """
                    ),
                    params,
                ).all()
                allowed_booth_nos.update([str(r.booth_no) for r in rows if r.booth_no is not None])

    q = db.query(VoterEnrichment).filter(VoterEnrichment.latitude.isnot(None), VoterEnrichment.longitude.isnot(None))
    if scope:
        filters = []
        if allowed_ward_codes:
            filters.append(VoterEnrichment.ward_code.in_(allowed_ward_codes))
        if allowed_booth_nos:
            filters.append(VoterEnrichment.booth_no.in_(allowed_booth_nos))
        if filters:
            q = q.filter(or_(*filters))
        else:
            q = q.filter(text("1=0"))

    if wardId:
        ward_cols = _get_table_columns(db, "public", "wards")
        ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
        ward_code_col = "ward_code" if "ward_code" in ward_cols else None
        if ward_id_col and ward_code_col:
            row = db.execute(
                text(
                    f"""
                    SELECT {ward_code_col} AS ward_code
                    FROM public.wards
                    WHERE {ward_id_col} = :ward_id
                    LIMIT 1
                    """
                ),
                {"ward_id": wardId},
            ).first()
            if row and row.ward_code is not None:
                q = q.filter(VoterEnrichment.ward_code == str(row.ward_code))
            else:
                q = q.filter(text("1=0"))

    q = _apply_enrichment_assembly_filter(q, db, assemblyCode)

    enrichments = q.all()
    if not enrichments:
        return api_success("Volunteer locations fetched", [])

    epics = [str(e.epic).strip().upper() for e in enrichments if e.epic]
    # 1. Fetch from public.voters (Master list)
    public_voter_map = {}
    if epics:
        clause, params = _build_in_clause("epic", epics, "pub_epic")
        rows_pub = db.execute(
            text(f"SELECT epic, name_en, rel_eng, gender, mobile FROM public.voters WHERE {clause}"),
            params
        ).all()
        public_voter_map = {str(r.epic).strip().upper(): r for r in rows_pub if r.epic}

    # 2. Fetch from data.voters (Tenant specific)
    tenant_voter_map: Dict[str, Voter] = {}
    if epics:
        voter_rows = db.query(Voter).filter(Voter.epic_no.in_(epics)).all()
        tenant_voter_map = {
            str(v.epic_no).strip().upper(): v for v in voter_rows if v.epic_no is not None
        }

    rows: List[Dict[str, Any]] = []
    gender_counts = {"male": 0, "female": 0, "other": 0}
    for enrichment in enrichments:
        if enrichment.latitude is None or enrichment.longitude is None:
            continue
        epic_key = str(enrichment.epic).strip().upper() if enrichment.epic else None
        v_ten = tenant_voter_map.get(epic_key)
        v_pub = public_voter_map.get(epic_key)
        
        gender = enrichment.gender or (v_ten.gender if v_ten else None) or (v_pub.gender if v_pub else None)
        gender_upper = str(gender or "").upper()
        if gender_upper.startswith("M"):
            gender_counts["male"] += 1
        elif gender_upper.startswith("F"):
            gender_counts["female"] += 1
        else:
            gender_counts["other"] += 1

        name_parts = [
            enrichment.first_middle_name_en or (v_ten.first_middle_name_en if v_ten else None),
            enrichment.last_name_en or (v_ten.last_name_en if v_ten else None),
        ]
        full_name = " ".join([part for part in name_parts if part]).strip()
        if not full_name and v_pub:
            full_name = v_pub.name_en
        if not full_name:
            full_name = "Unknown"

        rel_parts = [
            enrichment.relation_first_middle_name_en or (v_ten.relation_first_middle_name_en if v_ten else None),
            enrichment.relation_last_name_en or (v_ten.relation_last_name_en if v_ten else None),
        ]
        relation_name = " ".join([part for part in rel_parts if part]).strip()
        if not relation_name and v_pub:
            relation_name = v_pub.rel_eng

        rows.append(
            {
                "latitude": enrichment.latitude,
                "longitude": enrichment.longitude,
                "gender": gender,
                "name": full_name,
                "epic": enrichment.epic,
                "relationName": relation_name,
                "mobile": enrichment.mobile or (v_ten.mobile if v_ten else None) or (v_pub.mobile if v_pub else None),
            }
        )
    print(
        f"[volunteer-map] points={len(rows)} male={gender_counts['male']} female={gender_counts['female']} other={gender_counts['other']}"
    )
    return api_success("Volunteer locations fetched", rows)


@app.put(f"{CONTEXT_PATH}/api/user/block")
def block_user(payload: UserBlockRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD"))):
    target_first_name = resolve_user_first_name(payload.firstName, payload.userEmail)
    if not target_first_name:
        raise ValueError("firstName or userEmail is required")

    q = db.query(VolunteerUser).filter(VolunteerUser.first_name == target_first_name)
    if current.tenantId:
        q = q.filter(VolunteerUser.tenant_id == current.tenantId)
    user = q.first()
    if not user:
        return api_error("User not found", {"details": f"User not found with FirstName: '{target_first_name}'"})

    user.blocked = payload.block
    db.commit()
    action = "blocked" if payload.block else "unblocked"
    return api_success(f"User {action} successfully", {"firstName": target_first_name})


@app.get(f"{CONTEXT_PATH}/api/user")
def list_users(
    role: Optional[str] = None,
    page: int = 0,
    size: int = 10,
    search: Optional[str] = None,
    blocked: Optional[str] = None,
    deleted: Optional[str] = None,
    assignmentType: Optional[str] = None,
    sortBy: str = "firstName",
    direction: str = "asc",
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN")),
):
    # For non-super-admins, filter user list by their tenant
    q = db.query(VolunteerUser).filter(VolunteerUser.role == "USER")
    if current.tenantId is not None and current.role != "SUPER_ADMIN":
        q = q.filter(VolunteerUser.tenant_id == current.tenantId)

    blocked_filter = parse_optional_bool(blocked)
    deleted_filter = parse_optional_bool(deleted)
    assignment_type_filter = normalize_optional_text(assignmentType)

    if blocked_filter is not None:
        q = q.filter(VolunteerUser.blocked == blocked_filter)
    if deleted_filter is not None:
        q = q.filter(VolunteerUser.deleted == deleted_filter)
    if assignment_type_filter is not None:
        q = q.filter(VolunteerUser.assignment_type == assignment_type_filter)
    if search and search.strip():
        s = f"%{search.lower()}%"
        q = q.filter(or_(func.lower(VolunteerUser.first_name).like(s), VolunteerUser.phone.like(f"%{search}%")))

    sort_map = {
        "id": VolunteerUser.id,
        "firstName": VolunteerUser.first_name,
        "phone": VolunteerUser.phone,
        "role": VolunteerUser.role,
        "assignmentType": VolunteerUser.assignment_type,
        "assignmentId": VolunteerUser.assignment_id,
        "blocked": VolunteerUser.blocked,
        "deleted": VolunteerUser.deleted,
    }
    sort_col = sort_map.get(sortBy, VolunteerUser.first_name)
    q = q.order_by(desc(sort_col) if direction.lower() == "desc" else asc(sort_col))

    total = q.count()
    users = q.offset(page * size).limit(size).all()
    content = [to_user_details(u) for u in users]
    return build_page(content, page, size, total, sortBy, direction)


@app.put(f"{CONTEXT_PATH}/api/user/delete")
def delete_user(payload: UserDeleteRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD"))):
    target_first_name = resolve_user_first_name(payload.firstName, payload.userEmail)
    if not target_first_name:
        raise ValueError("firstName or userEmail is required")

    q = db.query(VolunteerUser).filter(VolunteerUser.first_name == target_first_name)
    if current.tenantId is not None and current.role != "SUPER_ADMIN":
        q = q.filter(VolunteerUser.tenant_id == current.tenantId)
    user = q.first()
    if not user:
        return api_error("User not found", {"details": f"User not found with FirstName: '{target_first_name}'"})

    user.deleted = payload.delete
    db.commit()
    action = "deleted" if payload.delete else "restored"
    return api_success(f"User {action} successfully", {"email": payload.model_dump()})


@app.put(f"{CONTEXT_PATH}/api/user/block/bulk")
def bulk_block(payload: UserBulkActionRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD"))):
    usernames = resolve_bulk_usernames(payload)
    q = db.query(VolunteerUser).filter(VolunteerUser.first_name.in_(usernames))
    if current.tenantId is not None and current.role != "SUPER_ADMIN":
        q = q.filter(VolunteerUser.tenant_id == current.tenantId)
    users = q.all()
    if len(users) != len(usernames):
        return api_error("Some users not found", {"details": "Some users not found"})

    for user in users:
        user.blocked = payload.action
    db.commit()
    action = "blocked" if payload.action else "unblocked"
    return api_success(f"Users {action} successfully", {"emails": usernames})


@app.put(f"{CONTEXT_PATH}/api/user/delete/bulk")
def bulk_delete(payload: UserBulkActionRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("ADMIN", "SUPER_ADMIN", "ASSEMBLY", "WARD"))):
    usernames = resolve_bulk_usernames(payload)
    q = db.query(VolunteerUser).filter(VolunteerUser.first_name.in_(usernames))
    if current.tenantId is not None and current.role != "SUPER_ADMIN":
        q = q.filter(VolunteerUser.tenant_id == current.tenantId)
    users = q.all()
    if len(users) != len(usernames):
        return api_error("Some users not found", {"details": "Some users not found"})

    for user in users:
        user.deleted = payload.action
    db.commit()
    action = "deleted" if payload.action else "restored"
    return api_success(f"Users {action} successfully", {"emails": usernames})


@app.get(f"{CONTEXT_PATH}/api/user/profile")
def get_profile(db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    user = db.query(User).filter(User.first_name == current.firstName, User.phone == current.phone).first()
    volunteer = None
    if not user:
        volunteer = db.query(VolunteerUser).filter(VolunteerUser.first_name == current.firstName, VolunteerUser.phone == current.phone).first()
        if not volunteer:
            raise ResourceNotFoundException("User", "username", current.firstName)

    target = user or volunteer
    presigned = ""
    if target.profile_pic_url:
        key = s3_extract_key(target.profile_pic_url)
        presigned = s3_presigned_url(key, 15, fallback_url=target.profile_pic_url)

    return {
        "firstName": target.first_name,
        "lastName": "",
        "userName": target.first_name,
        "phone": target.phone,
        "profilePicUrl": presigned,
        "tenantId": _resolve_tenant_id_for_entity(target),
        "role": target.role,
    }


@app.put(f"{CONTEXT_PATH}/api/user/profile")
def update_profile(payload: UserProfileDto, db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    user = db.query(User).filter(User.first_name == current.firstName, User.phone == current.phone).first()
    volunteer = None
    if not user:
        volunteer = db.query(VolunteerUser).filter(VolunteerUser.first_name == current.firstName, VolunteerUser.phone == current.phone).first()
        if not volunteer:
            raise ResourceNotFoundException("User", "username", current.firstName)

    target = user or volunteer
    target.first_name = payload.firstName
    target.phone = payload.phone
    db.commit()
    db.refresh(target)
    return get_profile(db, JwtUserDetails(phone=target.phone, firstName=target.first_name, role=target.role, tenantId=current.tenantId, assignmentType=current.assignmentType, assignmentId=current.assignmentId))


@app.post(f"{CONTEXT_PATH}/api/user/profile/upload")
def upload_profile(file: UploadFile = File(...), db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    user = db.query(User).filter(User.first_name == current.firstName, User.phone == current.phone).first()
    volunteer = None
    if not user:
        volunteer = db.query(VolunteerUser).filter(VolunteerUser.first_name == current.firstName, VolunteerUser.phone == current.phone).first()
        if not volunteer:
            raise ResourceNotFoundException("User", "username", current.firstName)

    ext = Path(file.filename or "").suffix
    target = user or volunteer
    tenant_segment = _resolve_tenant_id_for_entity(target) or "global"
    key = f"{PROFILE_UPLOAD_DIR}/{tenant_segment}/{target.first_name}/{uuid.uuid4()}{ext}"
    raw = file.file.read()
    s3_url = s3_upload_bytes(raw, file.content_type or "application/octet-stream", key)
    target.profile_pic_url = s3_url
    db.commit()

    presigned = s3_presigned_url(key, 15, fallback_url=s3_url)
    return {
        "firstName": target.first_name,
        "lastName": "",
        "userName": target.first_name,
        "phone": target.phone,
        "profilePicUrl": presigned,
        "tenantId": _resolve_tenant_id_for_entity(target),
        "role": target.role,
    }


@app.post(f"{CONTEXT_PATH}/api/tenant")
def create_tenant(payload: TenantDtoIn, db: Session = Depends(get_db), _: JwtUserDetails = Depends(require_roles("SUPER_ADMIN"))):
    existing = db.query(Tenant).filter(Tenant.contact_email == payload.contactEmail).first()
    if existing:
        raise ResourceAlreadyExistsException("Tenant", "contactEmail", payload.contactEmail)

    tenant = Tenant(
        name=payload.name,
        description=payload.description,
        contact_email=payload.contactEmail,
        contact_phone=payload.contactPhone,
        active=True,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        tenant_id=f"iswot-{int(datetime.utcnow().timestamp() * 1000) % 100_000_000:08d}",
    )
    db.add(tenant)
    db.flush()

    admin = User(
        role="ADMIN",
        tenant_ref=tenant.id,
        first_name="Tenant",
        phone=payload.contactPhone or "0000000000",
        blocked=False,
        deleted=False,
    )
    db.add(admin)
    db.commit()
    db.refresh(tenant)
    return to_tenant_dto(tenant)


@app.get(f"{CONTEXT_PATH}/api/tenant/{{tenantId}}")
def get_tenant(tenantId: str, db: Session = Depends(get_db), _: JwtUserDetails = Depends(require_roles("SUPER_ADMIN"))):
    tenant = db.query(Tenant).filter(Tenant.tenant_id == tenantId).first()
    if not tenant:
        raise ValueError(f"Tenant not found: {tenantId}")
    return to_tenant_dto(tenant)


@app.put(f"{CONTEXT_PATH}/api/tenant/{{tenantId}}")
def update_tenant(tenantId: str, payload: TenantDtoIn, db: Session = Depends(get_db), _: JwtUserDetails = Depends(require_roles("SUPER_ADMIN"))):
    tenant = db.query(Tenant).filter(Tenant.tenant_id == tenantId).first()
    if not tenant:
        raise ValueError(f"Tenant not found: {tenantId}")

    tenant.name = payload.name
    tenant.description = payload.description
    tenant.contact_phone = payload.contactPhone
    tenant.active = payload.active if payload.active is not None else tenant.active
    tenant.updated_at = datetime.utcnow()
    db.commit()
    return to_tenant_dto(tenant)


@app.get(f"{CONTEXT_PATH}/api/tenant")
def list_tenants(page: int = 0, size: int = 10, db: Session = Depends(get_db), _: JwtUserDetails = Depends(require_roles("SUPER_ADMIN"))):
    q = db.query(Tenant).order_by(Tenant.id)
    total = q.count()
    tenants = q.offset(page * size).limit(size).all()
    return build_page([to_tenant_dto(t) for t in tenants], page, size, total)


@app.get(f"{CONTEXT_PATH}/api/assignments")
def get_assignments(type: str = Query(...), db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER"))):
    scope = _resolve_access_scope_ids(db, current)
    if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
        return []
    t = type.upper()
    if t == "ASSEMBLY":
        q = db.query(Assembly)
        if current.tenantId is not None and current.role != "SUPER_ADMIN":
            q = q.filter(Assembly.tenant_id == current.tenantId)
        if scope and scope.get("allowed_assembly_ids"):
            q = q.filter(Assembly.assembly_id.in_(scope.get("allowed_assembly_ids")))
        rows = q.order_by(Assembly.assembly_name_en.asc()).all()
        return [{"id": r.assembly_id, "name": r.assembly_name_en} for r in rows]
    if t == "WARD":
        q = db.query(Ward)
        if current.tenantId is not None and current.role != "SUPER_ADMIN":
            q = q.filter(Ward.tenant_id == current.tenantId)
        if scope and scope.get("allowed_ward_ids"):
            q = q.filter(Ward.ward_id.in_(scope.get("allowed_ward_ids")))
        rows = q.order_by(Ward.ward_name_en.asc()).all()
        return [{"id": r.ward_id, "name": r.ward_name_en} for r in rows]
    if t == "BOOTH":
        q = db.query(Booth)
        if current.tenantId is not None and current.role != "SUPER_ADMIN":
            q = q.filter(Booth.tenant_id == current.tenantId)
        if scope and scope.get("allowed_booth_ids"):
            q = q.filter(Booth.booth_id.in_(scope.get("allowed_booth_ids")))
        rows = q.order_by(Booth.polling_station_adr_en.asc()).all()
        return [{"id": r.booth_id, "name": r.polling_station_adr_en} for r in rows]
    raise ValueError("Invalid assignment type")


@app.get(f"{CONTEXT_PATH}/api/booth")
def get_booths(db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER"))):
    scope = _resolve_access_scope_ids(db, current)
    if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
        return api_success("Booths fetched successfully", [])
    booth_cols = _get_table_columns(db, "public", "booths")
    booth_pk_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else "id")
    booth_ward_id_col = "ward_id" if "ward_id" in booth_cols else None

    where_parts: List[str] = []
    params: Dict[str, Any] = {}
    if scope:
        allowed_booth_ids = sorted([v for v in (scope.get("allowed_booth_ids") or []) if v is not None])
        allowed_ward_ids = sorted([v for v in (scope.get("allowed_ward_ids") or []) if v is not None])
        if allowed_booth_ids:
            clause, clause_params = _build_in_clause(f"b.{booth_pk_col}", allowed_booth_ids, "scope_booth")
            where_parts.append(clause)
            params.update(clause_params)
        elif allowed_ward_ids and booth_ward_id_col:
            clause, clause_params = _build_in_clause(f"b.{booth_ward_id_col}", allowed_ward_ids, "scope_ward")
            where_parts.append(clause)
            params.update(clause_params)

    where_clause = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
    rows = db.execute(
        text(
            f"""
            SELECT
                COALESCE(b.booth_no, b.{booth_pk_col}) AS id,
                b.booth_no AS booth_id,
                COALESCE(NULLIF(b.booth_add_en, ''), 'Booth ' || COALESCE(b.booth_no::text, b.{booth_pk_col}::text)) AS name_en
            FROM public.booths b
            {where_clause}
            ORDER BY
                b.booth_no ASC NULLS LAST,
                b.{booth_pk_col} ASC
            """
        ),
        params,
    ).all()
    dto = [{"id": int(r.id), "boothId": int(r.booth_id), "nameEn": r.name_en} for r in rows]
    return api_success("Booths fetched successfully", dto)


@app.get(f"{CONTEXT_PATH}/api/booths")
def get_booths_plural(
    assemblyCode: Optional[str] = None,
    wardId: Optional[int] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER")),
):
    assemblyCode = _normalize_assembly_code(assemblyCode)
    scope = _resolve_access_scope_ids(db, current)
    if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
        return []
    q = (
        db.query(Booth)
        .join(Ward, Booth.ward_id == Ward.ward_id)
        .join(Assembly, Ward.assembly_id == Assembly.assembly_id)
    )
    if wardId:
        q = q.filter(Booth.ward_id == wardId)
    if current.tenantId is not None and current.role != "SUPER_ADMIN":
        q = q.filter(Booth.tenant_id == current.tenantId)
    if scope:
        allowed_booth_ids = {v for v in (scope.get("allowed_booth_ids") or set()) if v is not None}
        allowed_ward_ids = {v for v in (scope.get("allowed_ward_ids") or set()) if v is not None}
        if allowed_booth_ids:
            q = q.filter(Booth.booth_id.in_(allowed_booth_ids))
        elif allowed_ward_ids:
            q = q.filter(Booth.ward_id.in_(allowed_ward_ids))
    if assemblyCode:
        q = q.filter(Assembly.assembly_code == assemblyCode)
    if wardId:
        if scope and (scope.get("allowed_ward_ids") or set()) and wardId not in (scope.get("allowed_ward_ids") or set()):
            return []
        q = q.filter(Ward.ward_id == wardId)

    booths = q.order_by(Booth.booth_no.asc()).all()
    # For super admins, we also want to optionally see the public list if data is sparse
    if booths and (current.role != "SUPER_ADMIN" or not wardId):
        return [
            {
                "boothId": b.booth_id,
                "boothNo": b.booth_no,
                "boothNameEn": b.polling_station_adr_en,
                "pollingStationAdrEn": b.polling_station_adr_en,
                "wardId": b.ward_id,
                "id": b.booth_id,
            }
            for b in booths
        ]

    booth_cols = _get_table_columns(db, "public", "booths")
    pk_col = "booth_id" if "booth_id" in booth_cols else ("id" if "id" in booth_cols else None)
    no_col = "booth_no" if "booth_no" in booth_cols else None
    name_col = "polling_station_adr_en" if "polling_station_adr_en" in booth_cols else ("booth_add_en" if "booth_add_en" in booth_cols else ("name_en" if "name_en" in booth_cols else None))
    ward_ref = "ward_id" if "ward_id" in booth_cols else ("ward_no" if "ward_no" in booth_cols else None)
    if not (pk_col or no_col) or not name_col:
        return []
    select_pk = f"{pk_col} AS booth_pk" if pk_col else "NULL AS booth_pk"
    select_no = f"{no_col} AS booth_no" if no_col else "NULL AS booth_no"
    where = []
    params: Dict[str, Any] = {}
    if wardId and ward_ref:
        where.append(f"{ward_ref} = :ward_id")
        params["ward_id"] = wardId
    where_clause = f"WHERE {' AND '.join(where)}" if where else ""
    rows = db.execute(
        text(
            f"""
            SELECT {select_pk}, {select_no}, {name_col} AS name, {ward_ref if ward_ref else 'NULL'} AS ward_id
            FROM public.booths
            {where_clause}
            ORDER BY {name_col}
            """
        ),
        params,
    ).all()
    if scope:
        allowed_booth_ids = scope.get("allowed_booth_ids") or set()
        allowed_ward_ids = scope.get("allowed_ward_ids") or set()
        allowed_assembly_ids = scope.get("allowed_assembly_ids") or set()
        
        # In fallback, r.ward_id and r.id (booth_no/id) are our best links.
        # We need to allow if it matches an allowed booth, OR an allowed ward.
        # Note: assembly level access is handled via ward_id expansion in scope.
        rows = [
            r
            for r in rows
            if (not allowed_booth_ids or int(getattr(r, "booth_pk", None) or getattr(r, "booth_no", None) or 0) in allowed_booth_ids)
            and (not allowed_ward_ids or (r.ward_id is not None and int(r.ward_id) in allowed_ward_ids))
        ]
    result = []
    for r in rows:
        booth_pk = getattr(r, "booth_pk", None)
        booth_no = getattr(r, "booth_no", None)
        ward_id_val = int(r.ward_id) if r.ward_id is not None else None
        if booth_pk is not None:
            resolved_id = int(booth_pk)
        elif ward_id_val is not None and booth_no is not None:
            try:
                resolved_id = ward_id_val * 10000 + int(booth_no)
            except (TypeError, ValueError):
                resolved_id = ward_id_val
        else:
            resolved_id = int(booth_no) if booth_no is not None else 0
        result.append(
            {
                "boothId": resolved_id,
                "pollingStationAdrEn": r.name,
                "pollingStation_adr_en": r.name,
                "boothNameEn": r.name,
                "pollingStationAdrLocal": None,
                "wardId": ward_id_val,
                "tenantId": None,
                "boothNo": int(booth_no) if booth_no is not None and str(booth_no).isdigit() else booth_no,
            }
        )
    return result


@app.get(f"{CONTEXT_PATH}/api/booths/public")
def get_public_booths(
    wardId: Optional[int] = None,
    assemblyId: Optional[str] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER", "ASSEMBLY", "WARD")),
):
    scope = _resolve_access_scope_ids(db, current)
    booth_cols = _get_table_columns(db, "public", "booths")
    real_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
    id_col = "booth_no" if "booth_no" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
    name_col = "booth_add_en" if "booth_add_en" in booth_cols else ("polling_station_adr_en" if "polling_station_adr_en" in booth_cols else None)
    ward_ref = "ward_id" if "ward_id" in booth_cols else ("ward_no" if "ward_no" in booth_cols else None)
    assembly_ref = "assembly_id" if "assembly_id" in booth_cols else ("assembly_no" if "assembly_no" in booth_cols else None)
    if not id_col or not name_col:
        return []
    where = []
    params: Dict[str, Any] = {}
    if assemblyId:
        normalized = normalize_assembly_code(assemblyId)
        unpadded = str(int(normalized)) if normalized and normalized.isdigit() else normalized
        where.append(f"({assembly_ref} = :assembly_id OR CAST({assembly_ref} AS TEXT) = :assembly_unpadded)")
        params["assembly_id"] = normalized
        params["assembly_unpadded"] = unpadded
    if wardId and ward_ref:
        where.append(f"{ward_ref} = :ward_id")
        params["ward_id"] = wardId
    where_clause = f"WHERE {' AND '.join(where)}" if where else ""
    rows = db.execute(
        text(
            f"""
            SELECT {real_id_col if real_id_col else id_col} as id, {id_col} AS booth_no, {name_col} AS booth_name, {ward_ref if ward_ref else 'NULL'} AS ward_id
            FROM public.booths
            {where_clause}
            ORDER BY {id_col}
            """
        ),
        params,
    ).all()
    if scope:
        allowed_booth_ids = scope.get("allowed_booth_ids") or set()
        allowed_ward_ids = scope.get("allowed_ward_ids") or set()
        rows = [
            r
            for r in rows
            if (not allowed_booth_ids or int(r.id) in allowed_booth_ids or int(r.booth_no) in allowed_booth_ids)
            and (not allowed_ward_ids or (r.ward_id is not None and int(r.ward_id) in allowed_ward_ids))
        ]
    return [
        {
            "id": r.id,
            "boothNo": int(r.booth_no) if r.booth_no is not None else None,
            "boothNameEn": r.booth_name,
            "wardId": int(r.ward_id) if r.ward_id is not None else None,
        }
        for r in rows
    ]


@app.get(f"{CONTEXT_PATH}/api/wards")
def get_wards(
    assemblyId: Optional[str] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER", "ASSEMBLY", "WARD")),
):
    assemblyId = _normalize_assembly_code(assemblyId)
    scope = _resolve_access_scope_ids(db, current)
    print(f"[GET_WARDS] assemblyId={assemblyId}, role={current.role}, scope={scope}")
    if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
        print("[GET_WARDS] Empty scope, returning []")
        return []
    q = db.query(Ward)
    if assemblyId:
        q = q.filter(Ward.assembly_id == assemblyId)
    if current.tenantId is not None and current.role != "SUPER_ADMIN":
        q = q.filter(Ward.tenant_id == current.tenantId)
    if scope:
        allowed_ward_ids = {v for v in (scope.get("allowed_ward_ids") or set()) if v is not None}
        if allowed_ward_ids:
            q = q.filter(Ward.ward_id.in_(allowed_ward_ids))
    
    wards = q.order_by(Ward.ward_id.asc()).all()
    # If no wards in data schema, OR we are Super Admin (who should see EVERYTHING), 
    # check public schema for the full master list.
    if wards and (current.role != "SUPER_ADMIN" or not assemblyId or len(wards) > 50):
        return [
            {
                "wardId": w.ward_id,
                "wardNameEn": w.ward_name_en,
                "wardNameLocal": w.ward_name_local,
                "wardCode": w.ward_code,
                "assemblyId": w.assembly_id,
                "id": w.ward_id,
            }
            for w in wards
        ]

    ward_cols = _get_table_columns(db, "public", "wards")
    id_col = "ward_id" if "ward_id" in ward_cols else ("ward_no" if "ward_no" in ward_cols else ("id" if "id" in ward_cols else None))
    name_col = "ward_name_en" if "ward_name_en" in ward_cols else ("name_en" if "name_en" in ward_cols else ("ward_name_local" if "ward_name_local" in ward_cols else ("name_kannada" if "name_kannada" in ward_cols else None)))
    assembly_ref = "assembly_id" if "assembly_id" in ward_cols else ("assembly_no" if "assembly_no" in ward_cols else None)
    if not id_col or not name_col:
        return []
    where = []
    params: Dict[str, Any] = {}
    if assemblyId and assembly_ref:
        normalized = _normalize_assembly_code(assemblyId)
        unpadded = str(int(normalized)) if normalized and normalized.isdigit() else normalized
        where.append(f"({assembly_ref} = :assembly_id OR CAST({assembly_ref} AS TEXT) = :assembly_unpadded)")
        params["assembly_id"] = normalized
        params["assembly_unpadded"] = unpadded
    if current.role != "SUPER_ADMIN" and "tenant_id" in ward_cols:
        if current.tenantId:
            where.append("(tenant_id = :tid OR tenant_id IS NULL)")
            params["tid"] = current.tenantId
        else:
            # If user has no tenant, they can see everything in public schema for that assembly
            pass
    where_clause = f"WHERE {' AND '.join(where)}" if where else ""
    rows = db.execute(
        text(
            f"""
            SELECT {id_col} AS id, {name_col} AS name, {assembly_ref if assembly_ref else 'NULL'} AS assembly_id
            FROM public.wards
            {where_clause}
            ORDER BY {id_col}
            """
        ),
        params,
    ).all()
    if scope:
        allowed_ward_ids = scope.get("allowed_ward_ids") or set()
        allowed_assembly_ids = scope.get("allowed_assembly_ids") or set()
        
        raw_count = len(rows)
        if allowed_ward_ids or allowed_assembly_ids:
            rows = [
                r for r in rows 
                if (not allowed_ward_ids or (r.id is not None and int(r.id) in allowed_ward_ids))
                or (not allowed_assembly_ids or (r.assembly_id is not None and int(r.assembly_id) in allowed_assembly_ids))
            ]
        else:
            rows = []
        print(f"[GET_WARDS] Filtered rows from {raw_count} to {len(rows)}")
    return [
        {
            "wardId": int(r.id),
            "wardNameEn": r.name,
            "assemblyId": int(r.assembly_id) if r.assembly_id is not None else None,
            "tenantId": None,
        }
        for r in rows
    ]


@app.get(f"{CONTEXT_PATH}/api/assemblies")
def get_assemblies_plural(
    assemblyCode: Optional[str] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "USER")),
):
    q = db.query(Assembly)
    if current.tenantId is not None and current.role != "SUPER_ADMIN":
        q = q.filter(Assembly.tenant_id == current.tenantId)
    if assemblyCode:
        q = q.filter(Assembly.assembly_code == assemblyCode)

    assemblies = q.order_by(Assembly.assembly_id.asc()).all()
    return [
        {
            "assemblyId": a.assembly_id,
            "assemblyCode": a.assembly_code,
            "assemblyNameEn": a.assembly_name_en,
            "assemblyNameLocal": a.assembly_name_local,
            "tenantId": a.tenant_id,
        }
        for a in assemblies
    ]


@app.put(f"{CONTEXT_PATH}/api/voters/{{voterId}}")
def update_voter(voterId: int, payload: VoterUpdatePayload, db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    req = dict(payload.updateRequest or {})
    req["voterId"] = voterId

    updated_by = db.query(User).filter(User.first_name == current.firstName, User.phone == current.phone).first()
    if not updated_by:
        raise ValueError(f"User not found with Email: {current.firstName}")

    voter = db.query(Voter).filter(Voter.voter_id == req.get("voterId")).first()
    if not voter:
        raise ValueError(f"Voter not found with ID: {req.get('voterId')}")

    field_map = {
        "houseNoEn": "house_no_en",
        "houseNoLocal": "house_no_local",
        "relationType": "relation_type",
        "firstMiddleNameEn": "first_middle_name_en",
        "lastNameEn": "last_name_en",
        "firstMiddleNameLocal": "first_middle_name_local",
        "lastNameLocal": "last_name_local",
        "gender": "gender",
        "age": "age",
        "dob": "dob",
        "mobile": "mobile",
        "addressEn": "address_en",
        "addressLocal": "address_local",
        "latitude": "latitude",
        "longitude": "longitude",
    }

    logs: List[VoterChangeLog] = []
    for k, v in req.items():
        if k in ("voterId", "updatedByUserId") or v is None:
            continue
        if k not in field_map:
            raise ValueError(f"Invalid field in update request: {k}")

        attr = field_map[k]
        old = getattr(voter, attr)
        if old != v:
            logs.append(
                VoterChangeLog(
                    tenant_id=current.tenantId,
                    voter_id=voter.voter_id,
                    updated_by=updated_by.id,
                    field_name=k,
                    old_value=str(old) if old is not None else None,
                    new_value=str(v),
                    updated_at=datetime.utcnow(),
                    update_latitude=payload.updateLocationLat,
                    update_longitude=payload.updateLocationLng,
                )
            )
            setattr(voter, attr, v)

    db.add(voter)
    for log in logs:
        db.add(log)
    db.commit()

    voter_dict = {c.name: getattr(voter, c.name) for c in Voter.__table__.columns}
    return api_success("Voter updated successfully", voter_dict)


@app.put(f"{CONTEXT_PATH}/api/voters/by-epic/{{epic}}")
def update_public_voter_by_epic(epic: str, payload: PublicVoterUpdatePayload, db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    normalized_epic = normalize_optional_text(epic)
    if not normalized_epic:
        raise ValueError("EPIC is required")

    voter_cols = _get_table_columns(db, "public", "voters")
    if not voter_cols:
        raise ValueError("public.voters table not found")

    req = dict(payload.updateRequest or {})
    if not req:
        raise ValueError("updateRequest is required")

    unsupported = sorted(k for k in req.keys() if k not in VOTER_ENRICHMENT_FIELD_MAP)
    if unsupported:
        raise ValueError(f"Invalid field in update request: {unsupported[0]}")

    normalized_values: Dict[str, Any] = {
        api_field: _serialize_enrichment_value(api_field, req.get(api_field))
        for api_field in req.keys()
    }
    if not normalized_values:
        raise ValueError("No supported fields provided in updateRequest")

    ward_code = normalize_optional_text(payload.wardCode)
    booth_no = normalize_optional_text(payload.boothNo)

    where_parts = ["epic = :epic"]
    params: Dict[str, Any] = {"epic": normalized_epic}
    if ward_code and "ward_code" in voter_cols:
        where_parts.append("CAST(ward_code AS TEXT) = :ward_code")
        params["ward_code"] = ward_code
    if booth_no and "booth_no" in voter_cols:
        where_parts.append("CAST(booth_no AS TEXT) = :booth_no")
        params["booth_no"] = booth_no

    public_voter = db.execute(
        text(
            f"""
            SELECT epic, ward_code, booth_no, sl, house, name_en, name_kannada, gender, age, rel_eng, rel_kannada, rel_type, mobile
            FROM public.voters
            WHERE {" AND ".join(where_parts)}
            LIMIT 1
            """
        ),
        params,
    ).first()
    if not public_voter and (ward_code or booth_no):
        public_voter = db.execute(
            text(
                """
                SELECT epic, ward_code, booth_no, sl, house, name_en, name_kannada, gender, age, rel_eng, rel_kannada, rel_type, mobile
                FROM public.voters
                WHERE epic = :epic
                LIMIT 1
                """
            ),
            {"epic": normalized_epic},
        ).first()
    if not public_voter:
        raise ResourceNotFoundException("Public voter", "epic", normalized_epic)

    updated_by = db.query(User).filter(User.first_name == current.firstName, User.phone == current.phone).first()
    if not updated_by:
        volunteer_user = db.query(VolunteerUser).filter(VolunteerUser.first_name == current.firstName, VolunteerUser.phone == current.phone).first()
        if volunteer_user:
            updated_by = _ensure_user_from_volunteer(db, volunteer_user)
    enrichment = db.query(VoterEnrichment).filter(VoterEnrichment.epic == normalized_epic).first()
    if not enrichment:
        enrichment = VoterEnrichment(
            epic=normalized_epic,
            ward_code=str(public_voter.ward_code) if public_voter.ward_code is not None else ward_code,
            booth_no=str(public_voter.booth_no) if public_voter.booth_no is not None else booth_no,
        )
        db.add(enrichment)

    updated_fields = _parse_updated_fields(enrichment.updated_fields)
    for api_field, serialized_value in normalized_values.items():
        setattr(enrichment, VOTER_ENRICHMENT_FIELD_MAP[api_field], serialized_value)
        updated_fields.add(api_field)

    enrichment.ward_code = str(public_voter.ward_code) if public_voter.ward_code is not None else enrichment.ward_code
    enrichment.booth_no = str(public_voter.booth_no) if public_voter.booth_no is not None else enrichment.booth_no
    enrichment.updated_fields = json.dumps(sorted(updated_fields))
    enrichment.updated_at = datetime.utcnow()
    enrichment.updated_by = updated_by.id if updated_by else None
    enrichment.updated_by_name = updated_by.first_name if updated_by else None
    enrichment.updated_by_phone = updated_by.phone if updated_by else None
    db.commit()
    db.refresh(enrichment)
    return api_success(
        "Public voter updated successfully",
        _merge_voter_payload_with_enrichment(_build_public_voter_result(public_voter), _build_voter_enrichment_payload(enrichment)),
    )


@app.get(f"{CONTEXT_PATH}/api/voters")
def get_voters(
    assemblyCode: str,
    page: int = 0,
    size: int = 500,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("ADMIN", "USER")),
):
    scope = _resolve_access_scope_ids(db, current)
    if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
        return []
    q = (
        db.query(Voter)
        .join(Booth, Voter.booth_id == Booth.booth_id)
        .join(Ward, Booth.ward_id == Ward.ward_id)
        .join(Assembly, Ward.assembly_id == Assembly.assembly_id)
        .filter(
            Voter.tenant_id == current.tenantId,
            Assembly.assembly_code == assemblyCode,
        )
        .order_by(Voter.voter_id.asc())
    )
    if scope:
        if scope.get("allowed_booth_ids"):
            q = q.filter(Booth.booth_id.in_(scope.get("allowed_booth_ids")))
        elif scope.get("allowed_ward_ids"):
            q = q.filter(Ward.ward_id.in_(scope.get("allowed_ward_ids")))

    total = q.count()
    voters = q.offset(page * size).limit(size).all()
    return [_build_voter_map(v) for v in voters]


@app.get(f"{CONTEXT_PATH}/api/voter-search")
@app.get(f"{CONTEXT_PATH}/api/voters/search")
def search_voters(
    assemblyCode: str,
    searchQuery: Optional[str] = None,
    wardId: Optional[int] = None,
    boothNumber: Optional[str] = None,
    mobileNumber: Optional[str] = None,
    epicId: Optional[str] = None,
    relationName: Optional[str] = None,
    houseNumber: Optional[str] = None,
    page: int = 0,
    size: int = 50,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "USER", "ADMIN")),
):
    scope = _resolve_access_scope_ids(db, current)
    if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
        empty_payload = api_success("Voter search fetched", [])
        empty_payload["data"]["meta"] = {
            "total": 0,
            "male": 0,
            "female": 0,
            "returned": 0,
            "page": page,
            "size": max(1, min(size, 2000)),
            "hasMore": False,
        }
        return empty_payload
    booth_cols = _get_table_columns(db, "public", "booths")
    voter_cols = _get_table_columns(db, "public", "voters")
    ward_cols = _get_table_columns(db, "public", "wards")

    booth_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
    booth_no_col = "booth_no" if "booth_no" in booth_cols else booth_id_col
    booth_ward_code_col = "ward_code" if "ward_code" in booth_cols else None
    booth_ward_id_col = "ward_id" if "ward_id" in booth_cols else None
    booth_name_en_col = "booth_add_en" if "booth_add_en" in booth_cols else ("polling_station_adr_en" if "polling_station_adr_en" in booth_cols else None)
    booth_name_local_col = "booth_add_local" if "booth_add_local" in booth_cols else ("polling_station_adr_local" if "polling_station_adr_local" in booth_cols else None)

    if not booth_id_col or not booth_no_col:
        return JSONResponse(status_code=404, content=api_error("Search failed", "public.booths missing required columns"))

    ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
    ward_code_col = "ward_code" if "ward_code" in ward_cols else ("ward_no" if "ward_no" in ward_cols else None)
    ward_name_en_col = "ward_name_en" if "ward_name_en" in ward_cols else ("name_en" if "name_en" in ward_cols else None)
    ward_name_local_col = "ward_name_local" if "ward_name_local" in ward_cols else ("name_kannada" if "name_kannada" in ward_cols else None)

    if scope and scope.get("allowed_assembly_ids"):
        assembly_cols = _ensure_public_assembly_code(db)
        assembly_pk_col = "id" if "id" in assembly_cols else ("assembly_id" if "assembly_id" in assembly_cols else None)
        assembly_no_col = "assembly_no" if "assembly_no" in assembly_cols else ("assembly_id" if "assembly_id" in assembly_cols else None)
        assembly_code_expr = "assembly_code" if "assembly_code" in assembly_cols else (
            f"LPAD(CAST({assembly_no_col} AS TEXT), 12, '0')" if assembly_no_col else "NULL"
        )
        # For the assembly lookup itself, we don't strictly filter by tenant_id here
        # because the user's assignment scope (allowed_assembly_ids) will restrict them below.
        assembly_row = db.execute(
            text(
                f"""
                SELECT {assembly_pk_col if assembly_pk_col else 'NULL'} AS assembly_pk,
                       {assembly_no_col if assembly_no_col else 'NULL'} AS assembly_no
                FROM public.assembly
                WHERE ({assembly_code_expr} = :assembly_code
                   OR CAST({assembly_no_col if assembly_no_col else assembly_code_expr} AS TEXT) = :assembly_no_text)
                LIMIT 1
                """
            ),
            {
                "assembly_code": normalize_assembly_code(assemblyCode),
                "assembly_no_text": str(int(normalize_assembly_code(assemblyCode))),
            },
        ).first()
        if not assembly_row:
            raise HTTPException(status_code=404, detail="Assembly not found")
        allowed_assembly_ids = {v for v in (scope.get("allowed_assembly_ids") or set()) if v is not None}
        assembly_id = assembly_row.assembly_pk if assembly_row.assembly_pk is not None else assembly_row.assembly_no
        if allowed_assembly_ids and assembly_id not in allowed_assembly_ids:
            raise HTTPException(status_code=403, detail="Access denied for requested assembly")

    ward_by_id: Dict[int, Dict[str, Any]] = {}
    if ward_id_col:
        assembly_ref_ward = "assembly_id" if "assembly_id" in ward_cols else ("assembly_no" if "assembly_no" in ward_cols else None)
        ward_rows = db.execute(
            text(
                f"""
                SELECT
                    {ward_id_col} AS ward_id,
                    {ward_code_col if ward_code_col else 'NULL'} AS ward_code,
                    {ward_name_en_col if ward_name_en_col else 'NULL'} AS ward_name_en,
                    {ward_name_local_col if ward_name_local_col else 'NULL'} AS ward_name_local
                FROM public.wards
                WHERE ({assembly_ref_ward} = :assembly_id OR CAST({assembly_ref_ward} AS TEXT) = :assembly_no_text)
                """
            ),
            {
                "assembly_id": normalize_assembly_code(assemblyCode),
                "assembly_no_text": str(int(normalize_assembly_code(assemblyCode))),
            }
        ).all()
        for r in ward_rows:
            if scope and scope.get("allowed_ward_ids") and int(r.ward_id) not in scope.get("allowed_ward_ids"):
                continue
            ward_by_id[int(r.ward_id)] = {
                "wardId": int(r.ward_id),
                "wardCode": str(r.ward_code) if r.ward_code is not None else None,
                "wardNameEn": r.ward_name_en,
                "wardNameLocal": r.ward_name_local,
            }

    ward_code_filter: Optional[str] = None
    if wardId is not None and wardId in ward_by_id:
        ward_code_filter = ward_by_id[wardId].get("wardCode")
    if scope and scope.get("allowed_ward_ids") and wardId is not None and wardId not in scope.get("allowed_ward_ids"):
        empty_payload = api_success("Voter search fetched", [])
        empty_payload["data"]["meta"] = {
            "total": 0,
            "male": 0,
            "female": 0,
            "returned": 0,
            "page": page,
            "size": max(1, min(size, 2000)),
            "hasMore": False,
        }
        return empty_payload

    # Booths don't have direct assembly_id, we must join with wards
    assembly_ref_ward = "assembly_id" if "assembly_id" in ward_cols else ("assembly_no" if "assembly_no" in ward_cols else None)
    booth_sql = f"""
        SELECT
            b.{booth_id_col} AS booth_id,
            b.{booth_no_col} AS booth_no,
            b.{booth_ward_code_col if booth_ward_code_col else 'NULL'} AS ward_code,
            b.{booth_ward_id_col if booth_ward_id_col else 'NULL'} AS ward_id,
            b.{booth_name_en_col if booth_name_en_col else 'NULL'} AS booth_name_en,
            b.{booth_name_local_col if booth_name_local_col else 'NULL'} AS booth_name_local
        FROM public.booths b
        JOIN public.wards w ON b.{booth_ward_id_col} = w.{ward_id_col}
        WHERE (w.{assembly_ref_ward} = :assembly_id OR CAST(w.{assembly_ref_ward} AS TEXT) = :assembly_no_text)
    """
    booth_rows = db.execute(text(booth_sql), {
        "assembly_id": normalize_assembly_code(assemblyCode),
        "assembly_no_text": str(int(normalize_assembly_code(assemblyCode))),
    }).all()
    booth_by_key: Dict[tuple, Dict[str, Any]] = {}
    booths_by_no: Dict[str, List[Dict[str, Any]]] = {}
    for b in booth_rows:
        if scope:
            allowed_booth_ids = {v for v in (scope.get("allowed_booth_ids") or set()) if v is not None}
            allowed_ward_ids = {v for v in (scope.get("allowed_ward_ids") or set()) if v is not None}
            booth_pk = int(b.booth_id)
            ward_pk = int(b.ward_id) if b.ward_id is not None else None
            if allowed_booth_ids and booth_pk not in allowed_booth_ids:
                continue
            if allowed_ward_ids and ward_pk is not None and ward_pk not in allowed_ward_ids:
                continue
        row = {
            "boothId": int(b.booth_id),
            "boothNo": str(b.booth_no if b.booth_no is not None else b.booth_id),
            "wardCode": str(b.ward_code) if b.ward_code is not None else None,
            "wardId": int(b.ward_id) if b.ward_id is not None else None,
            "boothNameEn": b.booth_name_en,
            "boothNameLocal": b.booth_name_local,
        }
        booth_by_key[(row["wardCode"], row["boothNo"])] = row
        booths_by_no.setdefault(row["boothNo"], []).append(row)

    voter_id_col = "voter_id" if "voter_id" in voter_cols else ("id" if "id" in voter_cols else None)
    voter_sr_col = "sl" if "sl" in voter_cols else ("sr_no" if "sr_no" in voter_cols else None)
    voter_epic_col = "epic" if "epic" in voter_cols else ("epic_no" if "epic_no" in voter_cols else None)
    voter_name_en_col = "name_en" if "name_en" in voter_cols else ("first_middle_name_en" if "first_middle_name_en" in voter_cols else None)
    voter_name_local_col = "name_kannada" if "name_kannada" in voter_cols else ("first_middle_name_local" if "first_middle_name_local" in voter_cols else None)
    voter_relation_en_col = "relation_name_en" if "relation_name_en" in voter_cols else ("relation_first_middle_name_en" if "relation_first_middle_name_en" in voter_cols else ("rel_eng" if "rel_eng" in voter_cols else None))
    voter_relation_local_col = "relation_name_local" if "relation_name_local" in voter_cols else ("relation_first_middle_name_local" if "relation_first_middle_name_local" in voter_cols else ("rel_kannada" if "rel_kannada" in voter_cols else None))
    voter_relation_type_col = "relation_type" if "relation_type" in voter_cols else ("rel_type" if "rel_type" in voter_cols else None)
    voter_age_col = "age" if "age" in voter_cols else None
    voter_house_col = "house" if "house" in voter_cols else ("house_no_en" if "house_no_en" in voter_cols else None)
    voter_relation_en_col = "relation_name_en" if "relation_name_en" in voter_cols else ("relation_first_middle_name_en" if "relation_first_middle_name_en" in voter_cols else ("rel_eng" if "rel_eng" in voter_cols else None))
    voter_relation_local_col = "relation_name_local" if "relation_name_local" in voter_cols else ("relation_first_middle_name_local" if "relation_first_middle_name_local" in voter_cols else ("rel_kannada" if "rel_kannada" in voter_cols else None))
    voter_relation_type_col = "relation_type" if "relation_type" in voter_cols else ("rel_type" if "rel_type" in voter_cols else None)
    voter_age_col = "age" if "age" in voter_cols else None
    voter_gender_col = "gender" if "gender" in voter_cols else None
    voter_mobile_col = "mobile" if "mobile" in voter_cols else None
    voter_booth_no_col = "booth_no" if "booth_no" in voter_cols else ("booth_id" if "booth_id" in voter_cols else None)
    voter_ward_code_col = "ward_code" if "ward_code" in voter_cols else None

    if not voter_booth_no_col:
        return JSONResponse(status_code=404, content=api_error("Search failed", "public.voters missing booth mapping column"))

    where_parts = ["1=1"]
    params: Dict[str, Any] = {"limit": max(1, min(size, 2000)), "offset": max(page, 0) * max(1, min(size, 2000))}
    
    if current.role != "SUPER_ADMIN" and "tenant_id" in voter_cols:
        if current.tenantId:
            where_parts.append("(tenant_id = :tid OR tenant_id IS NULL OR tenant_id = '')")
            params["tid"] = current.tenantId
        else:
            # For users with no tenant, we rely on the assembly-scoped ward/booth filters
            pass

    allowed_ward_codes = sorted({str(v.get("wardCode")) for v in ward_by_id.values() if v.get("wardCode")})
    allowed_booth_nos = sorted({row.get("boothNo") for row in booth_by_key.values() if row.get("boothNo")})
    
    if allowed_ward_codes and voter_ward_code_col:
        clause, clause_params = _build_in_clause(f"CAST({voter_ward_code_col} AS TEXT)", allowed_ward_codes, "scope_ward_code")
        where_parts.append(clause)
        params.update(clause_params)
    if allowed_booth_nos:
        clause, clause_params = _build_in_clause(f"CAST({voter_booth_no_col} AS TEXT)", allowed_booth_nos, "scope_booth_no")
        where_parts.append(clause)
        params.update(clause_params)

    if ward_code_filter and voter_ward_code_col:
        where_parts.append(f"{voter_ward_code_col} = :ward_code")
        params["ward_code"] = ward_code_filter
    if boothNumber:
        where_parts.append(f"CAST({voter_booth_no_col} AS TEXT) ILIKE :booth_number")
        params["booth_number"] = f"%{boothNumber.strip()}%"
    if mobileNumber and voter_mobile_col:
        where_parts.append(f"CAST({voter_mobile_col} AS TEXT) ILIKE :mobile_number")
        params["mobile_number"] = f"%{mobileNumber.strip()}%"
    if epicId and voter_epic_col:
        where_parts.append(f"CAST({voter_epic_col} AS TEXT) ILIKE :epic_id")
        params["epic_id"] = f"%{epicId.strip()}%"
    if houseNumber and voter_house_col:
        where_parts.append(f"CAST({voter_house_col} AS TEXT) ILIKE :house_number")
        params["house_number"] = f"%{houseNumber.strip()}%"
    if relationName and voter_relation_en_col:
        where_parts.append(f"CAST({voter_relation_en_col} AS TEXT) ILIKE :relation_name")
        params["relation_name"] = f"%{relationName.strip()}%"

    search_fields: List[str] = []
    if voter_name_en_col:
        search_fields.append(f"CAST({voter_name_en_col} AS TEXT)")
    if voter_epic_col:
        search_fields.append(f"CAST({voter_epic_col} AS TEXT)")
    if voter_mobile_col:
        search_fields.append(f"CAST({voter_mobile_col} AS TEXT)")
    if voter_sr_col:
        search_fields.append(f"CAST({voter_sr_col} AS TEXT)")
    search_fields.append(f"CAST({voter_booth_no_col} AS TEXT)")
    if voter_house_col:
        search_fields.append(f"CAST({voter_house_col} AS TEXT)")
    if voter_relation_en_col:
        search_fields.append(f"CAST({voter_relation_en_col} AS TEXT)")
    if searchQuery and search_fields:
        where_parts.append("(" + " OR ".join([f"{f} ILIKE :search_query" for f in search_fields]) + ")")
        params["search_query"] = f"%{searchQuery.strip()}%"

    where_sql = " AND ".join(where_parts)

    order_expr = f"CAST({voter_sr_col} AS INT)" if voter_sr_col else (f"CAST({voter_epic_col} AS TEXT) ASC NULLS LAST" if voter_epic_col else voter_booth_no_col)
    voter_id_expr = f"{voter_id_col}" if voter_id_col else (f"COALESCE(CAST({voter_sr_col} AS BIGINT), ROW_NUMBER() OVER ())" if voter_sr_col else "ROW_NUMBER() OVER ()")
    voters_rows = db.execute(
        text(
            f"""
            SELECT
                {voter_id_expr} AS voter_id,
                {voter_sr_col if voter_sr_col else 'NULL'} AS sl,
                {voter_epic_col if voter_epic_col else 'NULL'} AS epic_no,
                {voter_name_en_col if voter_name_en_col else 'NULL'} AS name_en,
                {voter_name_local_col if voter_name_local_col else 'NULL'} AS name_local,
                {voter_relation_en_col if voter_relation_en_col else 'NULL'} AS relation_name_en,
                {voter_relation_local_col if voter_relation_local_col else 'NULL'} AS relation_name_local,
                {voter_relation_type_col if voter_relation_type_col else 'NULL'} AS relation_type,
                {voter_house_col if voter_house_col else 'NULL'} AS house_no_en,
                {voter_gender_col if voter_gender_col else 'NULL'} AS gender,
                {voter_age_col if voter_age_col else 'NULL'} AS age,
                {voter_mobile_col if voter_mobile_col else 'NULL'} AS mobile,
                {voter_booth_no_col} AS booth_no,
                {voter_ward_code_col if voter_ward_code_col else 'NULL'} AS ward_code
            FROM public.voters
            WHERE {where_sql}
            ORDER BY {order_expr}
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).all()

    result: List[Dict[str, Any]] = []
    for v in voters_rows:
        ward_code = str(v.ward_code) if v.ward_code is not None else None
        booth_no = str(v.booth_no)
        booth = booth_by_key.get((ward_code, booth_no))
        if not booth:
            candidates = booths_by_no.get(booth_no, [])
            booth = candidates[0] if candidates else None

        ward_id_mapped = booth.get("wardId") if booth else None
        if wardId is not None and ward_id_mapped is not None and int(ward_id_mapped) != int(wardId):
            continue

        ward_info = ward_by_id.get(int(ward_id_mapped)) if ward_id_mapped is not None else None
        result.append(
            {
                "voterId": int(v.voter_id) if v.voter_id is not None else None,
                "sl": v.sl,
                "epicNo": v.epic_no,
                "firstMiddleNameEn": v.name_en,
                "lastNameEn": "",
                "firstMiddleNameLocal": v.name_local,
                "lastNameLocal": "",
                "relationFirstMiddleNameEn": v.relation_name_en,
                "relationFirstMiddleNameLocal": v.relation_name_local,
                "relationType": v.relation_type,
                "relationLastNameEn": "",
                "houseNoEn": str(v.house_no_en) if v.house_no_en is not None else None,
                "houseNoLocal": None,
                "gender": v.gender,
                "age": v.age,
                "mobile": str(v.mobile) if v.mobile is not None else None,
                "wardCode": ward_code,
                "boothNo": booth_no,
                "boothInfo": {
                    "boothId": booth.get("boothId") if booth else None,
                    "boothNo": booth.get("boothNo") if booth else booth_no,
                    "wardCode": booth.get("wardCode") if booth else ward_code,
                    "boothNameEn": booth.get("boothNameEn") if booth else None,
                    "boothNameLocal": booth.get("boothNameLocal") if booth else None,
                },
                "wardId": ward_id_mapped,
                "wardNameEn": ward_info.get("wardNameEn") if ward_info else None,
                "wardNameLocal": ward_info.get("wardNameLocal") if ward_info else None,
            }
        )
    _merge_voter_payloads_with_enrichment(db, result)

    count_sql = f"""
        SELECT
            COUNT(*)::int AS total_count,
            SUM(CASE WHEN UPPER(COALESCE({voter_gender_col}, '')) LIKE 'M%' THEN 1 ELSE 0 END)::int AS male_count,
            SUM(CASE WHEN UPPER(COALESCE({voter_gender_col}, '')) LIKE 'F%' THEN 1 ELSE 0 END)::int AS female_count
        FROM public.voters
        WHERE {where_sql}
    """ if voter_gender_col else f"""
        SELECT
            COUNT(*)::int AS total_count,
            0::int AS male_count,
            0::int AS female_count
        FROM public.voters
        WHERE {where_sql}
    """
    count_params = {k: v for k, v in params.items() if k not in {"limit", "offset"}}
    counts = db.execute(text(count_sql), count_params).first()
    total_count = int((counts.total_count if counts else 0) or 0)
    male_count = int((counts.male_count if counts else 0) or 0)
    female_count = int((counts.female_count if counts else 0) or 0)

    payload = api_success("Voter search fetched", result)
    payload["data"]["meta"] = {
        "total": total_count,
        "male": male_count,
        "female": female_count,
        "returned": len(result),
        "page": page,
        "size": max(1, min(size, 2000)),
        "hasMore": (max(page, 0) * max(1, min(size, 2000)) + len(result)) < total_count,
    }
    return payload


@app.get(f"{CONTEXT_PATH}/api/poll-day/config")
def get_poll_day_config(assemblyId: Optional[str] = None, wardId: Optional[str] = None, db: Session = Depends(get_db)):
    # Convert to int if provided
    aid = int(assemblyId) if assemblyId and str(assemblyId).isdigit() else None
    wid = int(wardId) if wardId and str(wardId).isdigit() else None
    
    # Check exact match for the specific level requested
    exact_config = db.query(PollDayConfig).filter(
        PollDayConfig.assembly_id == aid, 
        PollDayConfig.ward_id == wid, 
        PollDayConfig.enabled == True
    ).first()
    
    # Determine if functionality should be 'active' for this context (global or specific)
    is_active = exact_config is not None
    if not is_active and (aid is not None or wid is not None):
        global_config = db.query(PollDayConfig).filter(
            PollDayConfig.assembly_id == None, 
            PollDayConfig.ward_id == None, 
            PollDayConfig.enabled == True
        ).first()
        if global_config:
            is_active = True
            
    return {"enabled": exact_config is not None, "isActive": is_active}

@app.post(f"{CONTEXT_PATH}/api/poll-day/config")
def set_poll_day_config(payload: Dict, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN"))):
    assembly_id = payload.get("assemblyId") 
    ward_id = payload.get("wardId")         
    enabled = payload.get("enabled", False)
    
    aid = int(assembly_id) if assembly_id and str(assembly_id).isdigit() else None
    wid = int(ward_id) if ward_id and str(ward_id).isdigit() else None
    
    config = db.query(PollDayConfig).filter(PollDayConfig.assembly_id == aid, PollDayConfig.ward_id == wid).first()
    if not config:
        config = PollDayConfig(assembly_id=aid, ward_id=wid, enabled=enabled)
        db.add(config)
    else:
        config.enabled = enabled
    db.commit()
    return api_success("Poll day config updated", {"enabled": enabled})

@app.get(f"{CONTEXT_PATH}/api/voters/by-booth")
def get_voters_by_booth(
    boothId: int,
    wardId: Optional[int] = None,
    boothNo: Optional[int] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "USER", "ADMIN")),
):
    scope = _resolve_access_scope_ids(db, current)
    if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
        raise HTTPException(status_code=403, detail="No access scope defined for this user")
    booth_cols = _get_table_columns(db, "public", "booths")
    voter_cols = _get_table_columns(db, "public", "voters")
    ward_cols = _get_table_columns(db, "public", "wards")

    booth_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
    booth_no_col = "booth_no" if "booth_no" in booth_cols else booth_id_col
    booth_ward_code_col = "ward_code" if "ward_code" in booth_cols else None
    booth_ward_id_col = "ward_id" if "ward_id" in booth_cols else None
    booth_name_en_col = "booth_add_en" if "booth_add_en" in booth_cols else ("polling_station_adr_en" if "polling_station_adr_en" in booth_cols else None)
    booth_name_local_col = "booth_add_local" if "booth_add_local" in booth_cols else ("polling_station_adr_local" if "polling_station_adr_local" in booth_cols else None)

    if not booth_id_col or not booth_no_col:
        return JSONResponse(status_code=404, content=api_error("Booth not found", "public.booths missing required columns"))

    booth_select = f"""
            SELECT
                {booth_id_col} AS booth_id,
                {booth_no_col} AS booth_no,
                {booth_ward_code_col if booth_ward_code_col else 'NULL'} AS ward_code,
                {booth_ward_id_col if booth_ward_id_col else 'NULL'} AS ward_id,
                {booth_name_en_col if booth_name_en_col else 'NULL'} AS booth_name_en,
                {booth_name_local_col if booth_name_local_col else 'NULL'} AS booth_name_local
            FROM public.booths
            """
    booth_row = db.execute(
        text(f"{booth_select} WHERE {booth_id_col} = :booth_id LIMIT 1"),
        {"booth_id": boothId},
    ).first()
    def _lookup_by_ward_and_no(target_ward: int, target_no: int):
        if not (booth_ward_id_col and booth_no_col):
            return None
        return db.execute(
            text(
                f"""
                {booth_select}
                WHERE {booth_ward_id_col} = :ward_id
                  AND (
                    CAST({booth_no_col} AS TEXT) = :booth_no_txt
                    OR CAST({booth_no_col} AS INT) = :booth_no_int
                  )
                LIMIT 1
                """
            ),
            {
                "ward_id": target_ward,
                "booth_no_txt": str(target_no),
                "booth_no_int": target_no,
            },
        ).first()

    # Fallback: explicit ward + booth number (mobile snapshot / booth list)
    if not booth_row and wardId is not None and boothNo is not None:
        booth_row = _lookup_by_ward_and_no(int(wardId), int(boothNo))
    # Fallback: composite boothId from GET /booths (wardId * 10000 + boothNo)
    if not booth_row and boothId >= 10000:
        composite_ward = int(boothId) // 10000
        composite_no = int(boothId) % 10000
        booth_row = _lookup_by_ward_and_no(composite_ward, composite_no)
    # Fallback: data.booths ORM primary key
    if not booth_row:
        orm_booth = db.query(Booth).filter(Booth.booth_id == boothId).first()
        if orm_booth:
            booth_row = type("BoothRow", (), {
                "booth_id": orm_booth.booth_id,
                "booth_no": orm_booth.booth_no,
                "ward_code": orm_booth.ward_code,
                "ward_id": orm_booth.ward_id,
                "booth_name_en": orm_booth.polling_station_adr_en or orm_booth.booth_add_en,
                "booth_name_local": orm_booth.polling_station_adr_local or orm_booth.booth_add_local,
            })()
    if not booth_row:
        return JSONResponse(status_code=404, content=api_error("Booth not found", f"Invalid boothId: {boothId}"))
    if scope:
        allowed_booth_ids = scope.get("allowed_booth_ids") or set()
        allowed_ward_ids = scope.get("allowed_ward_ids") or set()
        if allowed_booth_ids and int(booth_row.booth_id) not in allowed_booth_ids:
            raise HTTPException(status_code=403, detail="Access denied for requested booth")
        if allowed_ward_ids and booth_row.ward_id is not None and int(booth_row.ward_id) not in allowed_ward_ids:
            raise HTTPException(status_code=403, detail="Access denied for requested booth")

    ward_name_en = None
    ward_name_local = None
    if booth_row.ward_id is not None and ("id" in ward_cols or "ward_id" in ward_cols):
        ward_id_col = "id" if "id" in ward_cols else "ward_id"
        ward_name_en_col = "ward_name_en" if "ward_name_en" in ward_cols else ("name_en" if "name_en" in ward_cols else None)
        ward_name_local_col = "ward_name_local" if "ward_name_local" in ward_cols else ("name_kannada" if "name_kannada" in ward_cols else None)
        if ward_name_en_col or ward_name_local_col:
            ward_row = db.execute(
                text(
                    f"""
                    SELECT
                        {ward_name_en_col if ward_name_en_col else 'NULL'} AS ward_name_en,
                        {ward_name_local_col if ward_name_local_col else 'NULL'} AS ward_name_local
                    FROM public.wards
                    WHERE {ward_id_col} = :ward_id
                    LIMIT 1
                    """
                ),
                {"ward_id": booth_row.ward_id},
            ).first()
            if ward_row:
                ward_name_en = ward_row.ward_name_en
                ward_name_local = ward_row.ward_name_local

    voter_id_col = "voter_id" if "voter_id" in voter_cols else ("id" if "id" in voter_cols else None)
    voter_sr_col = "sl" if "sl" in voter_cols else ("sr_no" if "sr_no" in voter_cols else None)
    voter_epic_col = "epic" if "epic" in voter_cols else ("epic_no" if "epic_no" in voter_cols else None)
    voter_name_en_col = "name_en" if "name_en" in voter_cols else ("first_middle_name_en" if "first_middle_name_en" in voter_cols else None)
    voter_name_local_col = "name_kannada" if "name_kannada" in voter_cols else ("first_middle_name_local" if "first_middle_name_local" in voter_cols else None)
    voter_relation_en_col = "relation_name_en" if "relation_name_en" in voter_cols else ("relation_first_middle_name_en" if "relation_first_middle_name_en" in voter_cols else ("rel_eng" if "rel_eng" in voter_cols else None))
    voter_relation_local_col = "relation_name_local" if "relation_name_local" in voter_cols else ("relation_first_middle_name_local" if "relation_first_middle_name_local" in voter_cols else ("rel_kannada" if "rel_kannada" in voter_cols else None))
    voter_relation_type_col = "relation_type" if "relation_type" in voter_cols else ("rel_type" if "rel_type" in voter_cols else None)
    voter_age_col = "age" if "age" in voter_cols else None
    voter_house_col = "house" if "house" in voter_cols else ("house_no_en" if "house_no_en" in voter_cols else None)
    voter_gender_col = "gender" if "gender" in voter_cols else None
    voter_booth_no_col = "booth_no" if "booth_no" in voter_cols else ("booth_id" if "booth_id" in voter_cols else None)
    voter_ward_code_col = "ward_code" if "ward_code" in voter_cols else None

    if not voter_booth_no_col:
        return JSONResponse(status_code=404, content=api_error("Voters not found", "public.voters missing booth mapping column"))

    where_clause = f"{voter_booth_no_col} = :booth_no"
    params: Dict[str, Any] = {"booth_no": booth_row.booth_no}
    if voter_ward_code_col and booth_row.ward_code is not None:
        where_clause += f" AND {voter_ward_code_col} = :ward_code"
        params["ward_code"] = booth_row.ward_code

    voter_id_expr = f"{voter_id_col}" if voter_id_col else (f"COALESCE(CAST({voter_sr_col} AS BIGINT), ROW_NUMBER() OVER ())" if voter_sr_col else "ROW_NUMBER() OVER ()")
    order_expr = (
        f"CAST({voter_sr_col} AS INT) ASC NULLS LAST, CAST({voter_epic_col} AS TEXT) ASC NULLS LAST"
        if (voter_sr_col and voter_epic_col)
        else (f"CAST({voter_sr_col} AS INT) ASC NULLS LAST" if voter_sr_col else (f"CAST({voter_epic_col} AS TEXT) ASC NULLS LAST" if voter_epic_col else voter_booth_no_col))
    )
    voters_rows = db.execute(
        text(
            f"""
            SELECT
                {voter_id_expr} AS voter_id,
                {voter_sr_col if voter_sr_col else 'NULL'} AS sl,
                {voter_epic_col if voter_epic_col else 'NULL'} AS epic_no,
                {voter_name_en_col if voter_name_en_col else 'NULL'} AS name_en,
                {voter_name_local_col if voter_name_local_col else 'NULL'} AS name_local,
                {voter_relation_en_col if voter_relation_en_col else 'NULL'} AS relation_name_en,
                {voter_relation_local_col if voter_relation_local_col else 'NULL'} AS relation_name_local,
                {voter_relation_type_col if voter_relation_type_col else 'NULL'} AS relation_type,
                {voter_house_col if voter_house_col else 'NULL'} AS house_no_en,
                {voter_gender_col if voter_gender_col else 'NULL'} AS gender,
                {voter_age_col if voter_age_col else 'NULL'} AS age
            FROM public.voters
            WHERE {where_clause}
            ORDER BY {order_expr}
            """
        ),
        params,
    ).all()

    voters = []
    male = 0
    female = 0
    for v in voters_rows:
        g = (v.gender or "").upper()
        if g.startswith("M"):
            male += 1
        if g.startswith("F"):
            female += 1
        voters.append(
            {
                "voterId": int(v.voter_id) if v.voter_id is not None else None,
                "sl": v.sl,
                "epicNo": v.epic_no,
                "firstMiddleNameEn": v.name_en,
                "lastNameEn": "",
                "firstMiddleNameLocal": v.name_local,
                "lastNameLocal": "",
                "relationFirstMiddleNameEn": v.relation_name_en,
                "relationFirstMiddleNameLocal": v.relation_name_local,
                "relationType": v.relation_type,
                "houseNoEn": str(v.house_no_en) if v.house_no_en is not None else None,
                "houseNoLocal": None,
                "gender": v.gender,
                "age": v.age,
                "dob": None,
                "mobile": None,
                "wardCode": str(booth_row.ward_code) if booth_row.ward_code is not None else None,
                "boothNo": str(booth_row.booth_no) if booth_row.booth_no is not None else None,
                "addressEn": None,
                "addressLocal": None,
                "status": None,
                "community": None,
                "caste": None,
                "residenceType": None,
                "civicIssue": None,
                "motherTongue": None,
                "team": None,
                "ownership": None,
                "education": None,
                "natureOfVoter": None,
                "latitude": None,
                "longitude": None,
                "wardId": booth_row.ward_id,
                "wardNameEn": ward_name_en,
                "wardNameLocal": ward_name_local,
            }
        )
    _merge_voter_payloads_with_enrichment(db, voters)

    return api_success(
        "Booth voters fetched",
        {
            "boothId": int(booth_row.booth_id),
            "boothNameEn": booth_row.booth_name_en,
            "boothNameLocal": booth_row.booth_name_local,
            "wardId": booth_row.ward_id,
            "wardCode": str(booth_row.ward_code) if booth_row.ward_code is not None else None,
            "wardNameEn": ward_name_en,
            "wardNameLocal": ward_name_local,
            "voterStats": {"total": len(voters), "male": male, "female": female},
            "voters": voters,
        },
    )


@app.get(f"{CONTEXT_PATH}/api/voters/snapshot")
def get_snapshot(
    assemblyCode: str,
    request: Request,
    includeVoters: bool = True,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "USER", "ADMIN")),
):
    assemblyCode = _normalize_assembly_code(assemblyCode)
    try:
        scope = _resolve_access_scope_ids(db, current)
        if scope and not (scope.get("allowed_assembly_ids") or scope.get("allowed_ward_ids") or scope.get("allowed_booth_ids")):
            raise HTTPException(status_code=403, detail="No access scope defined for this user")
        try:
            public_snapshot = _build_public_snapshot(
                assembly_code=assemblyCode,
                db=db,
                current=current,
                include_voters=includeVoters,
                allowed_assembly_ids=scope.get("allowed_assembly_ids") if scope else None,
                allowed_ward_ids=scope.get("allowed_ward_ids") if scope else None,
                allowed_booth_ids=scope.get("allowed_booth_ids") if scope else None,
            )
            snapshot_id = _cache_snapshot(public_snapshot)
            snapshot_url = f"{_external_base_url(request)}{CONTEXT_PATH}/api/voters/snapshot/content/{snapshot_id}"
            payload = api_success("Snapshot fetched successfully", snapshot_url)
            payload["snapshotMode"] = "link"
            return JSONResponse(content=payload, headers={"X-Snapshot-Mode": "link"})
        except Exception as e:
            print("[PUBLIC_SNAPSHOT_FAILED]", str(e))
            local_snapshot = _build_snapshot_from_data(db, assemblyCode, includeVoters, current)
            payload = api_success("Snapshot fetched successfully", local_snapshot)
            payload["snapshotMode"] = "direct"
            return JSONResponse(content=payload, headers={"X-Snapshot-Mode": "direct"})
    except HTTPException as ex:
        return JSONResponse(status_code=ex.status_code, content=api_error("Snapshot failed", str(ex.detail)))
    except ValueError as ex:
        detail = str(ex)
        print("[SNAPSHOT_404]", detail)
        return JSONResponse(status_code=404, content=api_error("No snapshot found", detail))
    except Exception as ex:
        print("[SNAPSHOT_ERROR]", ex)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content=api_error("Snapshot failed", str(ex)))


@app.get(f"{CONTEXT_PATH}/api/voters/snapshot/content/{{snapshot_id}}")
def get_snapshot_content(snapshot_id: str):
    _cleanup_snapshot_cache()
    cached = _snapshot_cache.get(snapshot_id)
    if not cached:
        raise HTTPException(status_code=404, detail="Snapshot link expired or not found")
    return cached["payload"]


@app.get(f"{CONTEXT_PATH}/api/association")
def list_association(boothId: int, db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    rows = (
        db.query(Association)
        .filter(Association.booth_id == boothId, Association.tenant_id == current.tenantId)
        .all()
    )
    dtos = [{"associationId": r.association_id, "associationName": r.association_name} for r in rows]
    return api_success("Associations fetched", dtos)


@app.post(f"{CONTEXT_PATH}/api/association")
def create_association(payload: CreateAssociationRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("ADMIN"))):
    booth = db.query(Booth).filter(Booth.booth_id == payload.boothId).first()
    if not booth:
        raise ValueError("Invalid booth ID")

    association = Association(
        association_name=payload.associationName,
        association_address=payload.associationAddress,
        association_head_name=payload.associationHeadName,
        phone=payload.phone,
        latitude=payload.latitude,
        longitude=payload.longitude,
        booth_id=booth.booth_id,
        tenant_id=current.tenantId,
    )
    db.add(association)
    db.commit()
    db.refresh(association)

    return api_success(
        "Association created",
        {
            "associationId": association.association_id,
            "associationName": association.association_name,
        },
    )


FAMILY_AVAILABILITY_BUCKETS = [
    {"key": "available", "label": "Available", "match": "Available"},
    {"key": "notAvailable", "label": "Not Available", "match": "Not Available"},
    {"key": "entryDenied", "label": "Entry Denied", "match": "Entry Denied"},
    {"key": "dataNotGiven", "label": "Data not Given", "match": "Data not Given"},
    {"key": "doorClosed", "label": "Door Closed", "match": "Door Closed"},
]

FAMILY_DETAIL_EXPORT_FIELDS = [
    ("roadName", "Road Name"),
    ("familyNumber", "Family Number"),
    ("flatNumber", "Flat No"),
    ("buildingNumber", "Building/Apartment No"),
    ("buildingName", "Building/Apartment Name"),
    ("buildingAddress", "Building/Apartment Address"),
    ("tagLeader", "Tag Leader"),
    ("familyAvailability", "Family Availability"),
    ("economicStatus", "Economic Status"),
    ("familyNature", "Family Nature"),
    ("points", "Points"),
    ("phone", "Phone"),
    ("hasAssociation", "Has Association"),
    ("associationName", "Association Name"),
    ("associationHeadName", "Association Head Name"),
    ("associationHeadPhone", "Association Head Phone"),
    ("headName", "Head of Family"),
    ("headEpicNo", "Head EPIC"),
    ("memberCount", "Member Count"),
]


def _resolve_db_user(db: Session, current: JwtUserDetails) -> Optional[Any]:
    return db.query(User).filter(User.first_name == current.firstName, User.phone == current.phone).first()


def _resolve_volunteer_user(db: Session, current: JwtUserDetails) -> Optional[VolunteerUser]:
    if not current.phone:
        return None
    return (
        db.query(VolunteerUser)
        .filter(
            func.lower(VolunteerUser.first_name) == (current.firstName or "").lower(),
            VolunteerUser.phone == current.phone,
        )
        .first()
    )


def _resolve_family_audit_actor(db: Session, current: JwtUserDetails) -> tuple[Optional[int], str]:
    """Return (actor_id, display_name). Volunteer actors use negative ids to avoid User id collisions."""
    user = _resolve_db_user(db, current)
    if user:
        return int(user.id), (user.first_name or current.firstName or "").strip()
    volunteer = _resolve_volunteer_user(db, current)
    if volunteer:
        return -int(volunteer.id), (volunteer.first_name or current.firstName or "").strip()
    return None, (current.firstName or "").strip()


def _apply_family_audit(db: Session, fam: Family, current: JwtUserDetails, is_create: bool = False) -> None:
    actor_id, agent_name = _resolve_family_audit_actor(db, current)
    now = datetime.now(timezone.utc)
    if is_create:
        fam.created_by = actor_id
        fam.created_by_name = agent_name or None
        fam.created_by_phone = current.phone
        fam.created_at = now
    fam.updated_by = actor_id
    fam.updated_by_name = agent_name or None
    fam.updated_by_phone = current.phone
    fam.updated_at = now


def _as_utc_aware(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _family_effective_updated(fam: Family) -> Optional[datetime]:
    raw = fam.updated_at or fam.created_at
    return _as_utc_aware(raw)


def _family_sort_timestamp(fam: Family) -> datetime:
    return _family_effective_updated(fam) or datetime.min.replace(tzinfo=timezone.utc)


def _family_agent_id(fam: Family) -> Optional[int]:
    return fam.updated_by or fam.created_by


def _normalize_agent_phone(phone: Optional[str]) -> str:
    digits = re.sub(r"\D", "", str(phone or "").strip())
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def _family_agent_bucket_key(fam: Family) -> str:
    """Single bucket per real agent — phone first so volunteer id vs legacy name rows merge."""
    phone = _normalize_agent_phone(fam.updated_by_phone or fam.created_by_phone)
    if phone:
        return f"phone:{phone}"
    user_id = _family_agent_id(fam)
    if user_id is not None and int(user_id) != 0:
        uid = int(user_id)
        if uid < 0:
            return f"volunteer:{-uid}"
        return f"user:{uid}"
    name = (fam.updated_by_name or fam.created_by_name or "").strip().lower()
    return f"name:{name or 'unknown'}"


def _family_availability_bucket(fam: Family) -> str:
    val = (fam.family_availability or "").strip()
    for bucket in FAMILY_AVAILABILITY_BUCKETS:
        if val == bucket["match"]:
            return bucket["key"]
    return "other"


def _family_building_key(fam: Family) -> Optional[str]:
    parts = [
        (fam.building_number or "").strip().lower(),
        (fam.building_name or "").strip().lower(),
        (fam.building_address or "").strip().lower(),
    ]
    if not any(parts):
        return None
    return "|".join(parts)


def _parse_family_analysis_date(value: Optional[str], end_of_day: bool = False) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except Exception:
        try:
            parsed = datetime.strptime(value, "%Y-%m-%d")
        except Exception:
            return None
    if end_of_day and parsed.hour == 0 and parsed.minute == 0 and parsed.second == 0 and len(value) <= 10:
        parsed = parsed + timedelta(hours=23, minutes=59, seconds=59)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed


def _family_in_date_range(fam: Family, from_dt: Optional[datetime], to_dt: Optional[datetime]) -> bool:
    ts = _family_effective_updated(fam)
    if not ts:
        return from_dt is None and to_dt is None
    if from_dt and ts < from_dt:
        return False
    if to_dt and ts > to_dt:
        return False
    return True


@dataclass
class _FamilyBoothContext:
    booth_id: int
    ward_id: Optional[int] = None
    ward_code: Optional[str] = None
    booth_no: Optional[str] = None


def _resolve_booth_ids_for_ward_param(db: Session, ward_id: int) -> List[int]:
    booth_ids: set[int] = set()
    ward_codes: set[str] = set()

    ward_cols = _get_table_columns(db, "public", "wards")
    pub_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
    code_col = "ward_code" if "ward_code" in ward_cols else None
    if pub_id_col:
        select_cols = [f"{pub_id_col} AS wid"]
        if code_col:
            select_cols.append(f"{code_col} AS ward_code")
        row = db.execute(
            text(f"SELECT {', '.join(select_cols)} FROM public.wards WHERE {pub_id_col} = :ward_id LIMIT 1"),
            {"ward_id": ward_id},
        ).first()
        if row:
            if getattr(row, "ward_code", None) is not None:
                ward_codes.add(str(row.ward_code))
            booth_cols = _get_table_columns(db, "public", "booths")
            b_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
            b_ward_col = "ward_id" if "ward_id" in booth_cols else None
            if b_id_col and b_ward_col:
                pub_rows = db.execute(
                    text(f"SELECT {b_id_col} AS booth_id FROM public.booths WHERE {b_ward_col} = :ward_id"),
                    {"ward_id": ward_id},
                ).all()
                booth_ids.update(int(r.booth_id) for r in pub_rows if r.booth_id is not None)

    data_ward = db.query(Ward).filter(Ward.ward_id == ward_id).first()
    if data_ward and data_ward.ward_code:
        ward_codes.add(str(data_ward.ward_code))

    for wc in ward_codes:
        for row in db.query(Booth.booth_id).filter(Booth.ward_code == wc).all():
            booth_ids.add(int(row[0]))
        booth_cols = _get_table_columns(db, "public", "booths")
        b_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
        b_code_col = "ward_code" if "ward_code" in booth_cols else None
        if b_id_col and b_code_col:
            pub_rows = db.execute(
                text(f"SELECT {b_id_col} AS booth_id FROM public.booths WHERE {b_code_col} = :ward_code"),
                {"ward_code": wc},
            ).all()
            booth_ids.update(int(r.booth_id) for r in pub_rows if r.booth_id is not None)

    for row in db.query(Booth.booth_id).filter(Booth.ward_id == ward_id).all():
        booth_ids.add(int(row[0]))

    return sorted(booth_ids)


def _assert_booth_in_ward(db: Session, booth_id: int, ward_id: int) -> None:
    allowed = _resolve_booth_ids_for_ward_param(db, int(ward_id))
    if not allowed or int(booth_id) not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"Booth {booth_id} does not belong to ward {ward_id}",
        )


def _super_admin_user_ids(db: Session) -> set:
    rows = db.query(User.id).filter(User.role.in_(["SUPER_ADMIN", "ROLE_SUPER_ADMIN", "ADMIN", "ROLE_ADMIN"])).all()
    return {int(r[0]) for r in rows if r[0] is not None}


HIDDEN_FAMILY_AGENT_NAMES = frozenset({"admin@iswot.io", "admin@iswot.in"})


def _normalize_family_agent_name(name: Optional[str]) -> str:
    return (name or "").strip().lower()


def _family_agent_name_is_hidden(name: Optional[str]) -> bool:
    normalized = _normalize_family_agent_name(name)
    if not normalized:
        return False
    if normalized in HIDDEN_FAMILY_AGENT_NAMES:
        return True
    return normalized.startswith("admin@iswot")


def _family_is_hidden_admin_entry(db: Session, fam: Family, super_admin_ids: Optional[set] = None) -> bool:
    """Hide families touched by super-admin / admin@iswot.io from field volunteer family tables."""
    admin_ids = super_admin_ids if super_admin_ids is not None else _super_admin_user_ids(db)
    for agent_id in (fam.created_by, fam.updated_by, _family_agent_id(fam)):
        if agent_id is not None and int(agent_id) > 0 and int(agent_id) in admin_ids:
            return True
    if _family_agent_name_is_hidden(fam.created_by_name) or _family_agent_name_is_hidden(fam.updated_by_name):
        return True
    return False


def _apply_exclude_admin_family_filter(q, db: Session, super_admin_ids: Optional[set] = None):
    admin_ids = super_admin_ids if super_admin_ids is not None else _super_admin_user_ids(db)
    if admin_ids:
        id_list = list(admin_ids)
        q = q.filter(
            or_(
                Family.created_by.is_(None),
                Family.created_by < 0,
                ~Family.created_by.in_(id_list),
            )
        )
        q = q.filter(
            or_(
                Family.updated_by.is_(None),
                Family.updated_by < 0,
                ~Family.updated_by.in_(id_list),
            )
        )
    for hidden_name in HIDDEN_FAMILY_AGENT_NAMES:
        q = q.filter(
            func.lower(func.coalesce(Family.created_by_name, "")) != hidden_name,
            func.lower(func.coalesce(Family.updated_by_name, "")) != hidden_name,
        )
    return q


def _resolve_booth_ids_for_ward_list(db: Session, ward_ids: Iterable[int]) -> List[int]:
    merged: set[int] = set()
    for ward_id in ward_ids:
        if ward_id is None:
            continue
        merged.update(_resolve_booth_ids_for_ward_param(db, int(ward_id)))
    return sorted(merged)


def _assembly_filter_values(assembly_code: Optional[str]) -> List[str]:
    """String assembly id/code variants (151, 000000000151) for safe TEXT SQL compares."""
    if assembly_code is None:
        return []
    raw = str(assembly_code).strip()
    if not raw:
        return []
    values: List[str] = [raw]
    padded = _normalize_assembly_code(raw) or normalize_assembly_code(raw)
    if padded and padded not in values:
        values.append(padded)
    if padded and str(padded).isdigit():
        unpadded = str(int(str(padded)))
        if unpadded not in values:
            values.append(unpadded)
    return values


def _resolve_public_ward_codes_for_assembly(db: Session, assembly_code: Optional[str]) -> set[str]:
    """Ward codes in public.voter_enrichment scope for the selected assembly context."""
    values = _assembly_filter_values(assembly_code)
    if not values:
        return set()
    ward_cols = _get_table_columns(db, "public", "wards")
    ward_code_col = "ward_code" if "ward_code" in ward_cols else None
    assembly_ref = "assembly_id" if "assembly_id" in ward_cols else ("assembly_no" if "assembly_no" in ward_cols else None)
    if not ward_code_col or not assembly_ref:
        return set()
    clauses = []
    params: Dict[str, Any] = {}
    for index, val in enumerate(values):
        key = f"asm_{index}"
        clauses.append(f"CAST({assembly_ref} AS TEXT) = :{key}")
        params[key] = str(val)
    rows = db.execute(
        text(
            f"""
            SELECT DISTINCT {ward_code_col} AS ward_code
            FROM public.wards
            WHERE ({' OR '.join(clauses)})
              AND {ward_code_col} IS NOT NULL
            """
        ),
        params,
    ).all()
    return {str(row.ward_code).strip() for row in rows if row.ward_code is not None}


def _resolve_booth_ids_for_assembly_param(db: Session, assembly_code: Optional[str]) -> List[int]:
    str_values = _assembly_filter_values(assembly_code)
    if not str_values:
        return []
    asm_ids: List[Any] = list(dict.fromkeys(str_values))
    for val in str_values:
        if val.isdigit():
            asm_ids.append(int(val))
    ward_ids = [
        int(row.ward_id)
        for row in db.query(Ward.ward_id).filter(Ward.assembly_id.in_(asm_ids)).all()
        if row.ward_id is not None
    ]
    if not ward_ids:
        return []
    return _resolve_booth_ids_for_ward_list(db, ward_ids)


def _apply_enrichment_assembly_filter(q, db: Session, assembly_code: Optional[str]):
    if assembly_code is None or not str(assembly_code).strip():
        return q
    codes = _resolve_public_ward_codes_for_assembly(db, assembly_code)
    if not codes:
        return q.filter(text("1=0"))
    return q.filter(VoterEnrichment.ward_code.in_(sorted(codes)))


def _booth_context_for_family(db: Session, fam: Family) -> _FamilyBoothContext:
    booth = db.query(Booth).filter(Booth.booth_id == fam.booth_id).first()
    if booth:
        return _FamilyBoothContext(
            booth_id=int(booth.booth_id),
            ward_id=booth.ward_id,
            ward_code=str(booth.ward_code) if booth.ward_code is not None else None,
            booth_no=str(booth.booth_no) if booth.booth_no is not None else None,
        )

    booth_cols = _get_table_columns(db, "public", "booths")
    b_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
    if not b_id_col:
        return _FamilyBoothContext(booth_id=int(fam.booth_id))

    select_cols = [f"{b_id_col} AS booth_id"]
    for col, alias in (("ward_id", "ward_id"), ("ward_code", "ward_code"), ("booth_no", "booth_no")):
        if col in booth_cols:
            select_cols.append(f"{col} AS {alias}")
    row = db.execute(
        text(f"SELECT {', '.join(select_cols)} FROM public.booths WHERE {b_id_col} = :booth_id LIMIT 1"),
        {"booth_id": fam.booth_id},
    ).first()
    if row:
        return _FamilyBoothContext(
            booth_id=int(getattr(row, "booth_id", fam.booth_id)),
            ward_id=getattr(row, "ward_id", None),
            ward_code=str(row.ward_code) if getattr(row, "ward_code", None) is not None else None,
            booth_no=str(row.booth_no) if getattr(row, "booth_no", None) is not None else None,
        )
    return _FamilyBoothContext(booth_id=int(fam.booth_id))


def _apply_family_list_filters(
    q,
    db: Session,
    current: JwtUserDetails,
    wardId: Optional[int] = None,
    boothId: Optional[int] = None,
    assemblyCode: Optional[str] = None,
):
    role = (current.role or "").replace("ROLE_", "")
    if current.tenantId and role != "SUPER_ADMIN":
        q = q.filter(Family.tenant_id == current.tenantId)

    scope = _resolve_access_scope_ids(db, current)
    if scope:
        allowed_ward_ids = sorted(scope.get("allowed_ward_ids") or [])
        allowed_booth_ids = sorted(scope.get("allowed_booth_ids") or [])
        scope_booth_ids: set[int] = {int(b) for b in allowed_booth_ids if b is not None}
        if allowed_ward_ids:
            scope_booth_ids.update(_resolve_booth_ids_for_ward_list(db, allowed_ward_ids))
        if scope_booth_ids:
            q = q.filter(Family.booth_id.in_(sorted(scope_booth_ids)))
        else:
            q = q.filter(text("1=0"))

    if wardId is not None:
        ward_booth_ids = _resolve_booth_ids_for_ward_param(db, int(wardId))
        if ward_booth_ids:
            q = q.filter(Family.booth_id.in_(ward_booth_ids))
        else:
            q = q.filter(text("1=0"))
    if boothId is not None:
        q = q.filter(Family.booth_id == int(boothId))
    if assemblyCode is not None and str(assemblyCode).strip():
        asm_booth_ids = _resolve_booth_ids_for_assembly_param(db, assemblyCode)
        if asm_booth_ids:
            q = q.filter(Family.booth_id.in_(asm_booth_ids))
        else:
            q = q.filter(text("1=0"))
    return q


def _load_families_for_analysis(
    db: Session,
    current: JwtUserDetails,
    wardId: Optional[int] = None,
    boothId: Optional[int] = None,
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    assemblyCode: Optional[str] = None,
) -> List[tuple]:
    q = db.query(Family).filter(Family.deleted.is_(False))
    q = _apply_family_list_filters(q, db, current, wardId=wardId, boothId=boothId, assemblyCode=assemblyCode)
    q = _apply_exclude_admin_family_filter(q, db)

    from_dt = _parse_family_analysis_date(updatedFrom)
    to_dt = _parse_family_analysis_date(updatedTo, end_of_day=True)
    families = q.all()
    super_admin_ids = _super_admin_user_ids(db)
    rows: List[tuple] = []
    for fam in families:
        if _family_is_hidden_admin_entry(db, fam, super_admin_ids):
            continue
        if from_dt or to_dt:
            if not _family_in_date_range(fam, from_dt, to_dt):
                continue
        rows.append((fam, _booth_context_for_family(db, fam)))
    return rows


def _lookup_master_voter_relation(
    db: Session, epic_no: Optional[str], cache: Optional[Dict[str, tuple]] = None
) -> tuple:
    """Fallback relation name/type from public.voters (rel_eng) or voter_enrichment."""
    epic = normalize_optional_text(epic_no)
    if not epic:
        return "", ""
    rel_cache = cache if cache is not None else {}
    if epic.upper() in rel_cache:
        return rel_cache[epic.upper()]
    relation_name = ""
    relation_type = ""
    try:
        row = db.execute(
            text("SELECT rel_eng, rel_type FROM public.voters WHERE UPPER(epic) = UPPER(:epic) LIMIT 1"),
            {"epic": epic},
        ).mappings().first()
        if row:
            relation_name = (row.get("rel_eng") or "").strip()
            relation_type = (row.get("rel_type") or "").strip()
    except Exception:
        pass
    if not relation_name:
        try:
            enr = db.query(VoterEnrichment).filter(VoterEnrichment.epic == epic).first()
            if enr:
                relation_name = f"{enr.relation_first_middle_name_en or ''} {enr.relation_last_name_en or ''}".strip()
                if not relation_type:
                    relation_type = (enr.relation_type or "").strip()
        except Exception:
            pass
    rel_cache[epic.upper()] = (relation_name, relation_type)
    return relation_name, relation_type


def _family_to_dto(db: Session, fam: Family, rel_cache: Optional[Dict[str, tuple]] = None) -> Dict[str, Any]:
    head_name = ""
    head_epic = ""
    m_dto = []
    try:
        members_data = db.query(FamilyMember, Voter).join(Voter, FamilyMember.voter_id == Voter.voter_id).filter(FamilyMember.family_id == fam.familyId).all()
        for member, voter in members_data:
            full_name = f"{voter.first_middle_name_en or ''} {voter.last_name_en or ''}".strip()
            if member.is_head:
                head_name = full_name
                head_epic = voter.epic_no
            relation_name = f"{voter.relation_first_middle_name_en or ''} {voter.relation_last_name_en or ''}".strip()
            if not relation_name:
                relation_name = f"{voter.relation_first_middle_name_local or ''} {voter.relation_last_name_local or ''}".strip()
            relation_type = (voter.relation_type or "").strip()
            if not relation_name or not relation_type:
                master_name, master_type = _lookup_master_voter_relation(db, voter.epic_no, rel_cache)
                if not relation_name:
                    relation_name = master_name
                if not relation_type:
                    relation_type = master_type
            m_dto.append(
                {
                    "memberId": member.member_id,
                    "head": bool(member.is_head),
                    "epicNo": voter.epic_no,
                    "voterName": full_name,
                    "relationName": relation_name,
                    "relationType": relation_type,
                    "relationFirstMiddleNameEn": voter.relation_first_middle_name_en,
                    "relationLastNameEn": voter.relation_last_name_en,
                    "relationFirstMiddleNameLocal": voter.relation_first_middle_name_local,
                    "relationLastNameLocal": voter.relation_last_name_local,
                    "rel_eng": relation_name,
                }
            )
    except Exception as e:
        print(f"Error in _family_to_dto members: {e}")

    ward_id = None
    ward_code = None
    booth_no = None
    try:
        booth = db.query(Booth).filter(Booth.booth_id == fam.booth_id).first()
        if booth:
            ward_id = booth.ward_id
            ward_code = booth.ward_code
            booth_no = booth.booth_no
    except Exception:
        ward_id = None
        ward_code = None
        booth_no = None

    last_updated = _family_effective_updated(fam)

    return {
        "familyId": fam.familyId,
        "tenantId": fam.tenant_id,
        "familyName": fam.family_name,
        "familyAddress": fam.family_address,
        "buildingName": fam.building_name,
        "buildingAddress": fam.building_address,
        "hasAssociation": fam.has_association,
        "associationName": fam.association_name,
        "associationHeadName": fam.association_head_name,
        "associationHeadPhone": fam.association_head_phone,
        "phone": fam.phone,
        "points": fam.points,
        "pointsProvided": fam.points_provided,
        "latitude": fam.latitude,
        "longitude": fam.longitude,
        "boothId": fam.booth_id,
        "boothNo": str(booth_no) if booth_no is not None else None,
        "wardId": ward_id,
        "wardCode": str(ward_code) if ward_code is not None else None,
        "createdAt": fam.created_at.isoformat() if fam.created_at else None,
        "lastUpdatedAt": last_updated.isoformat() if last_updated else None,
        "updatedByName": fam.updated_by_name,
        "updatedByPhone": fam.updated_by_phone,
        "agentName": fam.updated_by_name or fam.created_by_name,
        "agentPhone": fam.updated_by_phone or fam.created_by_phone,
        "associationId": fam.association_id,
        "headMemberId": fam.head_voter_id,
        "headName": head_name,
        "headEpicNo": head_epic,
        "memberCount": len(m_dto),
        "members": m_dto,
        "economicStatus": fam.economic_status,
        "familyNature": fam.family_nature,
        "roadName": fam.road_name,
        "buildingNumber": fam.building_number,
        "flatNumber": fam.flat_number,
        "familyNumber": fam.family_number,
        "tagLeader": fam.tag_leader,
        "familyAvailability": fam.family_availability,
    }


def _bootstrap_booth(db: Session, booth_id: int, tenant_id: str):
    # Check if booth exists in data.booths
    booth = db.query(Booth).filter(Booth.booth_id == booth_id).first()
    if booth:
        return booth

    # Try to find in public.booths
    res = db.execute(text("SELECT * FROM public.booths WHERE id = :bid AND (tenant_id = :tid OR tenant_id IS NULL)"), {"bid": booth_id, "tid": tenant_id}).first()
    if not res:
        return None

    # Try to find ward info in public.wards
    ward_info = db.execute(text("SELECT id, assembly_no, ward_name_en, ward_name_local FROM public.wards WHERE ward_code = :wc AND (tenant_id = :tid OR tenant_id IS NULL)"), {"wc": res.ward_code, "tid": tenant_id}).first()
    assembly_id = ward_info.assembly_no if ward_info and ward_info.assembly_no else 1
    ward_id = ward_info.id if ward_info else res.ward_id

    # Assembly
    assembly = db.query(Assembly).filter(Assembly.assembly_id == assembly_id).first()
    if not assembly:
        assembly = Assembly(
            assembly_id=assembly_id, # Explicit ID
            tenant_id=tenant_id,
            assembly_name_en=f"Assembly {assembly_id}",
            assembly_code=str(assembly_id)
        )
        db.add(assembly)
        db.flush()

    # Ward
    ward = db.query(Ward).filter(Ward.ward_id == ward_id).first()
    if not ward:
        ward = Ward(
            ward_id=ward_id, # Explicit ID
            assembly_id=assembly.assembly_id,
            tenant_id=tenant_id,
            ward_name_en=ward_info.ward_name_en if ward_info else res.ward_code,
            ward_name_local=ward_info.ward_name_local if ward_info else None,
            ward_code=res.ward_code
        )
        db.add(ward)
        db.flush()

    # Create Booth in data schema
    booth = Booth(
        booth_id=res.id, # Explicit ID
        ward_id=ward.ward_id,
        tenant_id=tenant_id,
        polling_station_adr_en=getattr(res, 'booth_add_en', ''),
        polling_station_adr_local=getattr(res, 'booth_add_local', ''),
        booth_no=res.booth_no,
        ward_code=res.ward_code
    )
    db.add(booth)
    db.flush()
    return booth

def _bootstrap_voter(db: Session, epic: str, booth: Booth, tenant_id: str):
    # Check in data.voters
    voters = db.query(Voter).filter(Voter.epic_no == epic, Voter.tenant_id == tenant_id).first()
    if voters:
        return voters

    # Find in public.voters
    res = db.execute(text("SELECT * FROM public.voters WHERE epic = :epic AND (tenant_id = :tid OR tenant_id IS NULL)"), {"epic": epic, "tid": tenant_id}).first()
    if not res:
        return None

    # Generate a voter_id since data.voters has no serial
    next_id = db.execute(text("SELECT COALESCE(MAX(voter_id), 0) + 1 FROM data.voters")).scalar()

    # Create Voter in data schema
    v = Voter(
        voter_id=next_id, # Explicit ID
        tenant_id=tenant_id,
        booth_id=booth.booth_id,
        epic_no=res.epic,
        sr_no=int(res.sl) if res.sl and res.sl.isdigit() else 0,
        first_middle_name_en=res.name_en,
        gender=res.gender,
        age=int(res.age) if res.age and res.age.isdigit() else 0,
        house_no_en=res.house,
        mobile=res.mobile
    )
    db.add(v)
    db.flush()
    return v


@app.post(f"{CONTEXT_PATH}/api/family")
def create_family(payload: CreateFamilyRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "ASSEMBLY", "WARD", "USER"))):
    tenant_id = current.tenantId
    # If admin/superadmin with no tenant, we need to infer it or use booth tenant
    # But since bootstrap needs a tenant, we'll try to find one from existing data or use a default "T1"
    if not tenant_id:
        # Fallback to a default tenant if none exists in current user
        # In a real system, you'd pick the relevant tenant for the booth.
        tenant_id = "T1" 

    booth = _bootstrap_booth(db, payload.boothId, tenant_id)
    if not booth:
        raise ValueError("Invalid booth")

    if payload.wardId is not None:
        _assert_booth_in_ward(db, int(payload.boothId), int(payload.wardId))

    association = None
    if payload.associationId is not None:
        association = db.query(Association).filter(Association.association_id == payload.associationId).first()
        if not association:
            raise ValueError("Invalid association")

    fam = Family(
        family_name=payload.familyName,
        family_address=payload.familyAddress,
        road_name=payload.roadName,
        building_number=payload.buildingNumber,
        building_name=payload.buildingName,
        flat_number=payload.flatNumber,
        family_number=payload.familyNumber,
        tag_leader=payload.tagLeader,
        family_availability=payload.familyAvailability,
        building_address=payload.buildingAddress,
        has_association=payload.hasAssociation,
        association_name=payload.associationName,
        association_head_name=payload.associationHeadName,
        association_head_phone=payload.associationHeadPhone,
        phone=payload.phone,
        points=payload.points,
        points_provided=payload.pointsProvided,
        latitude=payload.latitude,
        longitude=payload.longitude,
        economic_status=payload.economicStatus,
        family_nature=payload.familyNature,
        tenant_id=current.tenantId or booth.tenant_id,
        booth_id=booth.booth_id,
        association_id=association.association_id if association else None,
        deleted=False,
    )
    db.add(fam)
    db.flush()
    _apply_family_audit(db, fam, current, is_create=True)

    head_member_id = None
    for epic in payload.memberEpicNos:
        # Ensure voter exists in data schema
        voter = _bootstrap_voter(db, epic, booth, tenant_id)
        
        if not voter:
            raise ValueError(f"Voter not found: {epic}")

        is_head = payload.headEpicNo == epic
        member = FamilyMember(family_id=fam.familyId, voter_id=voter.voter_id, is_head=is_head)
        db.add(member)
        db.flush()
        if is_head:
            head_member_id = member.member_id

    fam.head_voter_id = head_member_id
    _apply_family_audit(db, fam, current, is_create=False)
    db.add(fam)
    db.commit()
    db.refresh(fam)

    return api_success("Family created", _family_to_dto(db, fam))


@app.put(f"{CONTEXT_PATH}/api/family/{{familyId}}")
def update_family(familyId: int, payload: UpdateFamilyRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "ASSEMBLY", "WARD", "USER"))):
    fam = db.query(Family).filter(Family.familyId == familyId).first()
    if not fam:
        raise ValueError(f"Family not found: {familyId}")

    if payload.wardId is not None:
        _assert_booth_in_ward(db, int(payload.boothId), int(payload.wardId))

    fam.family_name = payload.familyName
    fam.family_address = payload.familyAddress
    fam.road_name = payload.roadName
    fam.building_number = payload.buildingNumber
    fam.building_name = payload.buildingName
    fam.flat_number = payload.flatNumber
    fam.family_number = payload.familyNumber
    fam.tag_leader = payload.tagLeader
    fam.family_availability = payload.familyAvailability
    fam.building_address = payload.buildingAddress
    fam.has_association = payload.hasAssociation
    fam.association_name = payload.associationName
    fam.association_head_name = payload.associationHeadName
    fam.association_head_phone = payload.associationHeadPhone
    fam.phone = payload.phone
    fam.points = payload.points
    fam.points_provided = payload.pointsProvided
    fam.latitude = payload.latitude
    fam.longitude = payload.longitude
    fam.economic_status = payload.economicStatus
    fam.family_nature = payload.familyNature

    if payload.associationId is not None:
        association = db.query(Association).filter(Association.association_id == payload.associationId).first()
        if not association:
            raise ValueError("Invalid association")
        fam.association_id = association.association_id
    else:
        fam.association_id = None

    existing = db.query(FamilyMember).join(Voter, FamilyMember.voter_id == Voter.voter_id).filter(FamilyMember.family_id == fam.familyId).all()
    keep = set(payload.memberEpicNos)

    for m in existing:
        voter = db.query(Voter).filter(Voter.voter_id == m.voter_id).first()
        if voter and voter.epic_no not in keep:
            db.delete(m)

    db.flush()

    all_members = db.query(FamilyMember).join(Voter, FamilyMember.voter_id == Voter.voter_id).filter(FamilyMember.family_id == fam.familyId).all()
    existing_epic = {
        db.query(Voter).filter(Voter.voter_id == m.voter_id).first().epic_no: m
        for m in all_members
    }

    for epic in payload.memberEpicNos:
        if epic not in existing_epic:
            voter = _bootstrap_voter(db, epic, booth, fam.tenant_id)
            if not voter:
                raise ValueError(f"Voter not found: {epic}")
            member = FamilyMember(family_id=fam.familyId, voter_id=voter.voter_id, is_head=(epic == payload.headEpicNo))
            db.add(member)

    db.flush()

    members_after = db.query(FamilyMember).join(Voter, FamilyMember.voter_id == Voter.voter_id).filter(FamilyMember.family_id == fam.familyId).all()
    head_member_id = None
    for m in members_after:
        voter = db.query(Voter).filter(Voter.voter_id == m.voter_id).first()
        is_head = bool(voter and voter.epic_no == payload.headEpicNo)
        m.is_head = is_head
        if is_head:
            head_member_id = m.member_id

    fam.head_voter_id = head_member_id
    _apply_family_audit(db, fam, current, is_create=False)
    db.add(fam)
    db.commit()
    db.refresh(fam)

    return api_success("Family updated", _family_to_dto(db, fam))


@app.get(f"{CONTEXT_PATH}/api/family")
def list_families(
    boothId: Optional[int] = None,
    wardId: Optional[int] = None,
    page: int = 0,
    size: int = 10,
    search: Optional[str] = None,
    association: Optional[str] = None,
    assemblyCode: Optional[str] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "ASSEMBLY", "WARD", "USER")),
):
    q = db.query(Family).filter(Family.deleted.is_(False))
    q = _apply_family_list_filters(q, db, current, wardId=wardId, boothId=boothId, assemblyCode=assemblyCode)
    q = _apply_exclude_admin_family_filter(q, db)

    association_filter = parse_optional_bool(association)
    if association_filter is not None:
        if association_filter:
            q = q.filter(Family.association_id.is_not(None))
        else:
            q = q.filter(Family.association_id.is_(None))

    if search and search.strip():
        s = f"%{search.lower().strip()}%"
        q = (
            q.outerjoin(FamilyMember, FamilyMember.family_id == Family.familyId)
            .outerjoin(Voter, FamilyMember.voter_id == Voter.voter_id)
            .filter(
                or_(
                    func.lower(Family.family_name).like(s),
                    func.lower(Family.family_address).like(s),
                    func.lower(Voter.first_middle_name_en).like(s),
                    func.lower(Voter.epic_no).like(s),
                )
            )
            .distinct()
        )

    total = q.count()
    families = q.offset(page * size).limit(size).all()
    rel_cache: Dict[str, tuple] = {}
    return build_page([_family_to_dto(db, f, rel_cache) for f in families], page, size, total)


@app.get(f"{CONTEXT_PATH}/api/families/analysis")
def families_analysis(
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "ASSEMBLY", "WARD", "USER")),
    wardId: Optional[int] = None,
    boothId: Optional[int] = None,
    mode: Optional[str] = "agent",
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    assemblyCode: Optional[str] = None,
):
    rows = _load_families_for_analysis(
        db, current, wardId=wardId, boothId=boothId, updatedFrom=updatedFrom, updatedTo=updatedTo, assemblyCode=assemblyCode
    )
    if not rows:
        return api_success("Family analysis fetched", {"fields": FAMILY_AVAILABILITY_BUCKETS, "rows": [], "mode": (mode or "agent").lower()})

    analysis_fields = [{"key": item["key"], "label": item["label"]} for item in FAMILY_AVAILABILITY_BUCKETS]
    mode_key = (mode or "agent").lower()

    def _init_counts() -> Dict[str, int]:
        return {item["key"]: 0 for item in FAMILY_AVAILABILITY_BUCKETS}

    if mode_key == "agent":
        super_admin_ids = _super_admin_user_ids(db)
        counters: Dict[str, Dict[str, Any]] = {}
        for fam, booth in rows:
            if _family_is_hidden_admin_entry(db, fam, super_admin_ids):
                continue
            user_id = _family_agent_id(fam)
            if user_id is not None and int(user_id) > 0 and int(user_id) in super_admin_ids:
                continue
            bucket_key = _family_agent_bucket_key(fam)
            bucket = counters.setdefault(
                bucket_key,
                {
                    "userId": int(user_id) if user_id is not None else 0,
                    "counts": _init_counts(),
                    "buildings": set(),
                    "totalFamilies": 0,
                    "agentName": fam.updated_by_name or fam.created_by_name,
                    "phone": fam.updated_by_phone or fam.created_by_phone,
                    "lastUpdatedAt": None,
                },
            )
            bucket["totalFamilies"] += 1
            building_key = _family_building_key(fam)
            if building_key:
                bucket["buildings"].add(building_key)
            avail_key = _family_availability_bucket(fam)
            if avail_key in bucket["counts"]:
                bucket["counts"][avail_key] += 1
            ts = _family_effective_updated(fam)
            if ts and (not bucket.get("lastUpdatedAt") or ts > bucket["lastUpdatedAt"]):
                bucket["lastUpdatedAt"] = ts

        positive_user_ids = sorted({b["userId"] for b in counters.values() if b.get("userId", 0) > 0})
        volunteer_ids = sorted({-b["userId"] for b in counters.values() if b.get("userId", 0) < 0})
        phone_keys = sorted(
            {
                key.split(":", 1)[1]
                for key in counters.keys()
                if key.startswith("phone:") and key.split(":", 1)[1]
            }
        )
        user_rows = db.query(User).filter(User.id.in_(positive_user_ids)).all() if positive_user_ids else []
        volunteer_rows = db.query(VolunteerUser).filter(VolunteerUser.id.in_(volunteer_ids)).all() if volunteer_ids else []
        if phone_keys:
            phone_filters = [VolunteerUser.phone.in_(phone_keys)]
            phone_filters.extend(VolunteerUser.phone.like(f"%{phone}") for phone in phone_keys)
            volunteer_rows = list(
                {
                    v.id: v
                    for v in (
                        volunteer_rows + db.query(VolunteerUser).filter(or_(*phone_filters)).all()
                    )
                }.values()
            )
        user_map = {u.id: u for u in user_rows}
        volunteer_map = {v.id: v for v in volunteer_rows}
        volunteer_by_phone = {_normalize_agent_phone(v.phone): v for v in volunteer_rows if v.phone}
        results = []
        for bucket_key, bucket in counters.items():
            user_id = int(bucket.get("userId") or 0)
            user = user_map.get(user_id) if user_id > 0 else None
            volunteer = volunteer_map.get(-user_id) if user_id < 0 else None
            if bucket_key.startswith("phone:"):
                volunteer = volunteer or volunteer_by_phone.get(bucket_key.split(":", 1)[1])
                if volunteer:
                    user_id = -int(volunteer.id)
            agent_name = (
                user.first_name
                if user
                else volunteer.first_name
                if volunteer
                else (bucket.get("agentName") or ("Unknown Agent" if user_id == 0 else f"User {user_id}"))
            )
            display_phone = (
                user.phone
                if user
                else volunteer.phone
                if volunteer
                else (bucket.get("phone") or (bucket_key.split(":", 1)[1] if bucket_key.startswith("phone:") else ""))
            )
            results.append(
                {
                    "userId": user_id,
                    "agentName": agent_name,
                    "phone": display_phone or "",
                    "totalBuildings": len(bucket["buildings"]),
                    "totalFamilies": bucket["totalFamilies"],
                    "counts": bucket["counts"],
                    "lastUpdatedAt": bucket.get("lastUpdatedAt").isoformat() if bucket.get("lastUpdatedAt") else None,
                }
            )
        results.sort(key=lambda item: item.get("agentName") or "")
        return api_success("Family analysis fetched", {"fields": analysis_fields, "rows": results, "mode": mode_key})

    ward_name_map: Dict[str, str] = {}
    if mode_key == "ward":
        ward_cols = _get_table_columns(db, "public", "wards")
        ward_code_col = "ward_code" if "ward_code" in ward_cols else None
        ward_name_col = (
            "ward_name_en"
            if "ward_name_en" in ward_cols
            else ("name_en" if "name_en" in ward_cols else ("ward_name_local" if "ward_name_local" in ward_cols else None))
        )
        if ward_code_col and ward_name_col:
            t_clause, t_params = _build_public_tenant_filter(current)
            ward_rows = db.execute(
                text(
                    f"""
                    SELECT {ward_code_col} AS ward_code, {ward_name_col} AS ward_name
                    FROM public.wards
                    WHERE 1=1 {t_clause}
                    """
                ),
                t_params,
            ).all()
            ward_name_map = {
                str(r.ward_code): str(r.ward_name)
                for r in ward_rows
                if r.ward_code is not None and r.ward_name is not None
            }

    group_buckets: Dict[str, Dict[str, Any]] = {}
    for fam, booth in rows:
        if mode_key == "date":
            ts = _family_effective_updated(fam)
            if not ts:
                continue
            group_key = ts.date().isoformat()
            group_label = group_key
        elif mode_key == "ward":
            group_key = str(booth.ward_code or booth.ward_id or "")
            if not group_key:
                continue
            group_label = ward_name_map.get(group_key) or f"Ward {group_key}"
        elif mode_key == "booth":
            group_key = str(booth.booth_no or fam.booth_id or "")
            if not group_key:
                continue
            group_label = f"Booth {group_key}"
        else:
            group_key = "all"
            group_label = "All"

        bucket = group_buckets.setdefault(
            group_key,
            {
                "groupKey": group_key,
                "label": group_label,
                "counts": _init_counts(),
                "buildings": set(),
                "totalFamilies": 0,
                "agents": set(),
                "booths": set(),
                "lastUpdatedAt": None,
            },
        )
        bucket["totalFamilies"] += 1
        building_key = _family_building_key(fam)
        if building_key:
            bucket["buildings"].add(building_key)
        agent_id = _family_agent_id(fam)
        if agent_id:
            bucket["agents"].add(agent_id)
        if booth.booth_no:
            bucket["booths"].add(str(booth.booth_no))
        avail_key = _family_availability_bucket(fam)
        if avail_key in bucket["counts"]:
            bucket["counts"][avail_key] += 1
        ts = _family_effective_updated(fam)
        if ts and (not bucket.get("lastUpdatedAt") or ts > bucket["lastUpdatedAt"]):
            bucket["lastUpdatedAt"] = ts

    grouped_rows = []
    for bucket in group_buckets.values():
        grouped_rows.append(
            {
                "groupKey": bucket["groupKey"],
                "label": bucket["label"],
                "totalBuildings": len(bucket["buildings"]),
                "totalFamilies": bucket["totalFamilies"],
                "agentsWorked": len(bucket["agents"]),
                "boothsCovered": len(bucket["booths"]),
                "counts": bucket["counts"],
                "lastUpdatedAt": bucket.get("lastUpdatedAt").isoformat() if bucket.get("lastUpdatedAt") else None,
            }
        )
    grouped_rows.sort(key=lambda item: item.get("label") or "")
    return api_success("Family analysis fetched", {"fields": analysis_fields, "rows": grouped_rows, "mode": mode_key})


@app.get(f"{CONTEXT_PATH}/api/families/map-points")
def families_map_points(
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "ASSEMBLY", "WARD", "USER")),
    wardId: Optional[int] = None,
    boothId: Optional[int] = None,
    assemblyCode: Optional[str] = None,
):
    """Families with coordinates and enriched member relation fields for map tooltips."""
    rows = _load_families_for_analysis(db, current, wardId=wardId, boothId=boothId, assemblyCode=assemblyCode)
    rel_cache: Dict[str, tuple] = {}
    points: List[Dict[str, Any]] = []
    for fam, _booth in rows:
        dto = _family_to_dto(db, fam, rel_cache)
        lat = dto.get("latitude")
        lng = dto.get("longitude")
        if lat is None or lng is None:
            continue
        try:
            lat_f = float(lat)
            lng_f = float(lng)
        except (TypeError, ValueError):
            continue
        if not (lat_f == lat_f and lng_f == lng_f):  # NaN check without math import
            continue
        points.append(
            {
                "familyId": dto.get("familyId"),
                "latitude": lat_f,
                "longitude": lng_f,
                "familyName": dto.get("familyName"),
                "familyAvailability": dto.get("familyAvailability") or "Available",
                "roadName": dto.get("roadName"),
                "buildingNumber": dto.get("buildingNumber"),
                "buildingName": dto.get("buildingName"),
                "familyNumber": dto.get("familyNumber"),
                "flatNumber": dto.get("flatNumber"),
                "boothNo": dto.get("boothNo"),
                "wardId": dto.get("wardId"),
                "wardCode": dto.get("wardCode"),
                "members": dto.get("members") or [],
            }
        )
    return api_success("Family map points fetched", points)


@app.get(f"{CONTEXT_PATH}/api/families/details")
def families_details(
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "ASSEMBLY", "WARD", "USER")),
    wardId: Optional[int] = None,
    boothId: Optional[int] = None,
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    page: Optional[int] = None,
    size: Optional[int] = None,
    assemblyCode: Optional[str] = None,
):
    rows = _load_families_for_analysis(
        db, current, wardId=wardId, boothId=boothId, updatedFrom=updatedFrom, updatedTo=updatedTo, assemblyCode=assemblyCode
    )
    rows.sort(key=lambda item: _family_sort_timestamp(item[0]), reverse=True)

    if page is not None and size is not None:
        start = page * size
        page_rows = rows[start : start + size]
        start_idx = start + 1
    else:
        page_rows = rows
        start_idx = 1

    detail_fields = [{"key": key, "label": label} for key, label in FAMILY_DETAIL_EXPORT_FIELDS]
    result_rows: List[Dict[str, Any]] = []
    rel_cache: Dict[str, tuple] = {}
    for idx, (fam, booth) in enumerate(page_rows, start=start_idx):
        dto = _family_to_dto(db, fam, rel_cache)
        ordered: Dict[str, Any] = {
            "familyId": dto.get("familyId"),
            "serialNumber": idx,
            "familyName": dto.get("familyName"),
            "boothNo": dto.get("boothNo") or (str(booth.booth_no) if booth.booth_no is not None else None),
            "latitude": dto.get("latitude"),
            "longitude": dto.get("longitude"),
            "lastUpdatedAt": dto.get("lastUpdatedAt"),
            "members": dto.get("members") or [],
        }
        for key, _label in FAMILY_DETAIL_EXPORT_FIELDS:
            ordered[key] = dto.get(key)
        result_rows.append(ordered)

    return api_success(
        "Family details fetched",
        {"fields": detail_fields, "rows": result_rows, "total": len(rows)},
    )


@app.get(f"{CONTEXT_PATH}/api/family/suggestions")
def family_suggestions(
    type: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "ASSEMBLY", "WARD", "USER")),
):
    field = normalize_optional_text(type).lower()
    if field == "building":
        q = db.query(Family.building_name).filter(Family.building_name.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", [r[0] for r in results if r[0]])

    elif field == "association":
        # From Families
        q_fam = db.query(Family.association_name).filter(Family.association_name.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q_fam = q_fam.filter(Family.tenant_id == current.tenantId)
        
        # From Associations table
        q_assoc = db.query(Association.association_name)
        if current.tenantId:
            q_assoc = q_assoc.filter(Association.tenant_id == current.tenantId)
            
        final_names = set()
        for r in q_fam.distinct().all():
            if r[0]: final_names.add(r[0])
        for r in q_assoc.distinct().all():
            if r[0]: final_names.add(r[0])
            
        return api_success("Suggestions fetched", sorted(list(final_names))[:100])

    elif field == "road":
        q = db.query(Family.road_name).filter(Family.road_name.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", [r[0] for r in results if r[0]])

    elif field == "leader":
        q = db.query(Family.tag_leader).filter(Family.tag_leader.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", [r[0] for r in results if r[0]])

    elif field in ("family", "familyname", "name"):
        q = db.query(Family.family_name).filter(Family.family_name.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", sorted({r[0] for r in results if r[0]})[:100])

    elif field in ("buildingnumber", "building_number"):
        q = db.query(Family.building_number).filter(Family.building_number.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", [r[0] for r in results if r[0]])

    elif field == "flat":
        q = db.query(Family.flat_number).filter(Family.flat_number.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", [r[0] for r in results if r[0]])

    elif field in ("address", "buildingaddress"):
        q = db.query(Family.building_address).filter(Family.building_address.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", [r[0] for r in results if r[0]])

    elif field in ("associationhead", "association_head"):
        q = db.query(Family.association_head_name).filter(Family.association_head_name.isnot(None), Family.deleted.is_(False))
        if current.tenantId:
            q = q.filter(Family.tenant_id == current.tenantId)
        results = q.distinct().limit(100).all()
        return api_success("Suggestions fetched", [r[0] for r in results if r[0]])
    
    return api_success("Suggestions fetched", [])


def _template_to_dto(tpl: MessageTemplate) -> Dict[str, Any]:
    return {
        "templateId": tpl.template_id,
        "tenantId": tpl.tenant_id,
        "wardId": tpl.ward_id,
        "channel": tpl.channel,
        "authorityName": tpl.authority_name,
        "electionName": tpl.election_name,
        "assemblyLabel": tpl.assembly_label,
        "wardLabel": tpl.ward_label,
        "candidateName": tpl.candidate_name,
        "candidateParty": tpl.candidate_party,
        "candidateWardLabel": tpl.candidate_ward_label,
        "voteDate": tpl.vote_date,
        "voteTime": tpl.vote_time,
        "socialLink": tpl.social_link,
        "boothLocationLink": tpl.booth_location_link,
        "bannerUrl": tpl.banner_url,
        "showLogo": bool(tpl.show_logo) if tpl.show_logo is not None else True,
        "enabled": bool(tpl.enabled),
        "updatedAt": tpl.updated_at.isoformat() if tpl.updated_at else None,
    }


@app.get(f"{CONTEXT_PATH}/api/message-template")
def get_message_template(
    wardId: Optional[str] = None,
    channel: str = "WHATSAPP",
    epicNo: Optional[str] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER")),
):
    channel = channel.upper()
    resolved_ward_id = wardId
    
    # If epicNo provided, try to resolve wardId from voter data
    if epicNo and (not resolved_ward_id or resolved_ward_id == "null" or resolved_ward_id == ""):
        try:
            voter_cols = _get_table_columns(db, "public", "voters")
            booth_cols = _get_table_columns(db, "public", "booths")
            
            epic_col = "epic_no" if "epic_no" in voter_cols else ("voter_id" if "voter_id" in voter_cols else None)
            booth_no_col = "booth_no" if "booth_no" in voter_cols else None
            
            if epic_col and booth_no_col:
                voter_row = db.execute(text(f"SELECT {booth_no_col} as booth_no FROM public.voters WHERE {epic_col} = :epic"), {"epic": epicNo}).first()
                if voter_row:
                    b_no = voter_row.booth_no
                    b_ward_id_col = "ward_id" if "ward_id" in booth_cols else None
                    if b_ward_id_col:
                        # CRITICAL: Scope by tenant_id to avoid matching booth 1 from a different assembly
                        booth_row = db.execute(text(f"SELECT {b_ward_id_col} as ward_id FROM public.booths WHERE booth_no = :bno AND tenant_id = :tid"), {"bno": b_no, "tid": current.tenantId}).first()
                        if booth_row:
                            resolved_ward_id = str(booth_row.ward_id)
        except Exception as e:
            print(f"Ward resolution error for EPIC {epicNo}: {e}")

    q = db.query(MessageTemplate).filter(
        MessageTemplate.tenant_id == current.tenantId,
        MessageTemplate.channel == channel
    )
    if resolved_ward_id and resolved_ward_id != "" and resolved_ward_id != "null":
        try:
            q = q.filter(MessageTemplate.ward_id == int(resolved_ward_id))
        except (ValueError, TypeError):
             q = q.filter(MessageTemplate.ward_id.is_(None))
    else:
        q = q.filter(MessageTemplate.ward_id.is_(None))
    
    tpl = q.first()
    return api_success("Message template fetched", _template_to_dto(tpl) if tpl else None)


@app.get(f"{CONTEXT_PATH}/api/message-template/activated-wards")
def get_activated_wards(
    assemblyId: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER")),
):
    q = db.query(MessageTemplate).filter(
        MessageTemplate.tenant_id == current.tenantId, 
        MessageTemplate.enabled == True
    )
    
    if assemblyId:
        # Include Global (ward_id is NULL) OR Wards belonging to this assembly
        q = q.outerjoin(Ward, MessageTemplate.ward_id == Ward.ward_id).filter(
            or_(MessageTemplate.ward_id.is_(None), Ward.assembly_id == assemblyId)
        )

    templates = q.all()
    
    # Pre-fetch ward details for labeling
    ward_ids = [t.ward_id for t in templates if t.ward_id is not None]
    wards_map = {}
    if ward_ids:
        wards = db.query(Ward).filter(Ward.ward_id.in_(ward_ids)).all()
        wards_map = {w.ward_id: w for w in wards}

    out = []
    for t in templates:
        w_obj = wards_map.get(t.ward_id) if t.ward_id else None
        out.append({
            "wardId": t.ward_id,
            "wardLabel": t.ward_label or (w_obj.ward_name_en if w_obj else None),
            "wardNameEn": w_obj.ward_name_en if w_obj else None,
            "channel": t.channel,
        })
    return api_success("Activated wards fetched", out)


@app.put(f"{CONTEXT_PATH}/api/message-template")
def save_message_template(
    payload: MessageTemplatePayload,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN")),
):
    ward_id = payload.wardId
    channel = (payload.channel or "WHATSAPP").upper()
    q = db.query(MessageTemplate).filter(
        MessageTemplate.tenant_id == current.tenantId,
        MessageTemplate.ward_id == ward_id,
        MessageTemplate.channel == channel,
    )
    tpl = q.first()
    if not tpl:
        tpl = MessageTemplate(tenant_id=current.tenantId, ward_id=ward_id, channel=channel)
        db.add(tpl)
    tpl.authority_name = payload.authorityName
    tpl.election_name = payload.electionName
    tpl.assembly_label = payload.assemblyLabel
    tpl.ward_label = payload.wardLabel
    tpl.candidate_name = payload.candidateName
    tpl.candidate_party = payload.candidateParty
    tpl.candidate_ward_label = payload.candidateWardLabel
    tpl.vote_date = payload.voteDate
    tpl.vote_time = payload.voteTime
    tpl.social_link = payload.socialLink
    tpl.booth_location_link = payload.boothLocationLink
    if payload.bannerUrl:
        tpl.banner_url = payload.bannerUrl
    if payload.showLogo is not None:
        tpl.show_logo = payload.showLogo
    if payload.enabled is not None:
        tpl.enabled = payload.enabled
    tpl.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(tpl)
    return api_success("Message template saved", _template_to_dto(tpl))


@app.post(f"{CONTEXT_PATH}/api/message-template/banner")
def upload_message_template_banner(
    wardId: Optional[int] = Query(None),
    channel: str = "WHATSAPP",
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN")),
):
    raw = file.file.read()
    ext = Path(file.filename or "").suffix or ".jpg"
    key = f"message_banners/{current.tenantId}/{wardId}/{uuid.uuid4().hex}{ext}"
    url = s3_upload_bytes(raw, file.content_type or "application/octet-stream", key)
    channel = channel.upper()
    q = db.query(MessageTemplate).filter(MessageTemplate.tenant_id == current.tenantId, MessageTemplate.channel == channel)
    if wardId is not None and wardId != "":
        try:
            q = q.filter(MessageTemplate.ward_id == int(wardId))
        except (ValueError, TypeError):
            pass
    else:
        q = q.filter(MessageTemplate.ward_id.is_(None))
    
    tpl = q.first()
    if not tpl:
        tpl = MessageTemplate(tenant_id=current.tenantId, ward_id=wardId if wardId else None, channel=channel)
        db.add(tpl)
    tpl.banner_url = url
    tpl.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(tpl)
    return api_success("Banner uploaded", _template_to_dto(tpl))


@app.get(f"{CONTEXT_PATH}/api/voter-activation/verify")
def verify_voter_activation(
    epicNo: str,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "USER")),
):
    """
    Check if a specific voter (by EPIC) has messaging enabled based on their resolved ward
    and the current activation status of that ward.  Uses ORM so it hits the correct
    'data' schema — the same source of truth as the Promotions tab.
    """
    # Normalize EPIC (Upper case + Strip)
    normalized_epic = epicNo.strip().upper()
    
    # Step 1: resolve the voter's ward_id using ORM (data.voters → data.booths → data.wards)
    resolved_ward_id: Optional[int] = None

    voter_obj = db.query(Voter).filter(Voter.epic_no == normalized_epic).first()
    if voter_obj and voter_obj.booth_id:
        booth_obj = db.query(Booth).filter(Booth.booth_id == voter_obj.booth_id).first()
        if booth_obj:
            resolved_ward_id = booth_obj.ward_id

    # Step 2: check activation for each channel against metastore.message_templates
    def is_active(channel: str) -> bool:
        # a) Global template (ward_id IS NULL)
        global_tpl = db.query(MessageTemplate).filter(
            MessageTemplate.tenant_id == current.tenantId,
            MessageTemplate.channel == channel,
            MessageTemplate.ward_id.is_(None),
        ).first()
        if global_tpl and global_tpl.enabled:
            return True

        # b) Ward-specific template
        if resolved_ward_id is not None:
            ward_tpl = db.query(MessageTemplate).filter(
                MessageTemplate.tenant_id == current.tenantId,
                MessageTemplate.channel == channel,
                MessageTemplate.ward_id == resolved_ward_id,
            ).first()
            if ward_tpl and ward_tpl.enabled:
                return True

        return False

    return api_success("Voter activation status", {
        "epicNo": epicNo,
        "wardId": resolved_ward_id,
        "whatsapp": is_active("WHATSAPP"),
        "sms": is_active("SMS"),
        "print": is_active("PRINT"),
    })


@app.post(f"{CONTEXT_PATH}/api/message-template/deactivate-all")
def deactivate_all_templates(
    channel: str = "WHATSAPP",
    tenantId: Optional[str] = None,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN")),
):
    channel = channel.upper()
    target_tenant = tenantId or current.tenantId
    if not target_tenant and current.role != "SUPER_ADMIN":
         raise HTTPException(status_code=400, detail="Tenant ID is required")
         
    try:
        q = db.query(MessageTemplate).filter(MessageTemplate.channel == channel)
        if target_tenant:
            q = q.filter(MessageTemplate.tenant_id == target_tenant)
            
        rows = q.update({MessageTemplate.enabled: False}, synchronize_session=False)
        db.commit()
        return api_success(f"Deactivated {channel} for {rows} wards.")
    except Exception as e:
        db.rollback()
        print(f"Deactivate all error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get(f"{CONTEXT_PATH}/api/family/{{id}}")
def get_family(id: int, db: Session = Depends(get_db), _: JwtUserDetails = Depends(require_roles("ADMIN", "USER"))):
    fam = db.query(Family).filter(Family.familyId == id).first()
    if not fam:
        raise ValueError(f"Family not found: {id}")
    return _family_to_dto(db, fam)


@app.delete(f"{CONTEXT_PATH}/api/family/{{id}}")
def delete_family(id: int, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("ADMIN"))):
    fam = db.query(Family).filter(Family.familyId == id).first()
    if not fam:
        raise ValueError(f"Family not found: {id}")
    if fam.tenant_id != current.tenantId:
        raise ValueError("Unauthorized: Family does not belong to your tenant")

    fam.deleted = True
    db.commit()
    return api_success("Family deleted", {})


@app.post(f"{CONTEXT_PATH}/api/meetings")
def create_meeting(payload: MeetingCreateRequest, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN"))):
    t_id = current.tenantId
    if not t_id:
        # Fallback for global admins: try to find the first tenant
        first_tenant = db.query(Tenant).first()
        if first_tenant:
            t_id = first_tenant.tenant_id
        else:
            t_id = "000001" # Very safe fallback

    meeting = Meeting(
        tenant_id=t_id,
        title=payload.title,
        start_time=payload.start_time,
        end_time=payload.end_time,
        latitude=payload.latitude,
        longitude=payload.longitude,
        radius=payload.radius,
        recipients=payload.recipients,
        channels=payload.channels,
    )
    db.add(meeting)
    db.commit()
    return api_success("Meeting created", {"id": meeting.meeting_id})


@app.get(f"{CONTEXT_PATH}/api/meetings")
def list_meetings(db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "WARD", "BOOTH", "USER"))):
    q = db.query(Meeting)
    if current.tenantId:
        q = q.filter(Meeting.tenant_id == current.tenantId)
    meetings = q.all()
    out = []
    for m in meetings:
        out.append({
            "id": m.meeting_id,
            "title": m.title,
            "dateTime": m.start_time,
            "description": f"End: {m.end_time}" if m.end_time else "",
            "latitude": m.latitude or 0.0,
            "longitude": m.longitude or 0.0,
            "radius": m.radius or 0,
            "recipients": m.recipients,
            "channels": m.channels,
        })
    return out


@app.post(f"{CONTEXT_PATH}/api/meetings/{{id}}/attendance")
def record_meeting_attendance(id: int, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN", "ADMIN", "WARD", "BOOTH"))):
    meeting = db.query(Meeting).filter(Meeting.meeting_id == id).first()
    if not meeting or not meeting.latitude or not meeting.longitude or not meeting.radius:
        raise ValueError("Meeting not found or lacks location/radius")
    
    # Approximate degree conversion for radius
    # 1 degree lat is ~111km. 1 degree lon is ~111km * cos(lat)
    import math
    lat_deg = meeting.radius / 111320.0
    lon_deg = meeting.radius / (111320.0 * math.cos(math.radians(meeting.latitude)))
    
    q = db.query(Voter).filter(
        Voter.tenant_id == meeting.tenant_id,
        Voter.latitude.between(meeting.latitude - lat_deg, meeting.latitude + lat_deg),
        Voter.longitude.between(meeting.longitude - lon_deg, meeting.longitude + lon_deg)
    )
    
    voters = q.all()
    count = 0
    for v in voters:
        # Fine-grained distance check
        dist_deg = math.sqrt((v.latitude - meeting.latitude)**2 + (v.longitude - meeting.longitude)**2)
        # Using a simple average for degree to meter conversion here for speed
        if dist_deg * 111000 <= meeting.radius:
            att = db.query(MeetingAttendance).filter(MeetingAttendance.meeting_id == meeting.meeting_id, MeetingAttendance.voter_id == v.voter_id).first()
            if not att:
                att = MeetingAttendance(
                    meeting_id=meeting.meeting_id,
                    voter_id=v.voter_id,
                    distance=dist_deg * 111000,
                )
                db.add(att)
                count += 1
    db.commit()
    return api_success("Attendance recorded", {"added": count})


@app.post(f"{CONTEXT_PATH}/api/meetings/{{id}}/attend-self")
def attend_meeting_self(id: int, lat: Optional[float] = None, lng: Optional[float] = None, db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    meeting = db.query(Meeting).filter(Meeting.meeting_id == id).first()
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    existing = db.query(MeetingAttendance).filter(
        MeetingAttendance.meeting_id == id,
        MeetingAttendance.volunteer_phone == current.phone
    ).first()
    if existing:
        return api_success("Already attended", {})
    
    # Try to link to a voter record by phone (fuzzy match last 10 digits)
    v_phone_tail = current.phone[-10:] if current.phone and len(current.phone) >= 10 else current.phone
    voter = db.query(Voter).filter(Voter.mobile.like(f"%{v_phone_tail}")).first()
    
    distance = None
    if lat is not None and lng is not None and meeting.latitude is not None and meeting.longitude is not None:
        import math
        dist_deg = math.sqrt((lat - meeting.latitude)**2 + (lng - meeting.longitude)**2)
        distance = dist_deg * 111000

    att = MeetingAttendance(
        meeting_id=id,
        voter_id=voter.voter_id if voter else None,
        volunteer_name=current.firstName,
        volunteer_phone=current.phone,
        distance=distance,
        attended_at=func.now()
    )
    db.add(att)
    db.commit()
    return api_success("Self attendance recorded", {})


@app.get(f"{CONTEXT_PATH}/api/meetings/{{id}}/attendance")
def list_meeting_attendance(id: int, db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN"))):
    attendances = db.query(MeetingAttendance).filter(MeetingAttendance.meeting_id == id).all()
    out = []
    for att in attendances:
        v_name = att.volunteer_name
        v_phone = att.volunteer_phone
        v_epic = "-"
        
        # Priority 1: Linked voter_id
        if att.voter_id:
            v = db.query(Voter).filter(Voter.voter_id == att.voter_id).first()
            if v:
                v_name = f"{v.first_middle_name_en or ''} {v.last_name_en or ''}".strip()
                v_phone = v.mobile
                v_epic = v.epic_no
        # Priority 2: Lookup by volunteer_phone if EPIC is still missing
        elif v_phone:
            v_phone_tail = v_phone[-10:] if len(v_phone) >= 10 else v_phone
            v = db.query(Voter).filter(Voter.mobile.like(f"%{v_phone_tail}")).first()
            if v:
                if not v_name: v_name = f"{v.first_middle_name_en or ''} {v.last_name_en or ''}".strip()
                v_epic = v.epic_no

        out.append({
            "id": att.attendance_id,
            "name": v_name,
            "epic": v_epic,
            "phone": v_phone,
            "distance": att.distance,
            "at": att.attended_at.isoformat() if att.attended_at else ""
        })
    return out


@app.get(f"{CONTEXT_PATH}/api/volunteers/dropdown")
def volunteer_dropdown(level: str, parentId: Optional[int] = None, db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    level = level.upper()
    out = []

    if level == "ASSEMBLY":
        query = db.query(Assembly.id if hasattr(Assembly, "id") else Assembly.assembly_id, Assembly.assembly_name_en)
        if current.tenantId is not None and current.role != "SUPER_ADMIN":
            query = query.filter(Assembly.tenant_id == current.tenantId)
        rows = query.all()
        
        # Super Admins should always see the full list of assemblies with correct names from public schema
        if not rows or current.role == "SUPER_ADMIN":
            assembly_cols = _ensure_public_assembly_code(db)
            id_col = "id" if "id" in assembly_cols else ("assembly_id" if "assembly_id" in assembly_cols else ("assembly_no" if "assembly_no" in assembly_cols else None))
            name_col = "assembly_name_en" if "assembly_name_en" in assembly_cols else ("name_en" if "name_en" in assembly_cols else ("assembly_name_local" if "assembly_name_local" in assembly_cols else ("name_kannada" if "name_kannada" in assembly_cols else None)))
            code_expr = "assembly_code" if "assembly_code" in assembly_cols else "NULL"
            if id_col and name_col:
                t_clause, t_params = _build_public_tenant_filter(current)
                public_rows = db.execute(
                    text(f"SELECT {id_col} AS id, {name_col} AS name, {code_expr} AS code FROM public.assembly WHERE 1=1 {t_clause} ORDER BY {name_col}"),
                    t_params,
                ).all()
                # If public has more rows OR we are super admin (to get better names like KR Pura), use public
                if not rows or len(public_rows) >= len(rows) or current.role == "SUPER_ADMIN":
                    for row in public_rows:
                        out.append({"id": int(row.id), "code": String(row.code or row.id), "name": row.name})
                    return api_success("Assemblies fetched", out)
        for row in rows:
            out.append({"id": row[0], "code": String(row[0]), "name": row[1]})
        return api_success("Assemblies fetched", out)
    elif level == "WARD":
        if parentId is None:
            raise ValueError("assemblyId is required for WARD")
        query = db.query(Ward.ward_id, Ward.ward_name_en).filter(Ward.assembly_id == parentId)
        if current.tenantId is not None and current.role != "SUPER_ADMIN":
            query = query.filter(Ward.tenant_id == current.tenantId)
        rows = query.order_by(Ward.ward_name_en).all()

        if not rows or current.role == "SUPER_ADMIN":
            ward_cols = _get_table_columns(db, "public", "wards")
            id_col = "ward_id" if "ward_id" in ward_cols else ("ward_no" if "ward_no" in ward_cols else ("id" if "id" in ward_cols else None))
            name_col = "ward_name_en" if "ward_name_en" in ward_cols else ("name_en" if "name_en" in ward_cols else ("ward_name_local" if "ward_name_local" in ward_cols else ("name_kannada" if "name_kannada" in ward_cols else None)))
            assembly_ref = "assembly_id" if "assembly_id" in ward_cols else ("assembly_no" if "assembly_no" in ward_cols else None)
            if id_col and name_col:
                t_clause, t_params = _build_public_tenant_filter(current)
                where = f"WHERE {assembly_ref} = :assembly_id {t_clause}" if assembly_ref else (f"WHERE 1=1 {t_clause}")
                public_rows = db.execute(
                    text(f"SELECT {id_col} AS id, {name_col} AS name FROM public.wards {where} ORDER BY {name_col}"),
                    {"assembly_id": parentId, **t_params},
                ).all()
                if not rows or len(public_rows) > len(rows):
                    for row in public_rows:
                        out.append({"id": int(row.id), "code": str(row.id), "name": row.name})
                    return out
        for row in rows:
            out.append({"id": row[0], "code": str(row[0]), "name": row[1]})
    elif level == "BOOTH":
        if parentId is None:
            raise ValueError("wardId is required for BOOTH")
        query = db.query(Booth.booth_id, Booth.polling_station_adr_en).filter(Booth.ward_id == parentId)
        if current.tenantId is not None and current.role != "SUPER_ADMIN":
            query = query.filter(Booth.tenant_id == current.tenantId)
        rows = query.order_by(Booth.polling_station_adr_en).all()

        if not rows or current.role == "SUPER_ADMIN":
            booth_cols = _get_table_columns(db, "public", "booths")
            id_col = "booth_id" if "booth_id" in booth_cols else ("booth_no" if "booth_no" in booth_cols else ("id" if "id" in booth_cols else None))
            name_col = "polling_station_adr_en" if "polling_station_adr_en" in booth_cols else ("booth_add_en" if "booth_add_en" in booth_cols else ("name_en" if "name_en" in booth_cols else None))
            ward_ref = "ward_id" if "ward_id" in booth_cols else ("ward_no" if "ward_no" in booth_cols else None)
            if id_col and name_col:
                t_clause, t_params = _build_public_tenant_filter(current)
                where = f"WHERE {ward_ref} = :ward_id {t_clause}" if ward_ref else (f"WHERE 1=1 {t_clause}")
                public_rows = db.execute(
                    text(f"SELECT {id_col} AS id, {name_col} AS name FROM public.booths {where} ORDER BY {name_col}"),
                    {"ward_id": parentId, **t_params},
                ).all()
                if not rows or len(public_rows) > len(rows):
                    for row in public_rows:
                        out.append({"id": int(row.id), "code": str(row.id), "name": row.name})
                    return out
        for row in rows:
            out.append({"id": row[0], "code": str(row[0]), "name": row[1]})
    else:
        raise ValueError("Invalid level")

    return out


@app.get(f"{CONTEXT_PATH}/api/volunteers/stats")
def volunteer_stats(level: Optional[str] = None, id: Optional[int] = None, db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    if level is not None and id is None:
        raise ValueError("assignmentId is required when assignmentType is provided")

    q = (
        db.query(
            func.count(User.id),
            func.coalesce(func.sum(case((User.blocked.is_(False), 1), else_=0)), 0),
            func.coalesce(func.sum(case((User.blocked.is_(True), 1), else_=0)), 0),
        )
        .join(Tenant, User.tenant_ref == Tenant.id)
        .filter(Tenant.tenant_id == current.tenantId, User.role == "USER")
    )

    if level is not None:
        q = q.filter(User.assignment_type == level)
    if id is not None:
        q = q.filter(User.assignment_id == id)

    total, active, inactive = q.one()
    total = int(total or 0)
    active = int(active or 0)
    inactive = int(inactive or 0)
    pending = 0

    def pct(v: int, t: int) -> int:
        return round((v * 100.0) / t) if t else 0

    return {
        "total": total,
        "active": {"count": active, "percentage": pct(active, total)},
        "pending": {"count": pending, "percentage": pct(pending, total)},
        "inactive": {"count": inactive, "percentage": pct(inactive, total)},
        "lastUpdated": datetime.now(timezone.utc).isoformat(),
    }


@app.get(f"{CONTEXT_PATH}/api/voters/stats/gender")
def voter_gender_stats(db: Session = Depends(get_db), current: JwtUserDetails = Depends(get_current_user)):
    total = db.query(func.count(Voter.voter_id)).filter(Voter.tenant_id == current.tenantId).scalar() or 0

    rows = db.query(Voter.gender, func.count(Voter.voter_id)).filter(Voter.tenant_id == current.tenantId).group_by(Voter.gender).all()
    male = female = other = 0

    for gender, count in rows:
        g = (gender or "").upper()
        if g == "M":
            male = int(count)
        elif g == "F":
            female = int(count)
        else:
            other += int(count)

    def pct(v: int, t: int) -> float:
        return round((v * 10000.0 / t), 0) / 100 if t else 0.0

    return api_success(
        "Voter gender statistics fetched",
        {
            "totalVoters": int(total),
            "maleCount": male,
            "malePercentage": pct(male, int(total)),
            "femaleCount": female,
            "femalePercentage": pct(female, int(total)),
            "otherCount": other,
            "otherPercentage": pct(other, int(total)),
        },
    )


def _build_voter_map(v: Voter) -> Dict[str, Any]:
    return {
        "voterId": v.voter_id,
        "srNo": v.sr_no,
        "epicNo": v.epic_no,
        "firstMiddleNameEn": v.first_middle_name_en,
        "lastNameEn": v.last_name_en,
        "firstMiddleNameLocal": v.first_middle_name_local,
        "lastNameLocal": v.last_name_local,
        "relationType": v.relation_type,
        "relationFirstMiddleNameEn": v.relation_first_middle_name_en,
        "relationLastNameEn": v.relation_last_name_en,
        "relationFirstMiddleNameLocal": v.relation_first_middle_name_local,
        "relationLastNameLocal": v.relation_last_name_local,
        "houseNoEn": v.house_no_en,
        "houseNoLocal": v.house_no_local,
        "gender": v.gender,
        "age": v.age,
        "dob": v.dob.isoformat() if v.dob else None,
        "mobile": v.mobile,
        "addressEn": v.address_en,
        "addressLocal": v.address_local,
        "status": v.status,
        "community": v.community,
        "caste": v.caste,
        "residenceType": v.residence_type,
        "civicIssue": v.civic_issue,
        "motherTongue": v.mother_tongue,
        "team": v.team,
        "ownership": v.ownership,
        "education": v.education,
        "natureOfVoter": v.nature_of_voter,
        "latitude": v.latitude,
        "longitude": v.longitude,
    }

_schema_cache: Dict[str, set[str]] = {}

def _get_table_columns(db: Session, schema: str, table: str) -> set[str]:
    key = f"{schema}.{table}"
    if key in _schema_cache and _schema_cache[key]:
        return _schema_cache[key]
    try:
        rows = db.execute(
            text(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = :schema AND table_name = :table
                """
            ),
            {"schema": schema, "table": table},
        ).all()
        cols = {r[0] for r in rows}
        if cols:
            _schema_cache[key] = cols
        return cols
    except Exception as e:
        print(f"[SCHEMA_ERROR] Failed to fetch columns for {key}: {e}")
        return set()


def _get_public_voter_column_map(voter_cols: set[str]) -> Dict[str, str]:
    field_map: Dict[str, str] = {}
    if "mobile" in voter_cols:
        field_map["mobile"] = "mobile"
    if "gender" in voter_cols:
        field_map["gender"] = "gender"
    if "age" in voter_cols:
        field_map["age"] = "age"
    if "house" in voter_cols:
        field_map["houseNoEn"] = "house"
        field_map["houseNoLocal"] = "house"
    if "name_en" in voter_cols:
        field_map["firstMiddleNameEn"] = "name_en"
    if "name_kannada" in voter_cols:
        field_map["firstMiddleNameLocal"] = "name_kannada"
    if "rel_eng" in voter_cols:
        field_map["relationFirstMiddleNameEn"] = "rel_eng"
    if "rel_kannada" in voter_cols:
        field_map["relationFirstMiddleNameLocal"] = "rel_kannada"
    if "rel_type" in voter_cols:
        field_map["relationType"] = "rel_type"
    return field_map


VOTER_ENRICHMENT_FIELD_MAP: Dict[str, str] = {
    "firstMiddleNameEn": "first_middle_name_en",
    "lastNameEn": "last_name_en",
    "firstMiddleNameLocal": "first_middle_name_local",
    "lastNameLocal": "last_name_local",
    "relationType": "relation_type",
    "relationFirstMiddleNameEn": "relation_first_middle_name_en",
    "relationLastNameEn": "relation_last_name_en",
    "relationFirstMiddleNameLocal": "relation_first_middle_name_local",
    "relationLastNameLocal": "relation_last_name_local",
    "houseNoEn": "house_no_en",
    "houseNoLocal": "house_no_local",
    "gender": "gender",
    "age": "age",
    "dob": "dob",
    "mobile": "mobile",
    "addressEn": "address_en",
    "addressLocal": "address_local",
    "status": "status",
    "community": "community",
    "caste": "caste",
    "residenceType": "residence_type",
    "civicIssue": "civic_issue",
    "motherTongue": "mother_tongue",
    "team": "team",
    "ownership": "ownership",
    "education": "education",
    "natureOfVoter": "nature_of_voter",
    "voterPoints": "voter_points",
    "govtSchemeTracking": "govt_scheme_tracking",
    "engagementPotential": "engagement_potential",
    "ifShifted": "if_shifted",
    "notes": "notes",
    "presentAddress": "present_address",
    "newWard": "new_ward",
    "newBoothNo": "new_booth_no",
    "newSerialNo": "new_serial_no",
    "notAvailableReason": "not_available_reason",
    "latitude": "latitude",
    "longitude": "longitude",
}
VOTER_ENRICHMENT_JSON_FIELDS = {"govtSchemeTracking"}
VOTER_ENRICHMENT_INT_FIELDS = {"age"}
VOTER_ENRICHMENT_FLOAT_FIELDS = {"latitude", "longitude"}


def _parse_updated_fields(raw: Optional[str]) -> set[str]:
    if not raw:
        return set()
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return {str(item) for item in parsed if str(item).strip()}
    except Exception:
        return set()
    return set()


def _serialize_enrichment_value(api_field: str, value: Any) -> Any:
    if api_field in VOTER_ENRICHMENT_JSON_FIELDS:
        return json.dumps(value or [])
    if api_field in VOTER_ENRICHMENT_INT_FIELDS:
        if value in (None, ""):
            return None
        return int(value)
    if api_field in VOTER_ENRICHMENT_FLOAT_FIELDS:
        if value in (None, ""):
            return None
        return float(value)
    if isinstance(value, str):
        return normalize_optional_text(value)
    return value


def _ensure_user_from_volunteer(db: Session, volunteer: VolunteerUser) -> Optional[User]:
    if not volunteer:
        return None
    existing = db.query(User).filter(User.first_name == volunteer.first_name, User.phone == volunteer.phone).first()
    if existing:
        return existing

    tenant_ref = None
    if volunteer.tenant_id:
        tenant = db.query(Tenant).filter(Tenant.tenant_id == volunteer.tenant_id).first()
        tenant_ref = tenant.id if tenant else None

    assignment_id = None
    if volunteer.assignment_id and str(volunteer.assignment_id).isdigit():
        assignment_id = int(str(volunteer.assignment_id))

    user = User(
        first_name=volunteer.first_name,
        phone=volunteer.phone,
        role=volunteer.role or "USER",
        tenant_ref=tenant_ref,
        assignment_type=volunteer.assignment_type,
        assignment_id=assignment_id,
        blocked=bool(volunteer.blocked),
        deleted=bool(volunteer.deleted),
    )
    db.add(user)
    db.flush()
    return user


def _deserialize_enrichment_value(api_field: str, value: Any) -> Any:
    if api_field in VOTER_ENRICHMENT_JSON_FIELDS:
        if value in (None, ""):
            return []
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, list) else []
        except Exception:
            return []
    return value


def _build_voter_enrichment_payload(enrichment: VoterEnrichment) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "epicNo": enrichment.epic,
        "wardCode": enrichment.ward_code,
        "boothNo": enrichment.booth_no,
        "updatedFields": sorted(_parse_updated_fields(enrichment.updated_fields)),
        "updatedByName": enrichment.updated_by_name,
        "updatedByPhone": enrichment.updated_by_phone,
    }
    for api_field, column_name in VOTER_ENRICHMENT_FIELD_MAP.items():
        payload[api_field] = _deserialize_enrichment_value(api_field, getattr(enrichment, column_name))
    return payload


def _get_voter_enrichments(db: Session, epics: Iterable[str]) -> Dict[str, Dict[str, Any]]:
    epic_list = sorted({str(epic).strip() for epic in epics if epic and str(epic).strip()})
    if not epic_list:
        return {}
    rows = db.query(VoterEnrichment).filter(VoterEnrichment.epic.in_(epic_list)).all()
    return {row.epic: _build_voter_enrichment_payload(row) for row in rows}


def _merge_voter_payload_with_enrichment(voter_payload: Dict[str, Any], enrichment: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not enrichment:
        return voter_payload
    updated_fields = set(enrichment.get("updatedFields") or [])
    for api_field in updated_fields:
        if api_field in enrichment:
            voter_payload[api_field] = enrichment.get(api_field)
    if enrichment.get("wardCode") and not voter_payload.get("wardCode"):
        voter_payload["wardCode"] = enrichment.get("wardCode")
    if enrichment.get("boothNo") and not voter_payload.get("boothNo"):
        voter_payload["boothNo"] = enrichment.get("boothNo")
    return voter_payload


def _merge_voter_payloads_with_enrichment(db: Session, voters: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    enrichments = _get_voter_enrichments(db, [v.get("epicNo") for v in voters])
    for voter in voters:
        _merge_voter_payload_with_enrichment(voter, enrichments.get(voter.get("epicNo")))
    return voters


def _build_public_voter_result(row: Any) -> Dict[str, Any]:
    return {
        "epicNo": row.epic,
        "wardCode": str(row.ward_code) if row.ward_code is not None else None,
        "boothNo": str(row.booth_no) if row.booth_no is not None else None,
        "srNo": row.sl,
        "houseNoEn": row.house,
        "houseNoLocal": row.house,
        "firstMiddleNameEn": row.name_en,
        "lastNameEn": "",
        "firstMiddleNameLocal": row.name_kannada,
        "lastNameLocal": "",
        "relationFirstMiddleNameEn": row.rel_eng,
        "relationLastNameEn": "",
        "relationFirstMiddleNameLocal": row.rel_kannada,
        "relationLastNameLocal": "",
        "relationType": row.rel_type,
        "gender": row.gender,
        "age": row.age,
        "mobile": row.mobile,
    }


def _ensure_public_assembly_code(db: Session) -> set[str]:
    assembly_cols = _get_table_columns(db, "public", "assembly")
    if not assembly_cols:
        raise ValueError("public.assembly table not found")
    if "assembly_code" in assembly_cols:
        return assembly_cols

    assembly_no_col = "assembly_no" if "assembly_no" in assembly_cols else ("assembly_id" if "assembly_id" in assembly_cols else None)
    if not assembly_no_col:
        raise ValueError("public.assembly missing assembly_code and assembly_no/assembly_id")

    try:
        db.execute(text("ALTER TABLE public.assembly ADD COLUMN IF NOT EXISTS assembly_code VARCHAR(12)"))
        db.execute(
            text(
                f"""
                UPDATE public.assembly
                SET assembly_code = LPAD(CAST({assembly_no_col} AS TEXT), 12, '0')
                WHERE assembly_code IS NULL OR TRIM(assembly_code) = ''
                """
            )
        )
        db.commit()
        return _get_table_columns(db, "public", "assembly")
    except Exception:
        db.rollback()
        return assembly_cols


def _build_public_snapshot(
    assembly_code: str,
    db: Session,
    current: JwtUserDetails,
    include_voters: bool = True,
    allowed_assembly_ids: Optional[set[int]] = None,
    allowed_ward_ids: Optional[set[int]] = None,
    allowed_booth_ids: Optional[set[int]] = None,
) -> Dict[str, Any]:
    allowed_assembly_ids = {v for v in (allowed_assembly_ids or set()) if v is not None}
    allowed_ward_ids = {v for v in (allowed_ward_ids or set()) if v is not None}
    allowed_booth_ids = {v for v in (allowed_booth_ids or set()) if v is not None}
    assembly_cols = _ensure_public_assembly_code(db)
    booth_cols = _get_table_columns(db, "public", "booths")
    voter_cols = _get_table_columns(db, "public", "voters")
    ward_cols = _get_table_columns(db, "public", "wards")

    requested_assembly_code = normalize_assembly_code(assembly_code)
    assembly_pk_col = "id" if "id" in assembly_cols else ("assembly_id" if "assembly_id" in assembly_cols else None)
    assembly_no_col = "assembly_no" if "assembly_no" in assembly_cols else ("assembly_id" if "assembly_id" in assembly_cols else None)
    assembly_name_en_col = "assembly_name_en" if "assembly_name_en" in assembly_cols else ("name_en" if "name_en" in assembly_cols else None)
    assembly_name_local_col = "assembly_name_local" if "assembly_name_local" in assembly_cols else ("name_kannada" if "name_kannada" in assembly_cols else None)
    assembly_code_expr = "assembly_code" if "assembly_code" in assembly_cols else (
        f"LPAD(CAST({assembly_no_col} AS TEXT), 12, '0')" if assembly_no_col else "NULL"
    )

    # Lookup the assembly row in the public schema
    # We don't filter by tenant here because we use the assembly assignment (scope) to grant access later.
    assembly_row = db.execute(
        text(
            f"""
            SELECT
                {assembly_code_expr} AS assembly_code,
                {assembly_pk_col if assembly_pk_col else 'NULL'} AS assembly_pk,
                {assembly_no_col if assembly_no_col else 'NULL'} AS assembly_no,
                {assembly_name_en_col if assembly_name_en_col else 'NULL'} AS assembly_name_en,
                {assembly_name_local_col if assembly_name_local_col else 'NULL'} AS assembly_name_local
            FROM public.assembly
            WHERE ({assembly_code_expr} = :assembly_code
               OR CAST({assembly_no_col if assembly_no_col else assembly_code_expr} AS TEXT) = :assembly_no_text)
            LIMIT 1
            """
        ),
        {
            "assembly_code": requested_assembly_code,
            "assembly_no_text": str(int(requested_assembly_code)) if requested_assembly_code and requested_assembly_code.isdigit() else requested_assembly_code,
        },
    ).first()

    # Fallback if assembly row is missing but we have wards for it
    if not assembly_row:
        # Create a synthetic assembly row from the requested code
        class SyntheticRow:
            def __init__(self, code):
                self.assembly_code = code
                self.assembly_pk = int(code) if code and code.isdigit() else None
                self.assembly_no = self.assembly_pk
                self.assembly_name_en = f"Assembly {code}"
                self.assembly_name_local = None
        assembly_row = SyntheticRow(requested_assembly_code)
    if allowed_assembly_ids:
        assembly_pk = assembly_row.assembly_pk if assembly_row.assembly_pk is not None else assembly_row.assembly_no
        if assembly_pk not in allowed_assembly_ids and assembly_row.assembly_no not in allowed_assembly_ids:
            raise HTTPException(status_code=403, detail="Access denied for requested assembly")

    booth_id_col = "id" if "id" in booth_cols else ("booth_id" if "booth_id" in booth_cols else None)
    booth_ward_id_col = "ward_id" if "ward_id" in booth_cols else None
    booth_no_col = "booth_no" if "booth_no" in booth_cols else (booth_id_col or "id")
    booth_ward_code_col = "ward_code" if "ward_code" in booth_cols else None
    booth_name_en_col = "booth_add_en" if "booth_add_en" in booth_cols else ("polling_station_adr_en" if "polling_station_adr_en" in booth_cols else None)
    booth_name_local_col = "booth_add_local" if "booth_add_local" in booth_cols else ("polling_station_adr_local" if "polling_station_adr_local" in booth_cols else None)
    
    # Ward columns needed for booth filtering
    ward_id_col = "id" if "id" in ward_cols else ("ward_id" if "ward_id" in ward_cols else None)
    ward_code_col = "ward_code" if "ward_code" in ward_cols else ("ward_no" if "ward_no" in ward_cols else None)
    ward_name_en_col = "ward_name_en" if "ward_name_en" in ward_cols else ("name_en" if "name_en" in ward_cols else None)
    ward_name_local_col = "ward_name_local" if "ward_name_local" in ward_cols else ("name_kannada" if "name_kannada" in ward_cols else None)
    ward_assembly_id_col = "assembly_id" if "assembly_id" in ward_cols else None
    ward_assembly_no_col = "assembly_no" if "assembly_no" in ward_cols else None
    ward_assembly_code_col = "assembly_code" if "assembly_code" in ward_cols else None

    booth_rows = []
    if booth_id_col and booth_ward_id_col:
        booth_filters: List[str] = []
        booth_params: Dict[str, Any] = {}
        
        # Crucial: Filter booths by the requested assembly to avoid massive data transfer
        if ward_id_col and (ward_assembly_id_col or ward_assembly_no_col or ward_assembly_code_col):
            # We join with wards to restrict to the requested assembly
            booth_filters.append(f"{booth_ward_id_col} IN (SELECT {ward_id_col} FROM public.wards WHERE " + 
                                (f"{ward_assembly_id_col} = :assembly_pk" if ward_assembly_id_col and assembly_row.assembly_pk else
                                 (f"{ward_assembly_no_col} = :assembly_no" if ward_assembly_no_col and assembly_row.assembly_no else
                                  f"{ward_assembly_code_col} = :assembly_code")) + ")")
            if ward_assembly_id_col and assembly_row.assembly_pk:
                booth_params["assembly_pk"] = assembly_row.assembly_pk
            elif ward_assembly_no_col and assembly_row.assembly_no:
                booth_params["assembly_no"] = assembly_row.assembly_no
            else:
                booth_params["assembly_code"] = assembly_row.assembly_code or requested_assembly_code

        if allowed_booth_ids:
            clause, params = _build_in_clause(booth_id_col, sorted(allowed_booth_ids), "scope_booth")
            booth_filters.append(clause)
            booth_params.update(params)
        if allowed_ward_ids:
            clause, params = _build_in_clause(booth_ward_id_col, sorted(allowed_ward_ids), "scope_ward")
            booth_filters.append(clause)
            booth_params.update(params)
        
        if current.role != "SUPER_ADMIN" and "tenant_id" in booth_cols:
            if current.tenantId:
                booth_filters.append("(tenant_id = :tid OR tenant_id IS NULL)")
                booth_params["tid"] = current.tenantId
            else:
                pass

        booth_where = f"WHERE {' AND '.join(booth_filters)}" if booth_filters else ""
        booth_rows = db.execute(
            text(
                f"""
                SELECT
                    {booth_id_col} AS booth_id,
                    {booth_ward_id_col} AS ward_id,
                    {booth_no_col} AS booth_no,
                    {booth_ward_code_col if booth_ward_code_col else 'NULL'} AS ward_code,
                    {booth_name_en_col if booth_name_en_col else 'NULL'} AS booth_name_en,
                    {booth_name_local_col if booth_name_local_col else 'NULL'} AS booth_name_local
                FROM public.booths
                {booth_where}
                ORDER BY {booth_ward_id_col}, {booth_no_col}
                """
            ),
            booth_params,
        ).all()


    ward_map: Dict[Any, Dict[str, Any]] = {}
    if ward_id_col:
        ward_filters: List[str] = []
        ward_params: Dict[str, Any] = {}
        if ward_assembly_id_col and assembly_row.assembly_pk is not None:
            ward_filters.append(f"{ward_assembly_id_col} = :assembly_pk")
            ward_params["assembly_pk"] = assembly_row.assembly_pk
        elif ward_assembly_no_col and assembly_row.assembly_no is not None:
            ward_filters.append(f"{ward_assembly_no_col} = :assembly_no")
            ward_params["assembly_no"] = assembly_row.assembly_no
        elif ward_assembly_code_col:
            ward_filters.append(f"{ward_assembly_code_col} = :assembly_code")
            ward_params["assembly_code"] = assembly_row.assembly_code or requested_assembly_code
        if allowed_ward_ids:
            clause, params = _build_in_clause(ward_id_col, sorted(allowed_ward_ids), "scope_ward")
            ward_filters.append(clause)
            ward_params.update(params)

        if current.role != "SUPER_ADMIN" and "tenant_id" in ward_cols:
            if current.tenantId:
                ward_filters.append("(tenant_id = :tid OR tenant_id IS NULL)")
                ward_params["tid"] = current.tenantId
            else:
                # If user has no tenant, they can see everything in the requested assembly
                pass

        ward_where = f"WHERE {' AND '.join(ward_filters)}" if ward_filters else ""

        ward_rows = db.execute(
            text(
                f"""
                SELECT
                    {ward_id_col} AS ward_id,
                    {ward_code_col if ward_code_col else 'NULL'} AS ward_code,
                    {ward_name_en_col if ward_name_en_col else 'NULL'} AS ward_name_en,
                    {ward_name_local_col if ward_name_local_col else 'NULL'} AS ward_name_local
                FROM public.wards
                {ward_where}
                """
            ),
            ward_params,
        ).all()
        for w in ward_rows:
            ward_map[w.ward_id] = {
                "wardId": w.ward_id,
                "wardCode": str(w.ward_code) if w.ward_code is not None else None,
                "wardNameEn": w.ward_name_en or (f"Ward {w.ward_id}"),
                "wardNameLocal": w.ward_name_local,
                "booths": [],
            }
    allowed_ward_ids = set(ward_map.keys())

    voter_id_col = "voter_id" if "voter_id" in voter_cols else ("id" if "id" in voter_cols else None)
    voter_sr_col = "sl" if "sl" in voter_cols else ("sr_no" if "sr_no" in voter_cols else None)
    voter_epic_col = "epic" if "epic" in voter_cols else ("epic_no" if "epic_no" in voter_cols else None)
    voter_name_local_col = "name_kannada" if "name_kannada" in voter_cols else ("first_middle_name_local" if "first_middle_name_local" in voter_cols else None)
    voter_house_col = "house" if "house" in voter_cols else ("house_no_en" if "house_no_en" in voter_cols else None)
    voter_gender_col = "gender" if "gender" in voter_cols else None
    voter_booth_no_col = "booth_no" if "booth_no" in voter_cols else ("booth_id" if "booth_id" in voter_cols else None)
    voter_ward_code_col = "ward_code" if "ward_code" in voter_cols else None

    if not voter_booth_no_col:
        raise ValueError("public.voters missing booth mapping column")

    voters_by_key: Dict[tuple, List[Dict[str, Any]]] = {}
    counts_by_key: Dict[tuple, Dict[str, int]] = {}
    relevant_booths = [
        b
        for b in booth_rows
        if (not allowed_ward_ids or b.ward_id in allowed_ward_ids)
        and (not allowed_booth_ids or b.booth_id in allowed_booth_ids)
    ]
    allowed_ward_codes = sorted(
        {
            str(w.get("wardCode"))
            for w in ward_map.values()
            if w.get("wardCode") is not None and str(w.get("wardCode")).strip() != ""
        }
    )
    allowed_booth_nos = sorted({str(b.booth_no) for b in relevant_booths if b.booth_no is not None})

    voter_where_parts: List[str] = []
    voter_where_params: Dict[str, Any] = {}
    if voter_ward_code_col and allowed_ward_codes:
        clause, params = _build_in_clause(f"CAST({voter_ward_code_col} AS TEXT)", allowed_ward_codes, "ward_code")
        voter_where_parts.append(clause)
        voter_where_params.update(params)
    if allowed_booth_nos:
        clause, params = _build_in_clause(f"CAST({voter_booth_no_col} AS TEXT)", allowed_booth_nos, "booth_no")
        voter_where_parts.append(clause)
        voter_where_params.update(params)

    if not voter_where_parts:
        return {
            "assembly": {
                "assemblyId": assembly_row.assembly_pk,
                "assemblyCode": assembly_row.assembly_code or requested_assembly_code,
                "assemblyNameEn": assembly_row.assembly_name_en,
                "assemblyNameLocal": assembly_row.assembly_name_local,
                "wards": sorted(list(ward_map.values()), key=lambda w: (w.get("wardCode") or str(w["wardId"]))),
            }
        }

    if current.role != "SUPER_ADMIN" and "tenant_id" in voter_cols:
        if current.tenantId:
            voter_where_parts.append("(tenant_id = :tid OR tenant_id IS NULL)")
            voter_where_params["tid"] = current.tenantId
        else:
            pass

    voter_where_sql = " WHERE " + " AND ".join(voter_where_parts)

    if include_voters:
        voter_id_expr = f"{voter_id_col}" if voter_id_col else "ROW_NUMBER() OVER ()"
        voter_rows = db.execute(
            text(
                f"""
                SELECT
                    {voter_id_expr} AS voter_id,
                    {voter_sr_col if voter_sr_col else 'NULL'} AS sr_no,
                    {voter_epic_col if voter_epic_col else 'NULL'} AS epic_no,
                    {voter_name_en_col if voter_name_en_col else 'NULL'} AS name_en,
                    {voter_name_local_col if voter_name_local_col else 'NULL'} AS name_local,
                    {voter_house_col if voter_house_col else 'NULL'} AS house_no_en,
                    {voter_gender_col if voter_gender_col else 'NULL'} AS gender,
                    {voter_booth_no_col} AS booth_no,
                    {voter_ward_code_col if voter_ward_code_col else 'NULL'} AS ward_code
                FROM public.voters
                {voter_where_sql}
                """
            ),
            voter_where_params,
        ).all()

        for v in voter_rows:
            key = (str(v.ward_code) if v.ward_code is not None else None, str(v.booth_no))
            voters_by_key.setdefault(key, []).append({
                "voterId": v.voter_id,
                "sl": v.sr_no,
                "epicNo": v.epic_no,
                "firstMiddleNameEn": v.name_en,
                "firstMiddleNameLocal": v.name_local,
                "houseNoEn": v.house_no_en,
                "gender": v.gender,
            })
            stats = counts_by_key.setdefault(key, {"total": 0, "male": 0, "female": 0})
            stats["total"] += 1
            g = (v.gender or "").upper()
            if g.startswith("M"): stats["male"] += 1
            elif g.startswith("F"): stats["female"] += 1
    else:
        # Fetch stats only using aggregate query
        stats_rows = db.execute(
            text(
                f"""
                SELECT
                    {voter_ward_code_col if voter_ward_code_col else 'NULL'} AS ward_code,
                    {voter_booth_no_col} AS booth_no,
                    COUNT(*) AS total,
                    SUM(CASE WHEN UPPER({voter_gender_col if voter_gender_col else "'M'"}) LIKE 'M%' THEN 1 ELSE 0 END) AS male,
                    SUM(CASE WHEN UPPER({voter_gender_col if voter_gender_col else "'F'"}) LIKE 'F%' THEN 1 ELSE 0 END) AS female
                FROM public.voters
                {voter_where_sql}
                GROUP BY {voter_ward_code_col if voter_ward_code_col else 'NULL'}, {voter_booth_no_col}
                """
            ),
            voter_where_params,
        ).all()
        for s in stats_rows:
            key = (str(s.ward_code) if s.ward_code is not None else None, str(s.booth_no))
            counts_by_key[key] = {
                "total": int(s.total or 0),
                "male": int(s.male or 0),
                "female": int(s.female or 0),
            }

    for b in booth_rows:
        ward_id = b.ward_id
        if allowed_ward_ids and ward_id not in allowed_ward_ids:
            continue
        if allowed_booth_ids and b.booth_id not in allowed_booth_ids:
            continue
        if ward_id not in ward_map:
            ward_map[ward_id] = {
                "wardId": ward_id,
                "wardCode": str(b.ward_code) if b.ward_code is not None else None,
                "wardNameEn": f"Ward {ward_id}",
                "wardNameLocal": None,
                "booths": [],
            }

        key = (str(b.ward_code) if b.ward_code is not None else None, str(b.booth_no))
        booth_entry = {
            "boothId": int(b.booth_id),
            "boothNo": int(b.booth_no) if b.booth_no is not None else None,
            "boothNameEn": b.booth_name_en or (f"Booth {b.booth_no}"),
            "boothNameLocal": b.booth_name_local,
            "voterStats": counts_by_key.get(key, {"total": 0, "male": 0, "female": 0}),
        }
        if include_voters:
            booth_entry["voters"] = voters_by_key.get(key, [])
        else:
            booth_entry["voters"] = []
        ward_map[ward_id]["booths"].append(booth_entry)

    return {
        "assembly": {
            "assemblyId": assembly_row.assembly_pk,
            "assemblyCode": assembly_row.assembly_code or requested_assembly_code,
            "assemblyNameEn": assembly_row.assembly_name_en,
            "assemblyNameLocal": assembly_row.assembly_name_local,
            "wards": sorted(
                [
                    {**ward, "booths": sorted(ward.get("booths", []), key=lambda b: (b.get("boothNo") or 0, b.get("boothId") or 0))}
                    for ward in ward_map.values()
                ],
                key=lambda w: (w.get("wardCode") or str(w["wardId"])),
            ),
            }
        }


def _upload_and_save_snapshot(db: Session, data: Dict[str, Any], key: str, tenant_id: str, assembly_code: str, ward_code: Optional[str], booth_id: Optional[int], level: str) -> None:
    s3_url = s3_upload_bytes(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"), "application/json", key)

    existing = (
        db.query(VoterSnapshot)
        .filter(
            VoterSnapshot.tenant_id == tenant_id,
            VoterSnapshot.assembly_code == assembly_code,
            VoterSnapshot.ward_code == ward_code,
            VoterSnapshot.booth_id == booth_id,
            VoterSnapshot.snapshot_level == level,
        )
        .first()
    )
    if existing:
        existing.s3_url = s3_url
        existing.updated_at = datetime.utcnow()
        existing.version = (existing.version or 1) + 1
        db.add(existing)
    else:
        db.add(
            VoterSnapshot(
                tenant_id=tenant_id,
                assembly_code=assembly_code,
                ward_code=ward_code,
                booth_id=booth_id,
                snapshot_level=level,
                s3_url=s3_url,
                version=1,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
        )


def _build_booth_stats(db: Session, booth_ids: List[int], tenant_id: Optional[str]) -> Dict[int, Dict[str, int]]:
    if not booth_ids:
        return {}
    q = db.query(Voter.booth_id, Voter.gender, func.count(Voter.voter_id))
    q = q.filter(Voter.booth_id.in_(booth_ids))
    if tenant_id is not None:
        q = q.filter(Voter.tenant_id == tenant_id)
    q = q.group_by(Voter.booth_id, Voter.gender)
    rows = q.all()
    stats: Dict[int, Dict[str, int]] = {}
    for booth_id, gender, count in rows:
        booth_stat = stats.setdefault(int(booth_id), {"total": 0, "male": 0, "female": 0})
        booth_stat["total"] += int(count)
        g = (gender or "").upper()
        if g.startswith("M"):
            booth_stat["male"] += int(count)
        elif g.startswith("F"):
            booth_stat["female"] += int(count)
    return stats


def _build_assembly_json(
    db: Session,
    assembly: Assembly,
    wards: List[Ward],
    include_voters: bool,
    tenant_id: Optional[str],
    booth_stats: Optional[Dict[int, Dict[str, int]]] = None,
) -> Dict[str, Any]:
    assembly_map: Dict[str, Any] = {
        "assemblyId": assembly.assembly_id,
        "assemblyNameEn": assembly.assembly_name_en,
        "assemblyNameLocal": assembly.assembly_name_local,
    }

    ward_list: List[Dict[str, Any]] = []
    for ward in wards:
        ward_map: Dict[str, Any] = {
            "wardId": ward.ward_id,
            "wardNameEn": ward.ward_name_en,
            "wardNameLocal": ward.ward_name_local,
        }

        booths_q = db.query(Booth).filter(Booth.ward_id == ward.ward_id)
        if tenant_id is not None:
            booths_q = booths_q.filter(Booth.tenant_id == tenant_id)
        booths = booths_q.all()
        booth_list: List[Dict[str, Any]] = []
        for booth in booths:
            booth_map: Dict[str, Any] = {
                "boothId": booth.booth_id,
                "boothNameEn": booth.polling_station_adr_en,
                "boothNameLocal": booth.polling_station_adr_local,
            }
            if include_voters:
                voters_q = db.query(Voter).filter(Voter.booth_id == booth.booth_id)
                if tenant_id is not None:
                    voters_q = voters_q.filter(Voter.tenant_id == tenant_id)
                voters = voters_q.all()
                booth_map["voters"] = [_build_voter_map(v) for v in voters]
            elif booth_stats is not None:
                booth_map["voterStats"] = booth_stats.get(booth.booth_id, {"total": 0, "male": 0, "female": 0})
            booth_list.append(booth_map)

        ward_map["booths"] = booth_list
        ward_list.append(ward_map)

    assembly_map["wards"] = ward_list
    return {"assembly": assembly_map}


def _build_snapshot_from_data(
    db: Session,
    assembly_code: str,
    include_voters: bool,
    current: JwtUserDetails,
) -> Dict[str, Any]:
    tenant_id = current.tenantId
    normalized_code = normalize_assembly_code(assembly_code)
    assembly = (
        db.query(Assembly)
        .filter(Assembly.assembly_code == normalized_code)
        .filter(Assembly.tenant_id == tenant_id if tenant_id else True)
        .first()
    )
    if not assembly:
        raise ValueError(f"Assembly not found for assemblyCode: {assembly_code}")

    assignment = (current.assignmentType or "ASSEMBLY").upper()
    if assignment == "ASSEMBLY":
        wards = db.query(Ward).filter(Ward.assembly_id == assembly.assembly_id).all()
        if tenant_id:
            wards = [w for w in wards if w.tenant_id == tenant_id]
        booth_ids = []
        if not include_voters:
            booth_ids = [
                b.booth_id
                for b in db.query(Booth)
                .filter(Booth.ward_id.in_([w.ward_id for w in wards]) if wards else True)
                .filter(Booth.tenant_id == tenant_id if tenant_id else True)
                .all()
            ]
        stats = _build_booth_stats(db, booth_ids, tenant_id) if not include_voters else None
        return _build_assembly_json(db, assembly, wards, include_voters, tenant_id, stats)

    if assignment == "WARD":
        ward = (
            db.query(Ward)
            .filter(Ward.assembly_id == assembly.assembly_id)
            .filter(
                (Ward.ward_id == current.assignmentId) | (Ward.ward_code == str(current.assignmentId))
            )
            .first()
        )
        if not ward:
            raise ValueError("No ward snapshot found")
        booth_ids = []
        if not include_voters:
            booth_ids = [
                b.booth_id
                for b in db.query(Booth)
                .filter(Booth.ward_id == ward.ward_id)
                .filter(Booth.tenant_id == tenant_id if tenant_id else True)
                .all()
            ]
        stats = _build_booth_stats(db, booth_ids, tenant_id) if not include_voters else None
        return _build_assembly_json(db, assembly, [ward], include_voters, tenant_id, stats)

    if assignment == "BOOTH":
        booth_row = (
            db.query(Booth)
            .filter(Booth.booth_id == current.assignmentId)
            .filter(Booth.tenant_id == tenant_id if tenant_id else True)
            .first()
        )
        if not booth_row:
            raise ValueError("No booth snapshot found")
        ward = db.query(Ward).filter(Ward.ward_id == booth_row.ward_id).first()
        if not ward:
            raise ValueError("No ward found for booth")
        voters = []
        if include_voters:
            voters_q = db.query(Voter).filter(Voter.booth_id == booth_row.booth_id)
            if tenant_id:
                voters_q = voters_q.filter(Voter.tenant_id == tenant_id)
            voters = voters_q.all()
        stats = _build_booth_stats(db, [booth_row.booth_id], tenant_id) if not include_voters else None
        booth_payload = {
            "boothId": booth_row.booth_id,
            "boothNameEn": booth_row.polling_station_adr_en,
            "boothNameLocal": booth_row.polling_station_adr_local,
        }
        if include_voters:
            booth_payload["voters"] = [_build_voter_map(v) for v in voters]
        elif stats is not None:
            booth_payload["voterStats"] = stats.get(booth_row.booth_id, {"total": 0, "male": 0, "female": 0})
        return {
            "assembly": {
                "assemblyId": assembly.assembly_id,
                "assemblyNameEn": assembly.assembly_name_en,
                "assemblyNameLocal": assembly.assembly_name_local,
                "wards": [
                    {
                        "wardId": ward.ward_id,
                        "wardNameEn": ward.ward_name_en,
                        "wardNameLocal": ward.ward_name_local,
                        "booths": [booth_payload],
                    }
                ],
            }
        }

    raise ValueError(f"Invalid role: {current.assignmentType}")


def generate_snapshots(db: Session, assembly_id: int, tenant_id: str) -> None:
    assembly = db.query(Assembly).filter(Assembly.assembly_id == assembly_id).first()
    if not assembly:
        raise ValueError(f"Assembly not found: {assembly_id}")

    assembly_code = normalize_assembly_code(assembly.assembly_code)

    wards = db.query(Ward).filter(Ward.tenant_id == tenant_id, Ward.assembly_id == assembly.assembly_id).all()
    if wards:
        assembly_json = _build_assembly_json(db, assembly, wards, True, tenant_id)
        _upload_and_save_snapshot(db, assembly_json, f"snapshots/{tenant_id}/{assembly_code}/assembly.json", tenant_id, assembly_code, None, None, "ASSEMBLY")

    for ward in wards:
        ward_json = _build_assembly_json(db, assembly, [ward], True, tenant_id)
        _upload_and_save_snapshot(
            db,
            ward_json,
            f"snapshots/{tenant_id}/{assembly_code}/wards/ward_{ward.ward_code}.json",
            tenant_id,
            assembly_code,
            ward.ward_code,
            None,
            "WARD",
        )

    booths = (
        db.query(Booth)
        .join(Ward, Booth.ward_id == Ward.ward_id)
        .join(Assembly, Ward.assembly_id == Assembly.assembly_id)
        .filter(Assembly.assembly_code == assembly_code, Booth.tenant_id == tenant_id)
        .all()
    )

    for booth in booths:
        voters = db.query(Voter).filter(Voter.tenant_id == tenant_id, Voter.booth_id == booth.booth_id).all()
        if not voters:
            continue

        ward = db.query(Ward).filter(Ward.ward_id == booth.ward_id).first()
        if not ward:
            continue

        booth_json = {
            "assembly": {
                "assemblyId": assembly.assembly_id,
                "assemblyNameEn": assembly.assembly_name_en,
                "assemblyNameLocal": assembly.assembly_name_local,
                "wards": [
                    {
                        "wardId": ward.ward_id,
                        "wardNameEn": ward.ward_name_en,
                        "wardNameLocal": ward.ward_name_local,
                        "booths": [
                            {
                                "boothId": booth.booth_id,
                                "boothNameEn": booth.polling_station_adr_en,
                                "boothNameLocal": booth.polling_station_adr_local,
                                "voters": [_build_voter_map(v) for v in voters],
                            }
                        ],
                    }
                ],
            }
        }

        _upload_and_save_snapshot(
            db,
            booth_json,
            f"snapshots/{tenant_id}/{assembly_code}/booths/booth_{booth.booth_id}.json",
            tenant_id,
            assembly_code,
            ward.ward_code,
            booth.booth_id,
            "BOOTH",
        )


@app.post(f"{CONTEXT_PATH}/api/excel/upload")
def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db), current: JwtUserDetails = Depends(require_roles("ADMIN"))):
    try:
        wb = _load_workbook_from_upload(file)

        def sheet(name: str):
            if name not in wb.sheetnames:
                raise ValueError(f"Missing required sheet: {name}")
            return wb[name]

        def headers(ws) -> Dict[str, int]:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            return {str(v).strip().upper(): i for i, v in enumerate(first_row) if v is not None}

        ws_assembly = sheet("ASSEMBLY")
        ws_ward = sheet("WARD")
        ws_booth = sheet("BOOTH")
        ws_data = sheet("DATA")

        h_assembly = headers(ws_assembly)
        h_ward = headers(ws_ward)
        h_booth = headers(ws_booth)
        h_data = headers(ws_data)

        for col in ["ASSEMBLY_NO", "ASSEMBLY_NAME_EN", "ASSEMBLY_NAME_LOCAL"]:
            if col not in h_assembly:
                raise ValueError(f"Required column missing in sheet 'ASSEMBLY': {col}")

        arow = next(ws_assembly.iter_rows(min_row=2, max_row=2, values_only=True))
        assembly_no = int(arow[h_assembly["ASSEMBLY_NO"]])
        assembly_code = normalize_assembly_code(assembly_no)
        assembly_name_en = arow[h_assembly["ASSEMBLY_NAME_EN"]]
        assembly_name_local = arow[h_assembly["ASSEMBLY_NAME_LOCAL"]]

        assembly = db.query(Assembly).filter(Assembly.assembly_code == assembly_code, Assembly.tenant_id == current.tenantId).first()
        if not assembly:
            assembly = Assembly(
                assembly_id=assembly_no,
                assembly_code=assembly_code,
                assembly_name_en=assembly_name_en or f"Assembly {assembly_no}",
                assembly_name_local=assembly_name_local,
                tenant_id=current.tenantId,
            )
            db.add(assembly)
            db.flush()

        # Wards
        for i, row in enumerate(ws_ward.iter_rows(min_row=2, values_only=True)):
            ward_code = row[h_ward.get("WARD_CODE", -1)]
            if ward_code is None or str(ward_code).strip() == "":
                continue
            
            assembly_no_row = row[h_ward.get("ASSEMBLY_NO", -1)]
            if assembly_no_row is None:
                raise ValueError(f"Assembly number missing for ward: {ward_code}")
            
            cur_asm_no = int(assembly_no_row)
            w_id = (cur_asm_no * 1000) + (i + 1)
            
            exists = db.query(Ward).filter(Ward.ward_id == w_id).first()
            if exists:
                continue

            assembly_ref = db.query(Assembly).filter(Assembly.assembly_id == cur_asm_no).first()
            if not assembly_ref:
                assembly_ref = Assembly(
                    assembly_id=cur_asm_no,
                    assembly_code=normalize_assembly_code(cur_asm_no),
                    assembly_name_en=f"Assembly {cur_asm_no}",
                    tenant_id=current.tenantId,
                )
                db.add(assembly_ref)
                db.flush()

            db.add(
                Ward(
                    ward_id=w_id,
                    ward_code=str(ward_code),
                    ward_name_en=row[h_ward.get("WARD_NAME_EN", -1)] or f"Ward {ward_code}",
                    ward_name_local=row[h_ward.get("WARD_NAME_LOCAL", -1)],
                    tenant_id=current.tenantId,
                    assembly_id=assembly_ref.assembly_id,
                )
            )

        db.flush()

        # Booths
        for row in ws_booth.iter_rows(min_row=2, values_only=True):
            booth_no = row[h_booth.get("BOOTH_NO", -1)]
            ward_code = row[h_booth.get("WARD_CODE", -1)]
            if booth_no is None or ward_code is None:
                continue

            ward = db.query(Ward).filter(Ward.ward_code == str(ward_code), Ward.tenant_id == current.tenantId).first()
            if not ward:
                raise ValueError(f"Ward not found for booth {booth_no} (wardCode={ward_code})")

            b_id = (ward.ward_id * 1000) + int(booth_no)
            booth = db.query(Booth).filter(Booth.booth_id == b_id).first()
            if not booth:
                db.add(
                    Booth(
                        booth_id=b_id,
                        booth_no=str(booth_no),
                        ward_id=ward.ward_id,
                        polling_station_adr_en=row[h_booth.get("BOOTH_ADD_EN", -1)],
                        polling_station_adr_local=row[h_booth.get("BOOTH_ADD_LOCAL", -1)],
                        tenant_id=current.tenantId,
                    )
                )

        db.flush()

        # Voters
        seen = set()
        incoming = []
        for row in ws_data.iter_rows(min_row=2, values_only=True):
            sr_no = row[h_data.get("SL", -1)] if "SL" in h_data else None
            booth_no = row[h_data.get("BOOTH_NO", -1)] if "BOOTH_NO" in h_data else None
            epic = row[h_data.get("EPIC", -1)] if "EPIC" in h_data else None
            if sr_no is None or booth_no is None or epic is None:
                continue
            epic = str(epic)
            if epic in seen:
                continue
            seen.add(epic)
            incoming.append((row, int(booth_no), epic, int(sr_no)))

        epic_set = {epic for _, _, epic, _ in incoming}
        existing_epics = set(
            r[0]
            for r in db.query(Voter.epic_no)
            .filter(Voter.tenant_id == current.tenantId, Voter.epic_no.in_(list(epic_set)))
            .all()
        )

        saved = 0
        for row, booth_no, epic, sr_no in incoming:
            if epic in existing_epics:
                continue

            booth = db.query(Booth).filter(Booth.booth_id == booth_no).first()
            if not booth:
                raise ValueError(f"Booth not found: {booth_no}")

            mobile = str(row[h_data["MOBILE"]]) if "MOBILE" in h_data and row[h_data["MOBILE"]] is not None else None
            if mobile is not None:
                mobile = re.sub(r"\D+", "", mobile)
                if len(mobile) != 10:
                    mobile = None

            voter = Voter(
                tenant_id=current.tenantId,
                booth_id=booth.booth_id,
                sr_no=sr_no,
                epic_no=epic,
                house_no_en=str(row[h_data["HOUSE"]]) if "HOUSE" in h_data and row[h_data["HOUSE"]] is not None else None,
                first_middle_name_en=str(row[h_data["NAME_EN"]]) if "NAME_EN" in h_data and row[h_data["NAME_EN"]] is not None else None,
                first_middle_name_local=str(row[h_data["NAME_KANNADA"]]) if "NAME_KANNADA" in h_data and row[h_data["NAME_KANNADA"]] is not None else None,
                gender=str(row[h_data["GENDER"]]) if "GENDER" in h_data and row[h_data["GENDER"]] is not None else None,
                age=int(row[h_data["AGE"]]) if "AGE" in h_data and row[h_data["AGE"]] is not None else None,
                relation_type=str(row[h_data["REL_TYPE"]]) if "REL_TYPE" in h_data and row[h_data["REL_TYPE"]] is not None else None,
                relation_first_middle_name_en=str(row[h_data["REL_ENG"]]) if "REL_ENG" in h_data and row[h_data["REL_ENG"]] is not None else None,
                relation_first_middle_name_local=str(row[h_data["REL_KANNADA"]]) if "REL_KANNADA" in h_data and row[h_data["REL_KANNADA"]] is not None else None,
                mobile=mobile,
            )
            db.add(voter)
            saved += 1

        db.flush()
        if saved > 0:
            generate_snapshots(db, assembly_no, current.tenantId)
        db.commit()

        return api_success("Successfully uploaded", saved)
    except Exception as ex:
        db.rollback()
        return JSONResponse(status_code=500, content=api_error("Excel upload failed", str(ex)))


def _synthesize_booth_id(ward_id: int, booth_no: int) -> int:
    # *10000 avoids (ward_id*1000+booth) collisions across adjacent wards/booth numbers.
    return (ward_id * 10000) + int(booth_no)


_MASTER_ROLL_IMPORT_STATUS: Dict[str, Any] = {
    "active": False,
    "phase": "idle",
    "progress": 0,
    "assembly_no": None,
    "assembly_name_en": None,
    "inserted": {"assembly": 0, "wards": 0, "booths": 0, "voters": 0},
    "error": None,
}


def _load_workbook_from_upload(file: UploadFile, *, data_only: bool = True):
    """Read upload into memory; Starlette SpooledTemporaryFile lacks seekable() for openpyxl."""
    raw = file.file.read()
    if not raw:
        raise ValueError("Uploaded file is empty")
    return load_workbook(BytesIO(raw), data_only=data_only)


def _master_roll_status(
    phase: str,
    progress: int,
    *,
    assembly_no: Optional[int] = None,
    assembly_name_en: Optional[str] = None,
    inserted: Optional[Dict[str, int]] = None,
    error: Optional[str] = None,
    active: bool = True,
) -> None:
    prev_inserted = _MASTER_ROLL_IMPORT_STATUS.get("inserted") or {}
    next_inserted = dict(prev_inserted)
    if inserted is not None:
        next_inserted.update(inserted)
    _MASTER_ROLL_IMPORT_STATUS.update({
        "active": active,
        "phase": phase,
        "progress": max(0, min(100, int(progress))),
        "assembly_no": assembly_no if assembly_no is not None else _MASTER_ROLL_IMPORT_STATUS.get("assembly_no"),
        "assembly_name_en": assembly_name_en if assembly_name_en is not None else _MASTER_ROLL_IMPORT_STATUS.get("assembly_name_en"),
        "inserted": next_inserted,
        "error": error,
    })


@app.get(f"{CONTEXT_PATH}/api/admin/master-roll/import-status")
def master_roll_import_status(
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN")),
):
    return api_success("Master roll import status", dict(_MASTER_ROLL_IMPORT_STATUS))


@app.post(f"{CONTEXT_PATH}/api/admin/master-roll/upload")
def upload_master_roll(
    file: UploadFile = File(...),
    resume: bool = False,
    db: Session = Depends(get_db),
    current: JwtUserDetails = Depends(require_roles("SUPER_ADMIN"))
):
    # resume=true: continue import without deleting existing rows (upsert only).
    _ = resume
    _master_roll_status("starting", 1, inserted={}, error=None)
    try:
        wb = _load_workbook_from_upload(file)

        def sheet(name: str):
            for sn in wb.sheetnames:
                if sn.upper() == name.upper(): return wb[sn]
            raise ValueError(f"Missing required sheet: {name}")

        def headers(ws) -> Dict[str, int]:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            return {str(v).strip().upper(): i for i, v in enumerate(first_row) if v is not None}

        ws_assembly = sheet("ASSEMBLY")
        ws_ward = sheet("WARD")
        ws_booth = sheet("BOOTH")
        ws_data = sheet("DATA")

        h_assembly = headers(ws_assembly)
        h_ward = headers(ws_ward)
        h_booth = headers(ws_booth)
        h_data = headers(ws_data)

        # 1. Identify Assembly & Compute Tenant ID
        arow = next(ws_assembly.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if not arow: raise ValueError("Sheet ASSEMBLY is empty")
        
        def get_val(r, h, k):
            idx = h.get(k.upper())
            if idx is None or idx < 0: return None
            return r[idx]

        asm_no_val = get_val(arow, h_assembly, "ASSEMBLY_NO") or get_val(arow, h_assembly, "ASSEMBLY_ID")
        if asm_no_val is None: raise ValueError("Sheet ASSEMBLY missing ASSEMBLY_NO")
        
        assembly_no = int(asm_no_val)
        assembly_code = normalize_assembly_code(assembly_no)
        # Format: voter-000000000151
        tenant_id = f"voter-{assembly_code}"
        
        asm_name_en = get_val(arow, h_assembly, "ASSEMBLY_NAME_EN")
        asm_name_local = get_val(arow, h_assembly, "ASSEMBLY_NAME_LOCAL")

        _master_roll_status("assembly", 10, assembly_no=assembly_no, assembly_name_en=asm_name_en, inserted={"assembly": 0, "wards": 0, "booths": 0, "voters": 0})
        db.execute(text("""
            INSERT INTO public.assembly (assembly_no, assembly_name_en, assembly_name_local, assembly_code, tenant_id)
            VALUES (:no, :en, :local, :code, :tid)
            ON CONFLICT (assembly_no) DO UPDATE SET
                assembly_name_en = EXCLUDED.assembly_name_en,
                assembly_name_local = EXCLUDED.assembly_name_local,
                tenant_id = EXCLUDED.tenant_id
        """), {"no": assembly_no, "en": asm_name_en, "local": asm_name_local, "code": assembly_code, "tid": tenant_id})
        db.commit()
        _master_roll_status("assembly", 22, assembly_no=assembly_no, assembly_name_en=asm_name_en, inserted={"assembly": 1})

        _master_roll_status("wards", 28, assembly_no=assembly_no, assembly_name_en=asm_name_en)
        ward_count = 0
        booth_count = 0
        ward_map: Dict[str, int] = {}
        ward_index = 0
        for row in ws_ward.iter_rows(min_row=2, values_only=True):
            w_code = get_val(row, h_ward, "WARD_CODE") or get_val(row, h_ward, "WARD_COD")
            if w_code is None:
                continue

            ward_index += 1
            w_id = (assembly_no * 1000) + ward_index
            ward_map[str(w_code)] = w_id

            w_name_en = get_val(row, h_ward, "WARD_NAME_EN") or get_val(row, h_ward, "WARD_NAME")
            w_name_local = get_val(row, h_ward, "WARD_NAME_LOCAL") or get_val(row, h_ward, "WARD_NAME_L")

            db.execute(text("""
                INSERT INTO public.wards (id, ward_code, ward_name_en, ward_name_local, assembly_no, tenant_id)
                VALUES (:id, :code, :en, :local, :ano, :tid)
                ON CONFLICT (id) DO UPDATE SET
                    ward_name_en = EXCLUDED.ward_name_en,
                    ward_name_local = EXCLUDED.ward_name_local,
                    tenant_id = EXCLUDED.tenant_id
            """), {
                "id": w_id,
                "code": str(w_code),
                "en": w_name_en,
                "local": w_name_local,
                "ano": assembly_no,
                "tid": tenant_id,
            })
            ward_count += 1

        db.commit()
        _master_roll_status(
            "wards",
            38,
            assembly_no=assembly_no,
            assembly_name_en=asm_name_en,
            inserted={"assembly": 1, "wards": ward_count},
        )

        _master_roll_status("booths", 45, assembly_no=assembly_no, assembly_name_en=asm_name_en)
        seen_booth_keys: set[tuple[str, int]] = set()
        for row in ws_booth.iter_rows(min_row=2, values_only=True):
            b_no_raw = get_val(row, h_booth, "BOOTH_NO") or get_val(row, h_booth, "BOOTH_N")
            w_code = get_val(row, h_booth, "WARD_CODE") or get_val(row, h_booth, "WARD_COD")
            if b_no_raw is None or w_code is None:
                continue

            b_no = int(b_no_raw)
            w_code_str = str(w_code).strip()
            booth_key = (w_code_str, b_no)
            if booth_key in seen_booth_keys:
                continue
            seen_booth_keys.add(booth_key)

            w_id = ward_map.get(w_code_str)
            if not w_id:
                continue

            b_id = _synthesize_booth_id(w_id, b_no)
            db.execute(text("""
                INSERT INTO public.booths (id, booth_no, ward_code, ward_id, booth_add_en, booth_add_local, tenant_id)
                VALUES (:id, :no, :wc, :wid, :en, :local, :tid)
                ON CONFLICT (id) DO UPDATE SET
                    booth_add_en = EXCLUDED.booth_add_en,
                    booth_add_local = EXCLUDED.booth_add_local,
                    tenant_id = EXCLUDED.tenant_id
            """), {
                "id": b_id,
                "no": str(b_no),
                "wc": w_code_str,
                "wid": w_id,
                "en": get_val(row, h_booth, "BOOTH_ADD_EN") or get_val(row, h_booth, "POLLING_STATION_ADR_EN") or get_val(row, h_booth, "BOOTH_NAME_EN"),
                "local": get_val(row, h_booth, "BOOTH_ADD_LOCAL") or get_val(row, h_booth, "POLLING_STATION_ADR_LOCAL") or get_val(row, h_booth, "BOOTH_NAME_LOCAL"),
                "tid": tenant_id,
            })
            booth_count += 1

        db.commit()
        _master_roll_status(
            "booths",
            52,
            assembly_no=assembly_no,
            assembly_name_en=asm_name_en,
            inserted={"assembly": 1, "wards": ward_count, "booths": booth_count},
        )

        est_voter_rows = max(0, int(getattr(ws_data, "max_row", 2) or 2) - 1)
        _master_roll_status(
            "voters",
            55,
            assembly_no=assembly_no,
            assembly_name_en=asm_name_en,
            inserted={"assembly": 1, "wards": ward_count, "booths": booth_count, "voters": 0},
        )

        # Remote DB (e.g. AWS): large batches can exceed SSL/network timeouts — use smaller batches + no statement timeout.
        db.execute(text("SET statement_timeout = 0"))
        db.execute(text("SET lock_timeout = '120s'"))

        voter_insert_sql = text("""
            INSERT INTO public.voters (
                epic, sl, name_en, name_kannada, rel_eng, rel_kannada, rel_type,
                gender, age, house, booth_no, ward_code, mobile, tenant_id
            )
            VALUES (
                :epic, :sl, :name_en, :name_k, :rel_e, :rel_k, :rel_t,
                :gender, :age, :house, :bno, :wcode, :mobile, :tid
            )
            ON CONFLICT (epic) DO UPDATE SET
                name_en = EXCLUDED.name_en,
                name_kannada = COALESCE(EXCLUDED.name_kannada, public.voters.name_kannada),
                rel_eng = COALESCE(EXCLUDED.rel_eng, public.voters.rel_eng),
                rel_kannada = COALESCE(EXCLUDED.rel_kannada, public.voters.rel_kannada),
                rel_type = COALESCE(EXCLUDED.rel_type, public.voters.rel_type),
                gender = COALESCE(EXCLUDED.gender, public.voters.gender),
                age = COALESCE(EXCLUDED.age, public.voters.age),
                house = COALESCE(EXCLUDED.house, public.voters.house),
                booth_no = COALESCE(EXCLUDED.booth_no, public.voters.booth_no),
                ward_code = COALESCE(EXCLUDED.ward_code, public.voters.ward_code),
                mobile = COALESCE(public.voters.mobile, EXCLUDED.mobile),
                tenant_id = EXCLUDED.tenant_id
        """)
        voter_batch: List[Dict[str, Any]] = []
        seen_epics: set[str] = set()
        voter_count = 0
        batch_size = 400

        for row in ws_data.iter_rows(min_row=2, values_only=True):
            epic = get_val(row, h_data, "EPIC")
            if not epic:
                continue
            epic = str(epic).strip()
            if not epic or epic in seen_epics:
                continue
            seen_epics.add(epic)

            voter_batch.append({
                "epic": epic,
                "sl": str(get_val(row, h_data, "SL") or ""),
                "name_en": get_val(row, h_data, "NAME_EN"),
                "name_k": get_val(row, h_data, "NAME_KANNADA") or get_val(row, h_data, "NAME_LOCAL"),
                "rel_e": get_val(row, h_data, "REL_ENG"),
                "rel_k": get_val(row, h_data, "REL_KANNADA") or get_val(row, h_data, "REL_LOCAL"),
                "rel_t": get_val(row, h_data, "REL_TYPE"),
                "gender": get_val(row, h_data, "GENDER"),
                "age": str(get_val(row, h_data, "AGE") or ""),
                "house": str(get_val(row, h_data, "HOUSE") or ""),
                "bno": str(get_val(row, h_data, "BOOTH_NO") or get_val(row, h_data, "BOOTH_N") or ""),
                "wcode": str(get_val(row, h_data, "WARD_CODE") or get_val(row, h_data, "WARD_COD") or ""),
                "mobile": str(get_val(row, h_data, "MOBILE")) if get_val(row, h_data, "MOBILE") else None,
                "tid": tenant_id,
            })
            if len(voter_batch) >= batch_size:
                try:
                    db.execute(voter_insert_sql, voter_batch)
                    voter_count += len(voter_batch)
                    voter_batch.clear()
                    db.commit()
                except Exception as batch_ex:
                    db.rollback()
                    raise RuntimeError(
                        f"Voter batch failed after {voter_count} rows loaded. Re-upload the same Excel to continue (upsert). Original: {batch_ex}"
                    ) from batch_ex
                if est_voter_rows > 0:
                    voter_progress = 55 + int(40 * voter_count / est_voter_rows)
                else:
                    voter_progress = 90
                _master_roll_status(
                    "voters",
                    voter_progress,
                    assembly_no=assembly_no,
                    assembly_name_en=asm_name_en,
                    inserted={"assembly": 1, "wards": ward_count, "booths": booth_count, "voters": voter_count},
                )

        if voter_batch:
            db.execute(voter_insert_sql, voter_batch)
            voter_count += len(voter_batch)
            db.commit()

        _master_roll_status(
            "done",
            100,
            assembly_no=assembly_no,
            assembly_name_en=asm_name_en,
            inserted={"assembly": 1, "wards": ward_count, "booths": booth_count, "voters": voter_count},
            active=False,
        )
        return api_success(
            f"Master roll imported for assembly {assembly_no} ({asm_name_en or 'assembly'})",
            {
                "tenant_id": tenant_id,
                "assembly_no": assembly_no,
                "inserted": {
                    "assembly": 1,
                    "wards": ward_count,
                    "booths": booth_count,
                    "voters": voter_count,
                },
            },
        )
    except Exception as ex:
        db.rollback()
        err_text = f"{type(ex).__name__}: {str(ex)}"
        _master_roll_status("error", 0, error=err_text, active=False)
        print(f"[IMPORT_ERROR] {traceback.format_exc()}")
        return JSONResponse(status_code=500, content=api_error("Master roll upload failed", err_text))


@app.get("/health")
def health():
    return {"ok": True}


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app.main:app", host="0.0.0.0", port=int(os.getenv("PORT", "8081")), reload=True)
