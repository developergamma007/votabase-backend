import os
import re
import tempfile
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import boto3
import pdfplumber
import shutil
import subprocess
import uuid
from fastapi import APIRouter, File, Query, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

# Keep this module independent from main.py auth and models.
_ENGINE = None
_SESSION = None


def _get_engine():
    global _ENGINE, _SESSION
    if _ENGINE is not None:
        return _ENGINE
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        return None
    _ENGINE = create_engine(database_url, future=True, pool_pre_ping=True)
    _SESSION = sessionmaker(bind=_ENGINE, autoflush=False, autocommit=False, future=True)
    return _ENGINE


def _aws_clients():
    region = os.getenv("AWS_REGION", "ap-south-1")
    access_key = os.getenv("AWS_ACCESS_KEY_ID", "")
    secret_key = os.getenv("AWS_SECRET_ACCESS_KEY", "")
    session_token = os.getenv("AWS_SESSION_TOKEN", "")
    session_kwargs: Dict[str, Any] = {"region_name": region}
    if access_key and secret_key:
        session_kwargs["aws_access_key_id"] = access_key
        session_kwargs["aws_secret_access_key"] = secret_key
        if session_token:
            session_kwargs["aws_session_token"] = session_token
    s3 = boto3.client("s3", **session_kwargs)
    textract = boto3.client("textract", **session_kwargs)
    return s3, textract

router = APIRouter()


def _normalize_epic(raw: str) -> str:
    return re.sub(r"\s+", "", (raw or "").strip()).upper()


# Allow optional space/newline between letters and digits (Textract / OCR splits).
CARD_EPIC_RE = re.compile(r"[A-Z]{3}\s*\d{7}")
CARD_HOUSE_RE = re.compile(r"H\.?\s*No\.?\s*[:\-]?\s*([A-Za-z0-9\/\-]+)", re.IGNORECASE)
CARD_AGE_GENDER_RE = re.compile(r"Age\s*[:\-]?\s*(\d+)\s*.*Gender\s*[:\-]?\s*([A-Za-z]+)", re.IGNORECASE)
CARD_GENDER_RE = re.compile(r"Gender\s*[:\-]?\s*([A-Za-z]+)", re.IGNORECASE)
CARD_AGE_RE = re.compile(r"Age\s*[:\-]?\s*(\d+)", re.IGNORECASE)
CARD_NAME_BLOCK_RE = re.compile(
    r"Name\s*:\s*(.+?)(?=\s*(Father's Name|Husband's Name|Mother's Name|Other's Name|Age|Gender|H\.?\s*No\.?|$))",
    re.IGNORECASE,
)
CARD_REL_BLOCK_RE = re.compile(
    r"(Father's Name|Husband's Name|Mother's Name|Other's Name)\s*:\s*(.+?)(?=\s*(Name\s*:|Age|Gender|H\.?\s*No\.?|$))",
    re.IGNORECASE,
)
CARD_SERIAL_RE = re.compile(r"^\s*(\d{1,4})\b")
PHOTO_CLEAN_RE = re.compile(r"\bPhoto\b|\bAvailable\b|Photo\s+Available", re.IGNORECASE)


def ensure_extract_schema() -> None:
    engine = _get_engine()
    if engine is None:
        return
    with engine.begin() as conn:
        conn.execute(text("CREATE SCHEMA IF NOT EXISTS extract"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS extract.pdf_extract_jobs (
                    id SERIAL PRIMARY KEY,
                    filename TEXT,
                    status TEXT,
                    rows_count INTEGER,
                    error TEXT,
                    result_path TEXT,
                    created_at TIMESTAMP DEFAULT NOW(),
                    completed_at TIMESTAMP
                )
                """
            )
        )


def _parse_header_info(lines: List[str]) -> Dict[str, Optional[str]]:
    corporation = None
    ward_no = None
    ward_name_en = None
    booth_name_en = None
    booth_no = None

    for line in lines:
        if "Authority" in line and ":" in line:
            corporation = line.split(":", 1)[-1].strip()
        match = re.match(r"\s*(\d+)\s*[-–]\s*(.+)", line)
        if match:
            if not ward_no:
                ward_no = match.group(1).strip()
                ward_name_en = match.group(2).strip()
            elif not booth_name_en:
                booth_name_en = match.group(2).strip()
        part = re.search(r"Part\s*No\.?\s*[:\-]?\s*(\d+)", line, re.IGNORECASE)
        if part:
            booth_no = part.group(1).strip()
        if not booth_no:
            booth_m = re.search(r"(?:Booth|B\.?\s*O\.?)\s*No\.?\s*[:\-]?\s*(\d+)", line, re.IGNORECASE)
            if booth_m:
                booth_no = booth_m.group(1).strip()

    return {
        "corporation_name": corporation,
        "ward_no": ward_no,
        "ward_name_en": ward_name_en,
        "booth_no": booth_no,
        "booth_name_en": booth_name_en,
    }


def _lines_from_words(page) -> List[str]:
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
    if not words:
        return []
    # Group words into lines by rounded y0, then sort by x0
    lines: Dict[int, List[Dict[str, Any]]] = {}
    for w in words:
        y_key = int(round(w.get("top", 0)))
        lines.setdefault(y_key, []).append(w)
    rendered: List[str] = []
    for y in sorted(lines.keys()):
        row = sorted(lines[y], key=lambda item: item.get("x0", 0))
        text = " ".join([item.get("text", "").strip() for item in row if item.get("text")])
        if text:
            rendered.append(text)
    return rendered


def _flush_record(records: List[Dict[str, Any]], record: Dict[str, Any]) -> None:
    if record.get("epic") or record.get("name_en"):
        records.append(record.copy())


def _clean_value(text: str) -> str:
    cleaned = PHOTO_CLEAN_RE.sub("", text)
    cleaned = re.sub(r"\bName\s*:\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b(Father's|Husband's|Mother's|Other's)\s+Name\s*:\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = cleaned.replace("||", " ").replace("|", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip(" :-")


def _apply_name_blocks(line: str, current: Dict[str, Any], allow_overwrite: bool = False) -> None:
    name_matches = [m.group(1).strip() for m in CARD_NAME_BLOCK_RE.finditer(line)]
    if not name_matches:
        return
    first = _clean_value(name_matches[0])
    if first and (allow_overwrite or not current.get("name_en")):
        current["name_en"] = first
    if len(name_matches) > 1:
        second = _clean_value(name_matches[1])
        if second and (allow_overwrite or not current.get("name_local")):
            current["name_local"] = second


def _apply_relation_blocks(line: str, current: Dict[str, Any]) -> None:
    rel_matches = list(CARD_REL_BLOCK_RE.finditer(line))
    if not rel_matches:
        return
    rel_label = rel_matches[0].group(1).strip().lower()
    rel_value = _clean_value(rel_matches[0].group(2).strip())
    if "father" in rel_label:
        current["relation_type"] = "Father"
    elif "husband" in rel_label:
        current["relation_type"] = "Husband"
    elif "mother" in rel_label:
        current["relation_type"] = "Mother"
    else:
        current["relation_type"] = "Other"
    if rel_value:
        current["relation_name_en"] = current.get("relation_name_en") or rel_value
    if len(rel_matches) > 1:
        rel_value_local = _clean_value(rel_matches[1].group(2).strip())
        if rel_value_local:
            current["relation_name_local"] = current.get("relation_name_local") or rel_value_local

def _split_line_by_epic(line: str) -> List[Tuple[str, Optional[str], Optional[str]]]:
    """Split an OCR/text line into one chunk per voter card.

    Electoral rolls often put 3 cards on one row: [card1 fields][EPIC1][card2 fields][EPIC2]...
    Each card's *name, relation, age, ...* appear **before** its EPIC. Older logic used
    segments starting **at** each EPIC, which skipped card1's text (only used for serial)
    and merged multiple names into the wrong row.
    """
    matches = list(CARD_EPIC_RE.finditer(line))
    if not matches:
        return [(line, None, None)]
    segments: List[Tuple[str, Optional[str], Optional[str]]] = []
    for idx, match in enumerate(matches):
        chunk_start = 0 if idx == 0 else matches[idx - 1].end()
        chunk_end = match.start()
        card_text = line[chunk_start:chunk_end]
        epic = _normalize_epic(match.group(0))
        serial_match = CARD_SERIAL_RE.search(card_text)
        serial = serial_match.group(1) if serial_match else None
        segments.append((card_text.strip(), epic, serial))
    tail = line[matches[-1].end() :].strip()
    if tail:
        segments.append((tail, None, None))
    return segments


def _parse_cards_from_lines(lines: List[str], booth_no: Optional[str]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    current: Dict[str, Any] = {}
    pending_serial: Optional[str] = None

    for raw in lines:
        line = raw.strip()
        if not line:
            continue

        if line.isdigit() and len(line) <= 4:
            pending_serial = line

        segments = _split_line_by_epic(line)
        for segment, epic, serial in segments:
            segment = segment.strip()
            if epic:
                if current.get("epic") or current.get("name_en"):
                    _flush_record(records, current)
                    current = {}
                current["booth"] = booth_no
                current["serial"] = serial or pending_serial or current.get("serial")
                current["epic"] = epic
                pending_serial = None

            if not epic and current.get("name_en") and len(CARD_NAME_BLOCK_RE.findall(segment)) > 1:
                pass
            else:
                _apply_name_blocks(segment, current)
            _apply_relation_blocks(segment, current)

            house_match = CARD_HOUSE_RE.search(segment)
            if house_match:
                current["house"] = house_match.group(1).strip()

            age_gender_match = CARD_AGE_GENDER_RE.search(segment)
            if age_gender_match:
                current["age"] = age_gender_match.group(1).strip()
                gender = age_gender_match.group(2).strip()
                current["sex"] = (
                    "Male" if gender.lower().startswith("m") else "Female" if gender.lower().startswith("f") else gender
                )
            else:
                age_match = CARD_AGE_RE.search(segment)
                if age_match:
                    current["age"] = age_match.group(1).strip()
                gender_match = CARD_GENDER_RE.search(segment)
                if gender_match:
                    gender = gender_match.group(1).strip()
                    current["sex"] = (
                        "Male"
                        if gender.lower().startswith("m")
                        else "Female"
                        if gender.lower().startswith("f")
                        else gender
                    )

    _flush_record(records, current)
    return records


def _cluster_positions(values: List[float], tolerance: float = 0.06) -> List[float]:
    if not values:
        return []
    values_sorted = sorted(values)
    clusters: List[List[float]] = [[values_sorted[0]]]
    for v in values_sorted[1:]:
        if abs(v - clusters[-1][-1]) <= tolerance:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    return [sum(c) / len(c) for c in clusters]


def _cluster_with_fallback(values: List[float], primary_tol: float, fallback_tol: float, min_clusters: int) -> List[float]:
    clusters = _cluster_positions(values, tolerance=primary_tol)
    if len(clusters) < min_clusters:
        clusters = _cluster_positions(values, tolerance=fallback_tol)
    return clusters


def _edges_from_centers(centers: List[float]) -> List[Tuple[float, float]]:
    if not centers:
        return []
    centers = sorted(centers)
    edges: List[Tuple[float, float]] = []
    for idx, center in enumerate(centers):
        if idx == 0:
            left_edge = 0.0
        else:
            left_edge = (centers[idx - 1] + center) / 2
        if idx == len(centers) - 1:
            right_edge = 1.0
        else:
            right_edge = (center + centers[idx + 1]) / 2
        edges.append((left_edge, right_edge))
    return edges


def _group_words_into_lines(words: List[Dict[str, Any]], tol: float = 0.01) -> List[str]:
    if not words:
        return []
    words_sorted = sorted(words, key=lambda w: (w["top"], w["left"]))
    lines: List[List[Dict[str, Any]]] = []
    for word in words_sorted:
        if not lines:
            lines.append([word])
            continue
        if abs(word["top"] - lines[-1][0]["top"]) <= tol:
            lines[-1].append(word)
        else:
            lines.append([word])
    rendered: List[str] = []
    for line_words in lines:
        line_words_sorted = sorted(line_words, key=lambda w: w["left"])
        text = " ".join([w["text"] for w in line_words_sorted if w.get("text")])
        if text:
            rendered.append(text)
    return rendered


def _parse_cards_from_textract(
    words: List[Dict[str, Any]],
    booth_no: Optional[str],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    epic_words = [
        word
        for word in words
        if CARD_EPIC_RE.search(word["text"]) or CARD_EPIC_RE.search(_normalize_epic(word["text"]))
    ]
    if not epic_words:
        return [], []
    lefts = [word["left"] for word in epic_words]
    tops = [word["top"] for word in epic_words]
    col_centers = sorted(_cluster_with_fallback(lefts, primary_tol=0.03, fallback_tol=0.06, min_clusters=2))
    row_centers = sorted(_cluster_with_fallback(tops, primary_tol=0.035, fallback_tol=0.07, min_clusters=2))
    col_edges = _edges_from_centers(col_centers)
    row_edges = _edges_from_centers(row_centers)
    records: List[Dict[str, Any]] = []
    debug_rows: List[Dict[str, Any]] = []

    for epic_word in epic_words:
        epic_match = CARD_EPIC_RE.search(epic_word["text"]) or CARD_EPIC_RE.search(
            _normalize_epic(epic_word["text"])
        )
        if not epic_match:
            continue
        epic = _normalize_epic(epic_match.group(0))
        left = epic_word["left"]
        top = epic_word["top"]

        col_idx = min(range(len(col_centers)), key=lambda i: abs(col_centers[i] - left))
        row_idx = min(range(len(row_centers)), key=lambda i: abs(row_centers[i] - top))
        card_left, card_right = col_edges[col_idx]
        card_top, card_bottom = row_edges[row_idx]

        pad = 0.008
        card_words = [
            word
            for word in words
            if card_left + pad <= word["left"] <= card_right - pad
            and (word["left"] + word.get("width", 0)) <= card_right - pad
            and card_top + pad <= word["top"] <= card_bottom - pad
            and (word["top"] + word.get("height", 0)) <= card_bottom - pad
        ]
        text_lines = _group_words_into_lines(card_words)
        record: Dict[str, Any] = {"booth": booth_no, "epic": epic}

        # Serial number tends to be a small number box near top-left
        for line in text_lines[:3]:
            serial_match = CARD_SERIAL_RE.search(line)
            if serial_match:
                record["serial"] = serial_match.group(1)
                break

        combined_raw = " ".join(text_lines)
        combined = _clean_value(combined_raw)
        name_match = CARD_NAME_BLOCK_RE.search(combined)
        if name_match:
            record["name_en"] = _clean_value(name_match.group(1))

        rel_match = CARD_REL_BLOCK_RE.search(combined)
        if rel_match:
            rel_label = rel_match.group(1).strip().lower()
            rel_value = _clean_value(rel_match.group(2).strip())
            if "father" in rel_label:
                record["relation_type"] = "Father"
            elif "husband" in rel_label:
                record["relation_type"] = "Husband"
            elif "mother" in rel_label:
                record["relation_type"] = "Mother"
            else:
                record["relation_type"] = "Other"
            record["relation_name_en"] = rel_value

        house_match = CARD_HOUSE_RE.search(combined)
        if house_match:
            record["house"] = house_match.group(1).strip()

        age_gender_match = CARD_AGE_GENDER_RE.search(combined)
        if age_gender_match:
            record["age"] = age_gender_match.group(1).strip()
            gender = age_gender_match.group(2).strip()
            record["sex"] = (
                "Male"
                if gender.lower().startswith("m")
                else "Female"
                if gender.lower().startswith("f")
                else gender
            )
        else:
            age_match = CARD_AGE_RE.search(combined)
            if age_match:
                record["age"] = age_match.group(1).strip()
            gender_match = CARD_GENDER_RE.search(combined)
            if gender_match:
                gender = gender_match.group(1).strip()
                record["sex"] = (
                    "Male"
                    if gender.lower().startswith("m")
                    else "Female"
                    if gender.lower().startswith("f")
                    else gender
                )
        _flush_record(records, record)
        debug_rows.append({"epic": record.get("epic"), "text": combined_raw})
    return records, debug_rows


def _build_excel(
    records: List[Dict[str, Any]],
    booth_rows: List[Dict[str, Any]],
    debug_rows: Optional[List[Dict[str, Any]]] = None,
) -> str:
    from openpyxl import Workbook

    wb = Workbook()
    data_sheet = wb.active
    data_sheet.title = "DATA"

    data_headers = [
        "BOOTH",
        "SL",
        "HOUSE",
        "NAME",
        "NAME",
        "SEX",
        "AGE",
        "EPIC",
        "RELATION_TYPE",
        "RELATION",
        "RELATION",
        "MOBILE",
    ]
    data_sheet.append(data_headers)
    for row in records:
        data_sheet.append(
            [
                row.get("booth") or "",
                row.get("serial") or "",
                row.get("house") or "",
                row.get("name_en") or "",
                row.get("name_local") or "",
                row.get("sex") or "",
                row.get("age") or "",
                row.get("epic") or "",
                row.get("relation_type") or "",
                row.get("relation_name_en") or "",
                row.get("relation_name_local") or "",
                row.get("mobile") or "",
            ]
        )

    booth_sheet = wb.create_sheet("BOOTH")
    booth_headers = [
        "corporation_name",
        "ward_name",
        "ward_name_l1",
        "WARD",
        "BOOTH",
        "BOOTH NAME",
        "BOOTH NAME KANNADA",
        "LAT ",
        "LONG",
    ]
    booth_sheet.append(booth_headers)
    for row in booth_rows:
        booth_sheet.append(
            [
                row.get("corporation_name") or "",
                row.get("ward_name_en") or "",
                row.get("ward_name_local") or "",
                row.get("ward_no") or "",
                row.get("booth_no") or "",
                row.get("booth_name_en") or "",
                row.get("booth_name_local") or "",
                row.get("lat") or "",
                row.get("long") or "",
            ]
        )

    if not records:
        info_sheet = wb.create_sheet("INFO")
        info_sheet.append(["INFO"])
        info_sheet.append(["No card data found in PDF."])
        info_sheet.append(["If this is a scanned PDF, Textract or OCR is required."])
        info_sheet.append(["Ensure AWS credentials + S3 bucket are configured, or install tesseract + poppler."])

    if debug_rows:
        debug_sheet = wb.create_sheet("DEBUG")
        debug_sheet.append(["EPIC", "CARD_TEXT"])
        for row in debug_rows:
            debug_sheet.append([row.get("epic") or "", row.get("text") or ""])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _ocr_lines_from_pdf(pdf_path: str) -> List[str]:
    tmp_dir = tempfile.mkdtemp(prefix="extract_ocr_")
    try:
        pdftoppm = shutil.which("pdftoppm") or "/opt/homebrew/bin/pdftoppm"
        tesseract = shutil.which("tesseract") or "/opt/homebrew/bin/tesseract"
        if not os.path.exists(pdftoppm) or not os.path.exists(tesseract):
            return []
        prefix = os.path.join(tmp_dir, "page")
        subprocess.run([pdftoppm, "-r", "300", "-png", pdf_path, prefix], check=False, capture_output=True)
        png_files = sorted([os.path.join(tmp_dir, f) for f in os.listdir(tmp_dir) if f.endswith(".png")])
        if not png_files:
            return []
        lines: List[str] = []
        for png in png_files:
            proc = subprocess.run(
                [tesseract, png, "stdout", "-l", "eng"],
                check=False,
                capture_output=True,
                text=True,
            )
            text = proc.stdout or ""
            for line in text.splitlines():
                stripped = line.strip()
                if stripped:
                    lines.append(stripped)
        return lines
    finally:
        try:
            for name in os.listdir(tmp_dir):
                os.remove(os.path.join(tmp_dir, name))
            os.rmdir(tmp_dir)
        except Exception:
            pass


def _textract_lines_from_pdf(pdf_path: str, bucket: str) -> List[Dict[str, List[Dict[str, Any]]]]:
    s3, textract = _aws_clients()
    key = f"extract/{uuid.uuid4().hex}.pdf"
    print(f"[extract] Uploading PDF to s3://{bucket}/{key}")
    with open(pdf_path, "rb") as f:
        s3.put_object(Bucket=bucket, Key=key, Body=f, ContentType="application/pdf")

    try:
        start = textract.start_document_text_detection(DocumentLocation={"S3Object": {"Bucket": bucket, "Name": key}})
        job_id = start["JobId"]
        print(f"[extract] Textract job started: {job_id}")
        pages: Dict[int, Dict[str, List[Dict[str, Any]]]] = {}
        next_token = None
        attempts = 0
        while True:
            params = {"JobId": job_id}
            if next_token:
                params["NextToken"] = next_token
            resp = textract.get_document_text_detection(**params)
            status = resp.get("JobStatus")
            if attempts % 3 == 0:
                print(f"[extract] Textract status={status} pages_collected={len(pages)}")
            if status == "FAILED":
                break
            if status in {"IN_PROGRESS", "SUCCEEDED"}:
                attempts += 1
            for block in resp.get("Blocks", []):
                block_type = block.get("BlockType")
                if block_type not in {"LINE", "WORD"}:
                    continue
                page = int(block.get("Page", 1))
                bbox = block.get("Geometry", {}).get("BoundingBox", {})
                top = float(bbox.get("Top", 0))
                left = float(bbox.get("Left", 0))
                width = float(bbox.get("Width", 0))
                height = float(bbox.get("Height", 0))
                text = block.get("Text", "").strip()
                if not text:
                    continue
                entry = {
                    "top": top,
                    "left": left,
                    "width": width,
                    "height": height,
                    "text": text,
                }
                pages.setdefault(page, {"lines": [], "words": []})
                if block_type == "LINE":
                    pages[page]["lines"].append(entry)
                else:
                    pages[page]["words"].append(entry)
            next_token = resp.get("NextToken")
            if not next_token:
                if status == "SUCCEEDED":
                    break
                if status == "IN_PROGRESS":
                    time.sleep(1.2)
                    continue
                break
        # Order lines by top then left within each page
        ordered_pages: List[Dict[str, List[Dict[str, Any]]]] = []
        for page in sorted(pages.keys()):
            lines = sorted(pages[page]["lines"], key=lambda item: (item["top"], item["left"]))
            words = sorted(pages[page]["words"], key=lambda item: (item["top"], item["left"]))
            ordered_pages.append({"lines": lines, "words": words})
        return ordered_pages
    finally:
        try:
            s3.delete_object(Bucket=bucket, Key=key)
        except Exception:
            pass


def _run_pdf_extract(pdf_path: str, debug: bool = False) -> Tuple[str, int]:
    records: List[Dict[str, Any]] = []
    booth_rows: Dict[Tuple[str, str], Dict[str, Any]] = {}
    debug_rows: List[Dict[str, Any]] = []

    bucket = os.getenv("EXTRACT_S3_BUCKET", "extract")
    textract_pages: List[Dict[str, List[Dict[str, Any]]]] = []
    skip_textract = os.getenv("EXTRACT_SKIP_TEXTRACT", "").strip().lower() in ("1", "true", "yes")
    if not skip_textract:
        try:
            textract_pages = _textract_lines_from_pdf(pdf_path, bucket)
        except Exception:
            textract_pages = []

    if textract_pages:
        for page_data in textract_pages:
            page_lines = page_data.get("lines", [])
            page_words = page_data.get("words", [])
            ordered_lines = sorted(page_lines, key=lambda x: (x.get("top", 0), x.get("left", 0)))
            header_lines: List[str] = []
            for line in ordered_lines[:20]:
                t = (line.get("text") or "").strip()
                if t:
                    header_lines.append(t)
            header = _parse_header_info(header_lines)
            booth_key = (header.get("ward_no") or "", header.get("booth_no") or "")
            booth_rows.setdefault(
                booth_key,
                {
                    **header,
                    "ward_name_local": "",
                    "booth_name_local": "",
                    "lat": "",
                    "long": "",
                },
            )
            booth_no = header.get("booth_no")
            line_texts = [ln["text"].strip() for ln in ordered_lines if (ln.get("text") or "").strip()]
            page_records = _parse_cards_from_lines(line_texts, booth_no)
            page_debug: List[Dict[str, Any]] = []
            if not page_records and page_words:
                line_from_words = _group_words_into_lines(page_words)
                page_records = _parse_cards_from_lines(line_from_words, booth_no)
            if not page_records and page_words:
                page_records, page_debug = _parse_cards_from_textract(page_words, booth_no)
            records.extend(page_records)
            if debug:
                if page_debug:
                    debug_rows.extend(page_debug)
                elif page_records:
                    debug_rows.extend([{"epic": r.get("epic"), "text": ""} for r in page_records])
    else:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                lines = [line.strip() for line in text.splitlines() if line.strip()]
                if not lines:
                    lines = _lines_from_words(page)
                header = _parse_header_info(lines[:12])
                booth_key = (header.get("ward_no") or "", header.get("booth_no") or "")
                booth_rows.setdefault(
                    booth_key,
                    {
                        **header,
                        "ward_name_local": "",
                        "booth_name_local": "",
                        "lat": "",
                        "long": "",
                    },
                )

                booth_no = header.get("booth_no")
                page_records = _parse_cards_from_lines(lines, booth_no)
                records.extend(page_records)

    if not records:
        ocr_lines = _ocr_lines_from_pdf(pdf_path)
        if ocr_lines:
            header = _parse_header_info(ocr_lines[:12])
            booth_rows.setdefault(
                (header.get("ward_no") or "", header.get("booth_no") or ""),
                {**header, "ward_name_local": "", "booth_name_local": "", "lat": "", "long": ""},
            )
            records = _parse_cards_from_lines(ocr_lines, header.get("booth_no"))

    excel_path = _build_excel(records, list(booth_rows.values()), debug_rows if debug else None)
    return excel_path, len(records)


@router.post("/api/extract/pdf-to-excel")
def extract_pdf_to_excel(file: UploadFile = File(...), debug: bool = Query(False)):
    ensure_extract_schema()
    engine = _get_engine()
    job_id = None
    try:
        if engine is not None:
            with engine.begin() as conn:
                result = conn.execute(
                    text(
                        """
                        INSERT INTO extract.pdf_extract_jobs (filename, status)
                        VALUES (:filename, 'processing')
                        RETURNING id
                        """
                    ),
                    {"filename": file.filename},
                ).first()
                job_id = result[0] if result else None
    except Exception:
        job_id = None

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(file.file.read())
        pdf_path = tmp.name

    excel_path, row_count = _run_pdf_extract(pdf_path, debug)

    try:
        if engine is not None:
            with engine.begin() as conn:
                conn.execute(
                    text(
                        """
                        UPDATE extract.pdf_extract_jobs
                        SET status = 'completed', rows_count = :rows_count, result_path = :result_path, completed_at = :completed_at
                        WHERE id = :job_id
                        """
                    ),
                    {
                        "rows_count": row_count,
                        "result_path": excel_path,
                        "completed_at": datetime.utcnow(),
                        "job_id": job_id,
                    },
                )
    except Exception:
        pass

    return FileResponse(excel_path, filename="extract.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
